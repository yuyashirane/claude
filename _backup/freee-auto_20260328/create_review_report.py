from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

HEADER_FONT = Font(name='Meiryo UI', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
SUBHEADER_FILL = PatternFill('solid', fgColor='D6E4F0')
SUBHEADER_FONT = Font(name='Meiryo UI', bold=True, size=10)
NORMAL_FONT = Font(name='Meiryo UI', size=10)
BOLD_FONT = Font(name='Meiryo UI', bold=True, size=10)
RED_FONT = Font(name='Meiryo UI', size=10, color='FF0000')
RED_BOLD = Font(name='Meiryo UI', bold=True, size=10, color='FF0000')
TITLE_FONT = Font(name='Meiryo UI', bold=True, size=14, color='2F5496')
WARN_FILL = PatternFill('solid', fgColor='FFF2CC')
ERR_FILL = PatternFill('solid', fgColor='FCE4EC')
OK_FILL = PatternFill('solid', fgColor='E8F5E9')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
NUM_FMT = '#,##0'
PCT_FMT = '0.0%'

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data_area(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# Sheet 1: 概況
# ============================================================
ws1 = wb.active
ws1.title = '概況'
set_col_widths(ws1, [20, 18, 18, 18, 14])

ws1['A1'] = '㈱デイリーユニフォーム 期中レビュー結果'
ws1['A1'].font = TITLE_FONT
ws1.merge_cells('A1:E1')
ws1['A3'] = '基本情報'
ws1['A3'].font = SUBHEADER_FONT
ws1['A3'].fill = SUBHEADER_FILL
ws1.merge_cells('A3:E3')

info = [
    ['決算月', '5月（期首6/1〜期末5/31）'],
    ['対象期間', '2025年6月〜12月（7ヶ月）'],
    ['会計年度', '2026年5月期'],
    ['業種', '衣料卸売・繊維'],
    ['税務方式', '税抜経理'],
    ['freee事業所ID', '10794380'],
]
for i, (k, v) in enumerate(info, 5):
    ws1.cell(row=i, column=1, value=k).font = BOLD_FONT
    ws1.cell(row=i, column=2, value=v).font = NORMAL_FONT

r = 13
ws1.cell(row=r, column=1, value='概況サマリー（前年同期比較）').font = SUBHEADER_FONT
ws1.cell(row=r, column=1).fill = SUBHEADER_FILL
ws1.merge_cells(f'A{r}:E{r}')
r += 1
for c, h in enumerate(['指標', '当期(7ヶ月)', '前年同期(7ヶ月)', '増減額', '増減率'], 1):
    ws1.cell(row=r, column=c, value=h)
style_header_row(ws1, r, 5)
r += 1
summary = [
    ['売上高', 27137652, 40478000],
    ['売上総利益', 14550441, 21229875],
    ['営業利益', -2810516, 3035250],
    ['経常利益', -2699601, 5398328],
    ['当期純利益', -2785047, 4591956],
    ['総資産', 21502435, 35632328],
    ['純資産', 7212264, 10591560],
]
for i, (name, cur, prev) in enumerate(summary):
    row = r + i
    ws1.cell(row=row, column=1, value=name).font = BOLD_FONT
    ws1.cell(row=row, column=2, value=cur).number_format = NUM_FMT
    ws1.cell(row=row, column=3, value=prev).number_format = NUM_FMT
    ws1.cell(row=row, column=4).value = f'=B{row}-C{row}'
    ws1.cell(row=row, column=4).number_format = NUM_FMT
    ws1.cell(row=row, column=5).value = f'=IF(C{row}=0,"-",B{row}/C{row}-1)'
    ws1.cell(row=row, column=5).number_format = PCT_FMT
    if cur < 0:
        ws1.cell(row=row, column=2).font = RED_BOLD
style_data_area(ws1, r, r + len(summary) - 1, 5)

# ============================================================
# Sheet 2: PL比較
# ============================================================
ws2 = wb.create_sheet('PL前年同期比較')
set_col_widths(ws2, [22, 16, 16, 16, 12, 12, 12])
headers = ['項目', '当期(A)', '前年同期(B)', '増減(A-B)', '増減率', '売上比(当期)', '売上比(前年)']
for c, h in enumerate(headers, 1):
    ws2.cell(row=1, column=c, value=h)
style_header_row(ws2, 1, 7)

pl_data = [
    ['売上高', 27137652, 40478000, True],
    ['売上原価', 12587211, 19248125, False],
    ['　期首商品棚卸高', 338580, 2542830, False],
    ['　仕入高', 12248631, 16705295, False],
    ['売上総利益', 14550441, 21229875, True],
    ['販売管理費計', 17360957, 18194625, True],
    ['　役員報酬', 8600000, 11200000, False],
    ['　旅費交通費', 1543125, 348915, False],
    ['　法定福利費', 1395912, 1505784, False],
    ['　消耗品費', 1326203, 784237, False],
    ['　会議費', 1032318, 722434, False],
    ['　外注費', 790571, 577299, False],
    ['　地代家賃', 463640, 649096, False],
    ['　交際費', 416565, 109625, False],
    ['　支払報酬料', 368000, 435000, False],
    ['　支払手数料', 328198, 371437, False],
    ['　減価償却費', 328706, 478054, False],
    ['　その他', 767719, 1013344, False],
    ['営業利益', -2810516, 3035250, True],
    ['営業外収益', 177124, 2401792, False],
    ['　雑収入', 154201, 2400000, False],
    ['　受取利息', 22923, 1792, False],
    ['営業外費用', 66209, 38714, False],
    ['経常利益', -2699601, 5398328, True],
    ['税引前当期純利益', -2699601, 5398328, True],
    ['法人税等', 85446, 806372, False],
    ['当期純利益', -2785047, 4591956, True],
]
for i, (name, cur, prev, is_bold) in enumerate(pl_data):
    row = i + 2
    ws2.cell(row=row, column=1, value=name).font = BOLD_FONT if is_bold else NORMAL_FONT
    ws2.cell(row=row, column=2, value=cur).number_format = NUM_FMT
    ws2.cell(row=row, column=3, value=prev).number_format = NUM_FMT
    ws2.cell(row=row, column=4).value = f'=B{row}-C{row}'
    ws2.cell(row=row, column=4).number_format = NUM_FMT
    ws2.cell(row=row, column=5).value = f'=IF(C{row}=0,"-",B{row}/C{row}-1)'
    ws2.cell(row=row, column=5).number_format = PCT_FMT
    ws2.cell(row=row, column=6).value = f'=IF(B$2=0,"-",B{row}/B$2)'
    ws2.cell(row=row, column=6).number_format = PCT_FMT
    ws2.cell(row=row, column=7).value = f'=IF(C$2=0,"-",C{row}/C$2)'
    ws2.cell(row=row, column=7).number_format = PCT_FMT
    if cur < 0:
        ws2.cell(row=row, column=2).font = RED_FONT
    for c in range(1, 8):
        ws2.cell(row=row, column=c).border = THIN_BORDER

# ============================================================
# Sheet 3: BS比較
# ============================================================
ws3 = wb.create_sheet('BS比較')
set_col_widths(ws3, [24, 16, 16, 16, 16, 16])
bs_headers = ['項目', '当期12月末(A)', '前期末5月末(B)', '増減(A-B)', '前年12月末(C)', '増減(A-C)']
for c, h in enumerate(bs_headers, 1):
    ws3.cell(row=1, column=c, value=h)
style_header_row(ws3, 1, 6)

bs_data = [
    ['【流動資産】', None, None, None, True],
    ['現金', -4290, 0, 0, False],
    ['三菱UFJ', 9071379, 13237113, 15506246, False],
    ['GMOあおぞら', 8906103, 11728603, 1700756, False],
    ['売掛金', 1585525, 2773588, 12033387, False],
    ['商品', 0, 338580, 0, False],
    ['役員貸付金', -5003540, 0, 0, False],
    ['未収還付法人税等', 155000, 510500, 0, False],
    ['仮払消費税', 1732601, 0, 2569543, False],
    ['その他流動', 59424, 1298, 1298, False],
    ['流動資産計', 16502202, 28589682, 31811230, True],
    ['【固定資産】', None, None, None, True],
    ['有形固定資産', 993427, 1027496, 1406129, False],
    ['保険積立金', 4000000, 2600000, 2400000, False],
    ['長期前払費用', 6806, 11566, 14969, False],
    ['固定資産計', 5000233, 3639062, 3821098, True],
    ['資産合計', 21502435, 32228744, 35632328, True],
    ['【負債】', None, None, None, True],
    ['買掛金', 0, 3735666, 4108096, False],
    ['立替経費（佐藤）', 855279, 2252699, 930485, False],
    ['未払金', -3105511, 579153, 580817, False],
    ['JAL CARD', 214010, 429415, 455422, False],
    ['未払法人税等', 0, 35000, 0, False],
    ['未払消費税等', 0, 712700, 0, False],
    ['預り金', 706934, 611800, 44500, False],
    ['仮受消費税', 2669459, 0, 4000448, False],
    ['流動負債計', 1340171, 8356433, 10240768, True],
    ['長期借入金', 12950000, 13875000, 14800000, False],
    ['負債合計', 14290171, 22231433, 25040768, True],
    ['【純資産】', None, None, None, True],
    ['資本金', 1000000, 1000000, 1000000, False],
    ['繰越利益剰余金', 8997311, 8997311, 4999604, False],
    ['当期純損益', -2785047, None, 4591956, False],
    ['純資産合計', 7212264, 9997311, 10591560, True],
]
for i, (name, a, b, c_val, is_header) in enumerate(bs_data):
    row = i + 2
    ws3.cell(row=row, column=1, value=name).font = BOLD_FONT if is_header else NORMAL_FONT
    if a is not None:
        ws3.cell(row=row, column=2, value=a).number_format = NUM_FMT
        if a < 0:
            ws3.cell(row=row, column=2).font = RED_FONT
            ws3.cell(row=row, column=2).fill = ERR_FILL
    if b is not None:
        ws3.cell(row=row, column=3, value=b).number_format = NUM_FMT
    if a is not None and b is not None:
        ws3.cell(row=row, column=4).value = f'=B{row}-C{row}'
        ws3.cell(row=row, column=4).number_format = NUM_FMT
    if c_val is not None:
        ws3.cell(row=row, column=5, value=c_val).number_format = NUM_FMT
    if a is not None and c_val is not None:
        ws3.cell(row=row, column=6).value = f'=B{row}-E{row}'
        ws3.cell(row=row, column=6).number_format = NUM_FMT
    for col in range(1, 7):
        ws3.cell(row=row, column=col).border = THIN_BORDER

# ============================================================
# Sheet 4: 15分野記帳チェック
# ============================================================
ws4 = wb.create_sheet('記帳チェック')
set_col_widths(ws4, [6, 28, 12, 55])
check_headers = ['#', '分野', '判定', '詳細']
for c, h in enumerate(check_headers, 1):
    ws4.cell(row=1, column=c, value=h)
style_header_row(ws4, 1, 4)

checks = [
    [1, '現金・預金【GA-1】', 'NG', '現金 -4,290円（マイナス）。預金は正常。'],
    [2, '借入金【HB1-1】', '要確認', '長期借入金12,950千円。7ヶ月で925千円返済。返済予定表との照合が必要。支払利息65千円（利率確認要）。'],
    [3, '固定資産【GD-1】', '要確認', '消耗品費に106,366円（三越）あり。税抜96,696円で10万未満だが内容確認要。一括償却資産294,637円新規取得→台帳登録確認。'],
    [4, '家賃支払【HD-1】', 'OK', '地代家賃463,640円/7ヶ月=月66,234円。ただし前年月93千円から減額→契約変更確認。'],
    [5, '人件費【JC2-1】', '要確認', '役員報酬: 6-7月160万、8月以降108万に減額。期首3ヶ月以内の改定で適法だが議事録確認要。法定福利費比率16.2%（目安14-15%よりやや高）。'],
    [6, '士業・外注【HB2-1】', 'NG', '支払報酬料（税理士）に源泉徴収の記載なし→確認必須。外注費に個人外注先あり（佐藤壽久・久保布紀等）→源泉徴収の要否と役員との関係確認。'],
    [7, 'TPS9100/給与', 'OK', 'freee人事労務連携のため対象外。'],
    [8, '役員関係【JB-1】', 'NG', '役員貸付金 -5,003,540円（マイナス=実質役員借入金）。前期は「役員借入金」科目あり→科目修正要。立替経費（佐藤）855千円は精算進行中。'],
    [9, '売上・売掛金【HA-1】', '要確認', '売上高27,138千円（前年比-33%）。売掛金1,586千円は売上規模に対し妥当だが、売上減の原因確認要。'],
    [10, '仕入・買掛金【JC3-1】', 'OK', '原価率46.4%（卸売業平均と乖離→ブランド付加価値型事業と推定）。買掛金0円（全件決済済み確認）。'],
    [11, '在庫【JC3-3】', 'NG', '期首商品338,580円→原価振替後ゼロ。期末棚卸未計上。衣料卸売で在庫ゼロは通常ありえない。'],
    [12, 'その他経費【JC3-4】', '要確認', '会議費103万（月15万）・交際費42万→区分確認。旅費交通費154万（前年比+342%）→内容確認。未払金 -3,106千円→異常（給与unsettledが原因の可能性）。'],
    [13, '営業外損益【HC-1】', '要確認', '雑収入154千円の内訳確認要。前年の雑収入240万の内容も確認。'],
    [14, '税金【JC3-5】', 'OK', '未払法人税等・未払消費税等とも期首から納付済みでゼロ。受取利息の源泉所得税（15.315%）計上済み。予定納税85,446円計上済み。'],
    [15, 'その他の気付事項', '要確認', '給与取引7件がすべてunsettled（未決済）→未払金マイナスの原因の可能性大。決済登録のマッチング確認要。'],
]
for i, (num, area, result, detail) in enumerate(checks):
    row = i + 2
    ws4.cell(row=row, column=1, value=num).font = NORMAL_FONT
    ws4.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws4.cell(row=row, column=2, value=area).font = BOLD_FONT
    ws4.cell(row=row, column=3, value=result).font = BOLD_FONT
    ws4.cell(row=row, column=3).alignment = Alignment(horizontal='center')
    ws4.cell(row=row, column=4, value=detail).font = NORMAL_FONT
    ws4.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
    if result == 'NG':
        ws4.cell(row=row, column=3).fill = ERR_FILL
        ws4.cell(row=row, column=3).font = RED_BOLD
    elif result == '要確認':
        ws4.cell(row=row, column=3).fill = WARN_FILL
    else:
        ws4.cell(row=row, column=3).fill = OK_FILL
    for c in range(1, 5):
        ws4.cell(row=row, column=c).border = THIN_BORDER

# ============================================================
# Sheet 5: 消費税区分チェック
# ============================================================
ws5 = wb.create_sheet('消費税チェック')
set_col_widths(ws5, [22, 14, 28, 12, 45])
tax_headers = ['科目', 'tax_code', '区分', '判定', '備考']
for c, h in enumerate(tax_headers, 1):
    ws5.cell(row=1, column=c, value=h)
style_header_row(ws5, 1, 5)

tax_data = [
    ['役員報酬', '2', '不課税', 'OK', ''],
    ['法定福利費', '-', '不課税', 'OK', '推定'],
    ['外注費（インボイス登録済）', '136', '課税仕入10%（適格）', 'OK', '田原耕平'],
    ['外注費（インボイス非登録）', '189', '課税仕入10%（経過措置80%）', 'OK', '佐藤壽久、久保布紀、福岡洋平、高木洋平。2026/10以降は50%控除に変更要。'],
    ['支払報酬料（税理士）', '136', '課税仕入10%（適格）', 'OK', 'あしたの会計事務所税理士法人'],
    ['消耗品費（一般）', '136', '課税仕入10%（適格）', 'OK', ''],
    ['消耗品費（軽減税率）', '108', '課税仕入8%', 'OK', '407円のみ。飲食料品と推定。'],
    ['消耗品費（tax_code163）', '163', '要確認', '要確認', '8件2,703円。特定課税仕入等の定義確認要。金額僅少。'],
    ['受取利息', '23', '非課税売上', 'OK', ''],
    ['雑収入（保険返戻等）', '129', '課税売上10%', '要確認', '98,000円+36,200円。保険配当・返戻金なら不課税(2)が正しい。内容確認要。'],
    ['雑収入（前納減額金）', '23', '非課税', '要確認', '倒産防止共済9,900円。不課税(2)が適切では。少額。'],
    ['雑収入（予中間分）', '2', '不課税', 'OK', '22,300円。税の還付相当。'],
    ['支払利息', '-', '非課税仕入', 'OK', '推定'],
    ['保険料', '-', '非課税仕入', 'OK', '推定'],
    ['減価償却費', '-', '対象外', 'OK', ''],
    ['租税公課', '-', '不課税', 'OK', ''],
]
for i, (name, tc, cat, result, note) in enumerate(tax_data):
    row = i + 2
    ws5.cell(row=row, column=1, value=name).font = NORMAL_FONT
    ws5.cell(row=row, column=2, value=tc).font = NORMAL_FONT
    ws5.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws5.cell(row=row, column=3, value=cat).font = NORMAL_FONT
    ws5.cell(row=row, column=4, value=result).font = BOLD_FONT
    ws5.cell(row=row, column=4).alignment = Alignment(horizontal='center')
    ws5.cell(row=row, column=5, value=note).font = NORMAL_FONT
    ws5.cell(row=row, column=5).alignment = Alignment(wrap_text=True)
    if result == '要確認':
        ws5.cell(row=row, column=4).fill = WARN_FILL
    else:
        ws5.cell(row=row, column=4).fill = OK_FILL
    for c in range(1, 6):
        ws5.cell(row=row, column=c).border = THIN_BORDER

r2 = len(tax_data) + 4
ws5.cell(row=r2, column=1, value='消費税試算').font = SUBHEADER_FONT
ws5.cell(row=r2, column=1).fill = SUBHEADER_FILL
ws5.merge_cells(f'A{r2}:C{r2}')
ws5.cell(row=r2+1, column=1, value='仮受消費税').font = NORMAL_FONT
ws5.cell(row=r2+1, column=2, value=2669459).number_format = NUM_FMT
ws5.cell(row=r2+2, column=1, value='仮払消費税').font = NORMAL_FONT
ws5.cell(row=r2+2, column=2, value=1732601).number_format = NUM_FMT
ws5.cell(row=r2+3, column=1, value='差額（納付見込）').font = BOLD_FONT
ws5.cell(row=r2+3, column=2).value = f'=B{r2+1}-B{r2+2}'
ws5.cell(row=r2+3, column=2).number_format = NUM_FMT
ws5.cell(row=r2+3, column=2).font = BOLD_FONT
for rr in range(r2+1, r2+4):
    for cc in range(1, 3):
        ws5.cell(row=rr, column=cc).border = THIN_BORDER

# ============================================================
# Sheet 6: 確認事項一覧
# ============================================================
ws6 = wb.create_sheet('確認事項一覧')
set_col_widths(ws6, [5, 10, 30, 60])
issue_headers = ['#', '重要度', '項目', '内容・質問']
for c, h in enumerate(issue_headers, 1):
    ws6.cell(row=1, column=c, value=h)
style_header_row(ws6, 1, 4)

issues = [
    [1, '高', '未払金マイナス -311万', '給与がすべて未決済(unsettled)。振込の決済登録と未払金取引のマッチングが未了。仕訳の修正が必要。'],
    [2, '高', '役員貸付金マイナス -500万', '「役員借入金」科目に修正。立替経費（佐藤）85万円との相殺も決算時に検討。'],
    [3, '高', '支払報酬料の源泉徴収', '税理士報酬（月24千円税抜）に源泉10.21%が控除されているか確認。未控除なら是正必要。'],
    [4, '高', '外注費の源泉徴収・実態確認', '個人外注先への支払いに源泉徴収が必要か。「佐藤壽久」が役員と同一人物でないか確認。'],
    [5, '高', '売上33%減の原因', '月平均578万→388万と大幅減収。取引先別の分析と今後の見通し。'],
    [6, '中', '役員報酬改定の議事録', '160万→108万の減額。定時株主総会決議で行っているか確認。'],
    [7, '中', '期末棚卸高の計上', '12月末の在庫はゼロで正しいか。月次棚卸の実施を推奨。'],
    [8, '中', '雑収入の消費税区分', '98,000円+36,200円が課税売上10%。保険関連なら不課税に修正要。'],
    [9, '中', '旅費交通費+342%', '月22万円の出張費用の内容確認。'],
    [10, '中', '現金マイナス -4,290', '少額だが修正必要。'],
    [11, '低', '地代家賃の減額理由', '月93千→66千に減。契約変更の確認。'],
    [12, '低', '前年雑収入240万の内容', '当期に同様の収入がないため、一時的収入であれば当期の赤字が実力値。'],
]
for i, (num, level, item, content) in enumerate(issues):
    row = i + 2
    ws6.cell(row=row, column=1, value=num).font = NORMAL_FONT
    ws6.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    ws6.cell(row=row, column=2, value=level).font = BOLD_FONT
    ws6.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    ws6.cell(row=row, column=3, value=item).font = BOLD_FONT
    ws6.cell(row=row, column=4, value=content).font = NORMAL_FONT
    ws6.cell(row=row, column=4).alignment = Alignment(wrap_text=True)
    if level == '高':
        ws6.cell(row=row, column=2).fill = ERR_FILL
        ws6.cell(row=row, column=2).font = RED_BOLD
    elif level == '中':
        ws6.cell(row=row, column=2).fill = WARN_FILL
    else:
        ws6.cell(row=row, column=2).fill = OK_FILL
    for c in range(1, 5):
        ws6.cell(row=row, column=c).border = THIN_BORDER

# ============================================================
# Sheet 7: 役員報酬月次
# ============================================================
ws7 = wb.create_sheet('役員報酬月次')
set_col_widths(ws7, [10, 16, 16, 16, 16, 16, 16])
salary_headers = ['月', '役員報酬', '健保（預り）', '介護（預り）', '厚年（預り）', '源泉所得税', '住民税']
for c, h in enumerate(salary_headers, 1):
    ws7.cell(row=1, column=c, value=h)
style_header_row(ws7, 1, 7)

salary_data = [
    ['6月', 1600000, 78290, 12560, 118950, 123020, 44500],
    ['7月', 1600000, 78290, 12560, 118950, 123020, 44500],
    ['8月', 1080000, 78290, 12560, 118950, 35660, 0],
    ['9月', 1080000, 78290, 12560, 118950, 35660, 0],
    ['10月', 1080000, 78290, 12560, 118950, 35660, 0],
    ['11月', 1080000, 52524, 8426, 96990, 41940, 0],
    ['12月', 1080000, 52524, 8426, 96990, 41940, 82000],
]
for i, row_data in enumerate(salary_data):
    row = i + 2
    ws7.cell(row=row, column=1, value=row_data[0]).font = NORMAL_FONT
    ws7.cell(row=row, column=1).alignment = Alignment(horizontal='center')
    for c in range(1, 7):
        ws7.cell(row=row, column=c + 1, value=row_data[c]).number_format = NUM_FMT
        ws7.cell(row=row, column=c + 1).font = NORMAL_FONT
    for c in range(1, 8):
        ws7.cell(row=row, column=c).border = THIN_BORDER
    if row_data[1] != salary_data[0][1]:
        ws7.cell(row=row, column=2).fill = WARN_FILL

row_total = len(salary_data) + 2
ws7.cell(row=row_total, column=1, value='合計').font = BOLD_FONT
ws7.cell(row=row_total, column=1).alignment = Alignment(horizontal='center')
for c in range(2, 8):
    col_letter = get_column_letter(c)
    ws7.cell(row=row_total, column=c).value = f'=SUM({col_letter}2:{col_letter}{row_total-1})'
    ws7.cell(row=row_total, column=c).number_format = NUM_FMT
    ws7.cell(row=row_total, column=c).font = BOLD_FONT
    ws7.cell(row=row_total, column=c).border = THIN_BORDER
ws7.cell(row=row_total, column=1).border = THIN_BORDER

out = 'C:/Users/yuya_/claude/office/freee-auto/デイリーユニフォーム_期中レビュー_202506-12.xlsx'
wb.save(out)
print(f'Saved: {out}')
