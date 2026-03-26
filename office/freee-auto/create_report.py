from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Style definitions
title_font = Font(name="Meiryo UI", size=14, bold=True)
header_font = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2F5496")
red_fill = PatternFill("solid", fgColor="FFC7CE")
red_font = Font(name="Meiryo UI", size=10, color="9C0006")
yellow_fill = PatternFill("solid", fgColor="FFEB9C")
yellow_font = Font(name="Meiryo UI", size=10, color="9C6500")
green_fill = PatternFill("solid", fgColor="C6EFCE")
green_font = Font(name="Meiryo UI", size=10, color="006100")
normal_font = Font(name="Meiryo UI", size=10)
bold_font = Font(name="Meiryo UI", size=10, bold=True)
sub_font = Font(name="Meiryo UI", size=11, color="666666")
warn_font = Font(name="Meiryo UI", size=10, color="FF0000", bold=True)
section_font = Font(name="Meiryo UI", size=11, bold=True, color="2F5496")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
num_fmt = "#,##0"

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

# ===== Sheet 1: Summary =====
ws1 = wb.active
ws1.title = "調査サマリー"

ws1["A1"] = "Connectiv株式会社 源泉所得税チェックレポート"
ws1["A1"].font = title_font
ws1["A2"] = "対象期間: 2022年1月 - 2026年3月 / 作成日: 2026年3月25日"
ws1["A2"].font = sub_font

ws1["A4"] = "判定凡例"
ws1["A4"].font = bold_font
for r, (label, desc, font, fill) in enumerate([
    ("重大", "源泉徴収漏れの可能性が高い", red_font, red_fill),
    ("要確認", "業務内容により源泉対象の可能性あり", yellow_font, yellow_fill),
    ("問題なし", "源泉処理済み or 法人への支払い", green_font, green_fill),
], 5):
    c1 = ws1.cell(row=r, column=1, value=label)
    c1.font = font
    c1.fill = fill
    ws1.cell(row=r, column=2, value=desc).font = normal_font

headers = ["判定", "勘定科目", "取引先名", "個人/法人", "支払合計(税込)",
           "件数", "源泉徴収額", "源泉有無", "摘要/業務内容", "指摘事項/コメント"]
write_header_row(ws1, 9, headers)

data = [
    ["重大", "支払報酬料", "あしたの会計事務所", "個人(税理士)", 1817750, 46, 0, "なし",
     "税理士顧問料(月額22,330円等)", "税理士報酬は所得税法204条により源泉徴収必須。全期間にわたり源泉未処理。"],
    ["重大", "外注費", "須藤 一輝", "個人", 609000, 2, 0, "なし",
     "Snapshotデザイン&動画制作費/LP制作費", "デザイン料は所得税法204条対象。源泉徴収が必要。"],
    ["重大", "外注費", "ベトナムフリーランス", "個人/非居住者?", 1074750, 11, 0, "なし",
     "(摘要なし)", "非居住者への支払いの場合20.42%の源泉必要(所得税法212条)。居住地/役務提供地の確認要。"],
    ["要確認", "外注費", "星野萌音", "個人", 3616250, 7, 0, "なし",
     "開発費", "個人への高額支払い。システム開発なら204条対象外の可能性。契約書要確認。"],
    ["要確認", "外注費", "齊藤晶子", "個人(インボイス登録済)", 3121800, 14, 0, "なし",
     "システム開発", "個人への高額支払い。システム開発なら204条対象外の可能性。契約書要確認。"],
    ["要確認", "外注費", "タケダトキオ(デブオプワーカー)", "個人", 2664750, 14, 0, "なし",
     "SE稼働/インフラ開発費用", "SE業務なら204条対象外の可能性。契約書要確認。"],
    ["要確認", "外注費", "伊藤 源太(インボイス未登録)", "個人", 827250, 10, 0, "なし",
     "システム開発作業", "システム開発なら204条対象外の可能性。インボイス未登録のため経過措置確認も必要。"],
    ["要確認", "外注費", "Biteearth 徐基源", "個人", 510000, 4, 0, "なし",
     "開発協力", "個人への支払い。業務内容/非居住者該当性の確認要。"],
    ["要確認", "外注費", "満極 尚輝(インボイス未登録)", "個人", 172500, 3, 0, "なし",
     "(摘要なし)", "業務内容不明。摘要なしのため内容確認要。"],
    ["要確認", "支払報酬料", "東新宿総合法律事務所", "法律事務所", 7421, 1, 0, "なし",
     "(摘要なし)", "弁護士報酬なら少額でも源泉必要。実費精算の可能性もあり内容確認要。"],
    ["要確認", "支払報酬料", "社労士法人グランディール", "法人", 243150, 7, 0, "なし",
     "算定基礎届手続等", "社会保険労務士「法人」への支払い->源泉不要。ただし法人格の確認要。"],
    ["問題なし", "支払報酬料", "山下 聖志(顧問弁護士/士業)", "個人", 2665641, 47, 247419, "あり",
     "顧問料/株主間契約書レビュー等", "源泉徴収済み。税額の検算推奨。"],
    ["問題なし", "支払報酬料", "加藤来特許事務所(弁理士)", "個人", 389950, 1, 36194, "あり",
     "特許出願手続き", "源泉徴収済み。"],
    ["問題なし", "外注費", "今井友美(インボイス未登録)", "個人", 666050, 7, 61819, "あり",
     "デザイン制作/技術支援料", "デザイン料として源泉徴収済み。税額の検算推奨。"],
    ["問題なし", "支払報酬料", "高野経営労務事務所", "個人", 49500, 1, 4594, "あり",
     "(摘要なし)", "源泉徴収済み。"],
    ["問題なし", "外注費", "株式会社レイブリー", "法人", 2430994, 12, 0, "-",
     "NNA保守対応", "法人への支払い->源泉不要。"],
    ["問題なし", "外注費", "アイマツイフト(株)", "法人", 797500, 1, 0, "-",
     "Snapshot for Businessアプリ開発", "法人への支払い->源泉不要。"],
    ["問題なし", "外注費", "ペイオニアジャパン(株)", "法人", 384000, 4, 0, "-",
     "開発協力(送金サービス経由)", "法人への支払い->源泉不要。"],
    ["問題なし", "外注費", "齊藤憲生", "個人", 330000, 1, 0, "-",
     "返金", "返金取引のため源泉対象外。"],
]

for i, row in enumerate(data, 10):
    for col, val in enumerate(row, 1):
        cell = ws1.cell(row=i, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if col in (5, 7):
            cell.number_format = num_fmt
    j = row[0]
    pcell = ws1.cell(row=i, column=1)
    if j == "重大":
        pcell.font = red_font; pcell.fill = red_fill
    elif j == "要確認":
        pcell.font = yellow_font; pcell.fill = yellow_fill
    elif j == "問題なし":
        pcell.font = green_font; pcell.fill = green_fill

set_col_widths(ws1, [12, 12, 30, 20, 16, 6, 14, 10, 38, 55])

# ===== Sheet 2: Balance Trend =====
ws2 = wb.create_sheet("預り金残高推移")
ws2["A1"] = "預り金(源泉所得税等) 残高推移"
ws2["A1"].font = title_font

write_header_row(ws2, 3, ["会計年度", "期間", "期首残高", "借方(納付等)", "貸方(計上)", "期末残高", "コメント"])

balance_data = [
    ["第5期", "2021/9-2022/8", 100350, 2621375, 2552309, 31284, "期末残高31,284円。概ね適正に納付。"],
    ["第6期", "2022/9-2023/8", 31284, 2967144, 3038257, 102397, "期末残高102,397円。やや滞留。"],
    ["第7期", "2023/9-2024/8", 102397, 2790169, 2774222, 86450, "期末残高86,450円。前期とほぼ同水準。"],
    ["第8期", "2024/9-2025/8", 86450, 2717300, 2684234, 53384, "期末残高53,384円。概ね適正。"],
    ["第9期(途中)", "2025/9-2026/3", 53384, 436460, 499028, 115952, "進行期途中。期末残高115,952円。"],
]
for i, row in enumerate(balance_data, 4):
    for col, val in enumerate(row, 1):
        cell = ws2.cell(row=i, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        if 3 <= col <= 6:
            cell.number_format = num_fmt
            cell.alignment = Alignment(horizontal="right")

set_col_widths(ws2, [14, 18, 14, 16, 16, 14, 55])

ws2.cell(row=10, column=1, value="【分析】").font = bold_font
analysis = [
    "預り金の期末残高は各期とも概ね3万円-10万円で推移。大幅な滞留は見られない。",
    "ただし上記は給与の源泉所得税(年末調整分)を含む残高であり、",
    "「あしたの会計事務所」「須藤一輝」等の源泉漏れ分は含まれていない点に注意。",
    "源泉漏れ分を追加納付する場合、不納付加算税(10%/自主的5%)+延滞税が発生する可能性あり。",
]
for i, text in enumerate(analysis, 11):
    cell = ws2.cell(row=i, column=1, value=text)
    cell.font = warn_font if i == 13 else normal_font

# ===== Sheet 3: Action List =====
ws3 = wb.create_sheet("対応アクション")
ws3["A1"] = "源泉所得税 税務調査対応アクションリスト"
ws3["A1"].font = title_font

write_header_row(ws3, 3, ["優先度", "対象取引先", "アクション内容", "確認書類", "担当", "期限", "状況"])

actions = [
    ["最優先", "あしたの会計事務所", "源泉徴収漏れの確認。過去の支払調書との照合。自主修正納付の検討。", "請求書/支払調書/納付書", "", "", "未着手"],
    ["最優先", "須藤 一輝", "デザイン料の源泉徴収漏れ確認。自主修正納付の検討。", "請求書/契約書", "", "", "未着手"],
    ["最優先", "ベトナムフリーランス", "居住地/役務提供地の確認。非居住者該当性の判断。", "契約書/パスポート写し/送金記録", "", "", "未着手"],
    ["重要", "星野萌音", "業務内容の確認(204条該当性)。契約書/発注書の確認。", "業務委託契約書/発注書", "", "", "未着手"],
    ["重要", "齊藤晶子", "業務内容の確認(204条該当性)。", "業務委託契約書", "", "", "未着手"],
    ["重要", "タケダトキオ", "SE業務の内容確認(204条該当性)。", "業務委託契約書", "", "", "未着手"],
    ["重要", "伊藤 源太", "業務内容の確認。インボイス経過措置の適用確認。", "業務委託契約書", "", "", "未着手"],
    ["重要", "Biteearth 徐基源", "業務内容/非居住者該当性の確認。", "業務委託契約書", "", "", "未着手"],
    ["重要", "東新宿総合法律事務所", "支払い内容の確認(弁護士報酬 or 実費精算)。", "請求書", "", "", "未着手"],
    ["通常", "今井友美", "源泉税額の検算(10.21%)。", "請求書", "", "", "未着手"],
    ["通常", "山下 聖志", "源泉税額の検算(10.21%/20.42%)。100万円超部分の確認。", "請求書", "", "", "未着手"],
    ["通常", "全般", "預り金の納付状況確認(毎月10日 or 納期の特例)。", "納付書控え/e-Tax送信記録", "", "", "未着手"],
    ["通常", "全般", "支払調書の提出状況確認。", "支払調書控え/法定調書合計表", "", "", "未着手"],
]
for i, row in enumerate(actions, 4):
    for col, val in enumerate(row, 1):
        cell = ws3.cell(row=i, column=col, value=val)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    p = row[0]
    pcell = ws3.cell(row=i, column=1)
    if p == "最優先":
        pcell.font = red_font; pcell.fill = red_fill
    elif p == "重要":
        pcell.font = yellow_font; pcell.fill = yellow_fill

set_col_widths(ws3, [10, 24, 55, 38, 10, 12, 10])

# ===== Sheet 4: Checklist =====
ws4 = wb.create_sheet("源泉徴収チェックポイント")
ws4["A1"] = "源泉徴収 税務調査チェックポイント"
ws4["A1"].font = title_font

checklist = [
    ("1", "源泉徴収対象の報酬(所得税法204条)", ""),
    ("", "  (1) 弁護士/税理士/社労士/公認会計士等の報酬", "10.21%(100万円以下) / 20.42%(100万円超の部分)"),
    ("", "  (2) 司法書士の報酬", "(支払金額-10,000円) x 10.21%"),
    ("", "  (3) デザイン料/原稿料/翻訳料/講演料", "10.21% / 20.42%"),
    ("", "  (4) 外交員報酬", "(支払金額-12万円) x 10.21%"),
    ("", "  (5) 不動産の使用料等(個人への支払い)", "10.21% / 20.42%"),
    ("", "  (6) 非居住者への報酬(所得税法212条)", "原則 20.42%(租税条約による軽減あり)"),
    ("", "", ""),
    ("2", "源泉徴収が不要なもの", ""),
    ("", "  (1) 法人(株式会社/合同会社等)への支払い", ""),
    ("", "  (2) システム開発/プログラミング業務", "デザイン要素がなければ対象外"),
    ("", "  (3) 行政書士の報酬", "原則として源泉徴収不要"),
    ("", "  (4) 社会保険労務士法人への支払い", "法人格があれば源泉不要"),
    ("", "", ""),
    ("3", "消費税の取扱い", ""),
    ("", "  原則: 税込金額に対して源泉徴収", ""),
    ("", "  例外: 請求書で報酬額と消費税が明確に区分されている場合は税抜可", ""),
    ("", "", ""),
    ("4", "納付期限", ""),
    ("", "  原則: 翌月10日まで", ""),
    ("", "  納期の特例(常時10人未満): 1-6月分->7/10 / 7-12月分->翌1/20", ""),
    ("", "  ※士業以外の外注報酬には納期の特例なし", ""),
    ("", "", ""),
    ("5", "不納付加算税/延滞税", ""),
    ("", "  不納付加算税: 原則10%(自主納付の場合5%)", ""),
    ("", "  延滞税: 納付期限の翌日から年14.6%(2ヶ月以内は年7.3%)", ""),
]
for i, (num, item, note) in enumerate(checklist, 3):
    ws4.cell(row=i, column=1, value=num).font = section_font if num else normal_font
    ws4.cell(row=i, column=2, value=item).font = section_font if num else normal_font
    ws4.cell(row=i, column=3, value=note).font = normal_font

set_col_widths(ws4, [5, 58, 55])

output_path = "C:/Users/yuya_/claude/office/freee-auto/connectiv_withholding_tax_check.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
