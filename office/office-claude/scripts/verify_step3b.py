"""Step 3-B 検証②用: Excel 出力から G/F/E 列を抽出して検査。"""
from __future__ import annotations
import sys
from pathlib import Path
import openpyxl


def inspect_sheet(xlsx: Path, sheet_name: str, max_rows: int = 50) -> dict:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {"error": f"sheet '{sheet_name}' not found", "sheets": wb.sheetnames}
    ws = wb[sheet_name]
    rows_info = []
    for r in range(1, min(ws.max_row, max_rows) + 1):
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        e = ws.cell(row=r, column=5).value
        f = ws.cell(row=r, column=6).value
        g = ws.cell(row=r, column=7).value
        rows_info.append({
            "row": r, "A": a, "B": b, "C": c, "D": d, "E": e, "F": f, "G": g,
        })
    return {"sheet": sheet_name, "max_row": ws.max_row, "rows": rows_info}


def find_non_master_in_g(xlsx: Path) -> list:
    """全シート走査: G 列に非マスタ値が残っていないかチェック。"""
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    forbidden = {"課税売上の可能性", "別区分の可能性", "課税売上(要確認)",
                 "対象外の可能性", "非課仕入の可能性", "課税仕入の可能性",
                 "要判断", "要判断(軽減税率の可能性)"}
    hits = []
    for s in wb.sheetnames:
        ws = wb[s]
        for r in range(1, ws.max_row + 1):
            g = ws.cell(row=r, column=7).value
            if isinstance(g, str) and g in forbidden:
                hits.append({"sheet": s, "row": r, "G": g})
    return hits


def main():
    xlsx = Path(sys.argv[1])
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    print(f"FILE: {xlsx}")
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    print(f"SHEETS: {wb.sheetnames}")
    print()
    print("=== 非マスタ値の G 列残存チェック ===")
    hits = find_non_master_in_g(xlsx)
    print(f"  HITS: {len(hits)}")
    for h in hits[:20]:
        print(f"    {h}")
    print()
    if sheet:
        info = inspect_sheet(xlsx, sheet, max_rows=80)
        print(f"=== sheet: {sheet} (max_row={info.get('max_row')}) ===")
        for r in info.get("rows", []):
            # Compact printing
            line = f"  r{r['row']:>3} | F={r['F']!r:>20} | G={r['G']!r:>30} | E={str(r['E'])[:120]!r}"
            print(line)


if __name__ == "__main__":
    main()
