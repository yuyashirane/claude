"""Phase 6 完了検証スクリプト。

Phase 6 の実装が完了しているかを確認する。
- ファイル存在チェック
- import 可能チェック
- 基本動作チェック(Excel 生成)
- Phase 1〜5 の既存 tests が壊れていないことの確認(pytest 実行)
"""
from __future__ import annotations

import subprocess
import sys
import tempfile
from pathlib import Path

PROJECT_ROOT = Path(__file__).parent.parent

# ─────────────────────────────────────────────────────────────────────

def check(label: str, condition: bool) -> bool:
    status = "OK" if condition else "FAIL"
    print(f"  [{status}] {label}")
    return condition


def section(title: str) -> None:
    print(f"\n{'='*60}")
    print(f"  {title}")
    print('='*60)


# ─────────────────────────────────────────────────────────────────────
# 1. ファイル存在チェック
# ─────────────────────────────────────────────────────────────────────

def check_files() -> bool:
    section("1. ファイル存在チェック")
    required_files = [
        "skills/export/excel_report/SKILL.md",
        "skills/export/excel_report/__init__.py",
        "skills/export/excel_report/exporter.py",
        "skills/export/excel_report/sheet_builder.py",
        "skills/export/excel_report/styles.py",
        "skills/export/excel_report/sort_priority_map.py",
        "skills/export/excel_report/references/area-sheet-mapping.json",
        "tests/unit/test_excel_export.py",
        "scripts/verify_phase6_excel.py",
    ]
    all_ok = True
    for f in required_files:
        path = PROJECT_ROOT / f
        ok = path.exists()
        if not ok:
            all_ok = False
        check(f, ok)
    return all_ok


# ─────────────────────────────────────────────────────────────────────
# 2. Import チェック
# ─────────────────────────────────────────────────────────────────────

def check_imports() -> bool:
    section("2. import チェック")
    all_ok = True

    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))

    # exporter
    try:
        from skills.export.excel_report.exporter import export_to_excel
        ok = callable(export_to_excel)
        all_ok = all_ok and ok
        check("skills.export.excel_report.exporter.export_to_excel", ok)
    except ImportError as e:
        check(f"skills.export.excel_report.exporter: IMPORT ERROR - {e}", False)
        all_ok = False

    # styles
    try:
        from skills.export.excel_report.styles import SEVERITY_ROW_COLORS, DETAIL_COLUMNS, SUMMARY_COLUMNS
        ok = len(DETAIL_COLUMNS) == 19 and len(SUMMARY_COLUMNS) == 10
        all_ok = all_ok and ok
        check(f"styles: DETAIL_COLUMNS={len(DETAIL_COLUMNS)}, SUMMARY_COLUMNS={len(SUMMARY_COLUMNS)}", ok)
    except ImportError as e:
        check(f"skills.export.excel_report.styles: IMPORT ERROR - {e}", False)
        all_ok = False

    # sort_priority_map
    try:
        from skills.export.excel_report.sort_priority_map import SORT_PRIORITY_MAP, get_sort_priority
        ok = len(SORT_PRIORITY_MAP) >= 30
        all_ok = all_ok and ok
        check(f"sort_priority_map: {len(SORT_PRIORITY_MAP)} entries", ok)
        ok2 = get_sort_priority("TC-03a") == 1
        all_ok = all_ok and ok2
        check("get_sort_priority('TC-03a') == 1", ok2)
    except ImportError as e:
        check(f"sort_priority_map: IMPORT ERROR - {e}", False)
        all_ok = False

    return all_ok


# ─────────────────────────────────────────────────────────────────────
# 3. 基本動作チェック
# ─────────────────────────────────────────────────────────────────────

def check_basic_operation() -> bool:
    section("3. 基本動作チェック")
    all_ok = True

    if str(PROJECT_ROOT) not in sys.path:
        sys.path.insert(0, str(PROJECT_ROOT))

    # schema ロード
    import importlib.util
    if "schema" not in sys.modules:
        schema_path = (
            PROJECT_ROOT / "skills" / "verify" / "V1-3-rule"
            / "check-tax-classification" / "schema.py"
        )
        spec = importlib.util.spec_from_file_location("schema", schema_path)
        mod = importlib.util.module_from_spec(spec)
        sys.modules["schema"] = mod
        spec.loader.exec_module(mod)

    schema = sys.modules["schema"]
    from skills.export.excel_report.exporter import export_to_excel
    from openpyxl import load_workbook

    # 3-1. 空 findings → サマリーのみ
    with tempfile.TemporaryDirectory() as tmpdir:
        out = Path(tmpdir) / "empty.xlsx"
        result = export_to_excel([], out)
        ok = result == out and out.exists()
        all_ok = all_ok and ok
        check("空 findings: ファイル生成", ok)
        if ok:
            wb = load_workbook(out)
            ok2 = wb.sheetnames == ["サマリー"]
            all_ok = all_ok and ok2
            check("空 findings: サマリーシートのみ", ok2)

    # 3-2. Finding 1件 → サマリー + 詳細シート生成
    with tempfile.TemporaryDirectory() as tmpdir:
        f = schema.Finding(
            tc_code="TC-07",
            sub_code="TC-07a",
            severity="🔴 High",
            error_type="direct_error",
            review_level="🔴必修",
            area="A10",
            sort_priority=12,
            wallet_txn_id="verify-001",
            current_value="課対仕入10%",
            suggested_value="対象外",
            confidence=90,
            message="検収テスト: 慶弔見舞金が課税仕入になっています。",
            show_by_default=True,
        )
        out = Path(tmpdir) / "single.xlsx"
        result = export_to_excel([f], out)
        ok = result == out and out.exists()
        all_ok = all_ok and ok
        check("Finding 1件: ファイル生成", ok)
        if ok:
            wb = load_workbook(out)
            ok2 = "サマリー" in wb.sheetnames and "A10 その他経費" in wb.sheetnames
            all_ok = all_ok and ok2
            check("Finding 1件: シート構成正常", ok2)
            if ok2:
                ws = wb["A10 その他経費"]
                ok3 = ws.cell(1, 1).value == "優先度"
                ok4 = ws.cell(2, 2).value == "TC-07a"
                ok5 = ws.freeze_panes == "A2"
                all_ok = all_ok and ok3 and ok4 and ok5
                check("Finding 1件: ヘッダー「優先度」", ok3)
                check("Finding 1件: TC-07a がデータ行に配置", ok4)
                check("Finding 1件: ヘッダー固定(A2)", ok5)

    # 3-3. ValueError チェック(存在しない親ディレクトリ)
    with tempfile.TemporaryDirectory() as tmpdir:
        invalid = Path(tmpdir) / "nonexistent" / "out.xlsx"
        try:
            export_to_excel([], invalid)
            check("ValueError: 存在しない親ディレクトリ", False)
            all_ok = False
        except ValueError:
            check("ValueError: 存在しない親ディレクトリ", True)
        except Exception as e:
            check(f"ValueError: 予期しない例外 {type(e).__name__}: {e}", False)
            all_ok = False

    return all_ok


# ─────────────────────────────────────────────────────────────────────
# 4. pytest 実行
# ─────────────────────────────────────────────────────────────────────

def run_pytest() -> bool:
    section("4. pytest 実行")

    python_exec = PROJECT_ROOT / ".venv" / "Scripts" / "python"
    if not python_exec.exists():
        python_exec = sys.executable

    result = subprocess.run(
        [str(python_exec), "-m", "pytest", "tests/", "-q", "--tb=short"],
        cwd=str(PROJECT_ROOT),
        capture_output=True,
        text=True,
        encoding="utf-8",
    )

    print(result.stdout[-3000:] if len(result.stdout) > 3000 else result.stdout)
    if result.stderr:
        print(result.stderr[-500:])

    passed = result.returncode == 0
    check(f"pytest 全体: returncode={result.returncode}", passed)
    return passed


# ─────────────────────────────────────────────────────────────────────
# main
# ─────────────────────────────────────────────────────────────────────

def main() -> None:
    print("\n" + "="*60)
    print("  Phase 6 Excel 出力 完了検証")
    print("="*60)

    results = [
        check_files(),
        check_imports(),
        check_basic_operation(),
        run_pytest(),
    ]

    section("検証結果サマリー")
    labels = [
        "1. ファイル存在",
        "2. import チェック",
        "3. 基本動作",
        "4. pytest",
    ]
    all_passed = True
    for label, ok in zip(labels, results):
        check(label, ok)
        if not ok:
            all_passed = False

    print()
    if all_passed:
        print("[ALL OK]  Phase 6 検証 全 OK")
    else:
        print("[FAILED]  Phase 6 検証 FAILED -- 上記の FAIL 項目を修正してください")
    print()


if __name__ == "__main__":
    main()
