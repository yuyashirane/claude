"""サマリーシート・詳細シート生成モジュール。

仕様書 §11.2〜§11.7 に基づき、Finding リストから openpyxl ワークシートを組み立てる。
Phase 6.11a v2: rev.xlsx を仕様書として正確に再現
    - サマリー: 凡例右側配置（L-M列）、TC コードと名称を別列、Row 20〜下部テーブル
    - 詳細: Row1=タイトル / Row2=空 / Row3=ヘッダー / Row4+=データ（20列）
    - severity 表示: 重大 / 要注意 / 要確認（問題なし廃止）
    - 「同上」: 直前行と同一 sub_code の場合に観点・チェック結果を省略

責務はシート構造の生成のみ。ファイル保存は exporter.py が行う。
"""
from __future__ import annotations

import json
import re
from datetime import date as _date_cls
from pathlib import Path

from openpyxl.styles import Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from skills.export.excel_report.styles import (
    AMOUNT_ALIGNMENT,
    AMOUNT_NUMBER_FORMAT,
    DATA_FONT,
    DATA_FONT_BOLD,
    DEFAULT_ALIGNMENT,
    DETAIL_COLUMNS,
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    MEDIUM_BOTTOM_BORDER,
    SECTION_FONT,
    SEVERITY_DISPLAY,
    THIN_BORDER,
    TITLE_ALIGNMENT,
    TITLE_FONT,
    TOTAL_ROW_FILL,
    get_row_fill,
)
from skills.export.excel_report.sort_priority_map import get_sort_priority

# ─────────────────────────────────────────────────────────────────────
# 定数
# ─────────────────────────────────────────────────────────────────────

_REFS_DIR = Path(__file__).parent / "references"

# サマリーシートの固定行レイアウト
_SUMMARY_TABLE_HEADER_ROW = 20  # 下部テーブルヘッダーの行
_SUMMARY_DATA_START_ROW   = 21  # データ行の開始行

# TC 別集計マトリクスの定義（コード・名称を分離）
_TC_NAMES: list[tuple[str, str]] = [
    ("TC-01", "売上の税区分"),
    ("TC-02", "土地/住宅の非課税"),
    ("TC-03", "給与/人件費"),
    ("TC-04", "非課税/対象外の収益"),
    ("TC-05", "非課税/対象外の費用"),
    ("TC-06", "税金/租税公課"),
    ("TC-07", "福利厚生"),
]

# TC コード → 表示名（詳細シート項目名列・下部テーブル TC 名称列で使用）
_TC_DISPLAY: dict[str, str] = {code: name for code, name in _TC_NAMES}


def _load_area_mapping() -> dict[str, str]:
    path = _REFS_DIR / "area-sheet-mapping.json"
    data = json.loads(path.read_text(encoding="utf-8"))
    return data.get("mappings", {})


AREA_SHEET_MAPPING: dict[str, str] = _load_area_mapping()
AREA_ORDER: list[str] = ["A4", "A5", "A8", "A10", "A11", "A12"]

# ─────────────────────────────────────────────────────────────────────
# Finding 属性ヘルパー
# ─────────────────────────────────────────────────────────────────────

def _get_area(finding) -> str:
    return getattr(finding, "area", "")


def _get_sort_priority(finding) -> int:
    sp = getattr(finding, "sort_priority", 0)
    if sp and sp > 0:
        return sp
    return get_sort_priority(getattr(finding, "sub_code", ""))


def _get_transaction_date(finding) -> str:
    lh = getattr(finding, "link_hints", None)
    if lh is None:
        return "-"
    ps = getattr(lh, "period_start", None)
    if ps is None:
        return "-"
    return ps.strftime("%Y/%m/%d") if hasattr(ps, "strftime") else str(ps)


def _get_account_name(finding) -> str:
    lh = getattr(finding, "link_hints", None)
    if lh is None:
        return "-"
    return getattr(lh, "account_name", None) or "-"


def _get_wallet_txn_id(finding) -> str:
    v = getattr(finding, "wallet_txn_id", None)
    return v if v else "-"


def _is_medium_or_orange(severity: str) -> bool:
    return "🟡" in severity or "🟠" in severity


def _severity_display(severity: str) -> str:
    """severity 内部値（絵文字プレフィックス付き文字列）→ Excel 表示ラベルに変換。"""
    for prefix, label in SEVERITY_DISPLAY.items():
        if severity.startswith(prefix):
            return label
    return severity


def _get_tc_name(tc_code: str) -> str:
    """TC コード → TC 名称。未登録は空文字。"""
    return _TC_DISPLAY.get(tc_code, "")


# ─────────────────────────────────────────────────────────────────────
# 内部ヘルパー
# ─────────────────────────────────────────────────────────────────────

def _write_header_row(ws, columns: list, row: int = 1) -> None:
    """ヘッダー行を指定 row に書き込む（濃紺背景）。"""
    for col_idx, header, _ in columns:
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _write_section_title(ws, row: int, text: str) -> None:
    """セクション見出し（13pt bold）を書き込む。"""
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT


def _apply_data_style(cell, alignment=None) -> None:
    """データセルに標準フォント・配置・罫線を適用する。"""
    cell.font = DATA_FONT
    cell.alignment = alignment if alignment is not None else DEFAULT_ALIGNMENT
    cell.border = THIN_BORDER


def _make_fill(color: str) -> PatternFill:
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def _write_sev_header_cell(ws, row: int, col: int, val: str, color: str | None) -> None:
    """severity ヘッダーセル: 色あり=severity 色, 色なし=濃紺。"""
    cell = ws.cell(row=row, column=col, value=val)
    cell.fill = _make_fill(color) if color else HEADER_FILL
    cell.font = DATA_FONT_BOLD
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER


# ─────────────────────────────────────────────────────────────────────
# 詳細シート生成
# ─────────────────────────────────────────────────────────────────────

def build_detail_sheet(ws, findings: list, sheet_title: str = "") -> None:
    """詳細シートを生成する。

    レイアウト:
        Row 1: シートタイトル
        Row 2: 空行
        Row 3: ヘッダー行（20列、濃紺）
        Row 4+: データ行
    """
    # Row 1: シートタイトル
    if sheet_title:
        cell = ws.cell(row=1, column=1, value=sheet_title)
        cell.font = SECTION_FONT

    # Row 3: ヘッダー行
    _write_header_row(ws, DETAIL_COLUMNS, row=3)

    # 列幅設定
    for col_idx, _, width in DETAIL_COLUMNS:
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ヘッダー固定（Row 4 から freeze）
    ws.freeze_panes = "A4"

    # オートフィルタ（ヘッダー行 Row 3 を基準）
    last_col = get_column_letter(len(DETAIL_COLUMNS))
    ws.auto_filter.ref = f"A3:{last_col}3"

    # データ行（Row 4 から）
    prev_sub_code = ""
    for row_offset, finding in enumerate(findings):
        _write_detail_data_row(ws, row_offset + 4, finding, prev_sub_code)
        prev_sub_code = getattr(finding, "sub_code", "")

    # 確認状況列(R=18)にプルダウン
    if findings:
        last_data_row = len(findings) + 3
        dv = DataValidation(
            type="list",
            formula1='"〇,×,保留"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = f"R4:R{last_data_row}"
        ws.add_data_validation(dv)


def _write_detail_data_row(ws, row: int, finding, prev_sub_code: str = "") -> None:
    """詳細シートのデータ行を書き込む。

    同一 sub_code が直前行と連続する場合、観点(D)・チェック結果(E)を「同上」に置換。
    """
    severity      = getattr(finding, "severity", "")
    sub_code      = getattr(finding, "sub_code", "")
    tc_code       = getattr(finding, "tc_code", sub_code[:5] if len(sub_code) >= 5 else "")
    message       = getattr(finding, "message", "")
    current_value = getattr(finding, "current_value", "")
    suggested     = getattr(finding, "suggested_value", "")
    confidence    = getattr(finding, "confidence", 0)
    error_type    = getattr(finding, "error_type", "")
    txn_date      = _get_transaction_date(finding)
    account       = _get_account_name(finding)
    wallet_txn_id = _get_wallet_txn_id(finding)
    row_fill      = get_row_fill(severity)
    tc_name       = _get_tc_name(tc_code)

    # 「同上」: 直前行と同じ sub_code が連続している場合のみ適用
    same_as_prev = bool(prev_sub_code) and (sub_code == prev_sub_code)
    kanten_val       = "同上" if same_as_prev else (message[:40] if message else "")
    check_result_val = "同上" if same_as_prev else message

    data_values = [
        (1,  _severity_display(severity)),  # A: 優先度
        (2,  sub_code),                     # B: 項目
        (3,  tc_name),                      # C: 項目名
        (4,  kanten_val),                   # D: 観点
        (5,  check_result_val),             # E: チェック結果
        (6,  current_value),                # F: 現在の税区分
        (7,  suggested),                    # G: 推奨税区分
        (8,  txn_date),                     # H: 取引日
        (9,  account),                      # I: 勘定科目
        (10, "-"),                          # J: 取引先
        (11, ""),                           # K: 摘要（Phase 6.11b）
        (12, "-"),                          # L: 借方金額
        (13, "-"),                          # M: 貸方金額
        (14, ""),                           # N: 🔗総勘定元帳
        (15, ""),                           # O: 🔗仕訳帳
        (16, confidence),                   # P: 確信度
        (17, error_type),                   # Q: error_type
        (18, ""),                           # R: 確認状況
        (19, ""),                           # S: 担当者メモ
        (20, wallet_txn_id),               # T: walletTxnId
    ]

    for col_idx, value in data_values:
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.fill = row_fill
        cell.font = DATA_FONT
        cell.border = THIN_BORDER

        if col_idx in (12, 13):
            cell.alignment = AMOUNT_ALIGNMENT
            if isinstance(value, (int, float)):
                cell.number_format = AMOUNT_NUMBER_FORMAT
        elif col_idx == 5:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        else:
            cell.alignment = DEFAULT_ALIGNMENT

    # 修正 4: show_by_default=False でも非表示にしない（初期全行表示）


# ─────────────────────────────────────────────────────────────────────
# サマリーシート生成
# ─────────────────────────────────────────────────────────────────────

def build_summary_sheet(
    ws,
    all_findings: list,
    area_sheet_names: dict[str, str],
    company_name: str = "",
    period: str = "",
) -> None:
    """サマリーシートを生成する。

    レイアウト（固定行番号）:
        Row  1: タイトル（A1:K1 結合）+ L1="【判定凡例】"
        Row  2: 空行
        Row  3: 対象会社 | company_name  /  L="重大"(赤bg) M=説明文
        Row  4: 対象月   | period        /  L="要注意"(黄bg) M=説明文
        Row  5: チェック実行日 | 日付    /  L="要確認"(緑bg) M=説明文
        Row  6: 空行
        Row  7: 指摘サマリー（チェック項目別）
        Row  8: 空行
        Row  9: TC 集計ヘッダー（A=項目, B=チェック項目, D=重大, E=要注意, F=要確認, G=合計）
        Row 10-16: TC-01〜TC-07
        Row 17: 合計行（薄青背景）
        Row 18-19: 空行
        Row 20: 下部テーブルヘッダー
        Row 21+: データ行
    """
    # ── Row 1: レポートタイトル + 凡例ヘッダー ──────────────────────
    title = (
        f"{company_name} 消費税区分チェックレポート"
        if company_name
        else "消費税区分チェックレポート"
    )
    ws.merge_cells("A1:K1")
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = TITLE_ALIGNMENT

    # L1: 凡例ヘッダー
    leg_hdr = ws.cell(row=1, column=12, value="【判定凡例】")
    leg_hdr.font = DATA_FONT_BOLD

    # ── Row 3-5: メタ情報 + 凡例本文 ──────────────────────────────
    ws.cell(row=3, column=1, value="対象会社").font = DATA_FONT
    ws.cell(row=3, column=2, value=company_name).font = DATA_FONT

    ws.cell(row=4, column=1, value="対象月").font = DATA_FONT
    ws.cell(row=4, column=2, value=period if period else "").font = DATA_FONT

    ws.cell(row=5, column=1, value="チェック実行日").font = DATA_FONT
    ws.cell(row=5, column=2, value=_date_cls.today().strftime("%Y/%m/%d")).font = DATA_FONT

    # 凡例（L-M 列、severity 背景色）
    _legend = [
        (3, "重大",    "消費税区分の誤りの可能性が高い", "FFC7CE"),
        (4, "要注意",  "取引内容等により判断が必要",     "FFEB9C"),
        (5, "要確認",  "影響度少ない / 整合性の観点",    "C6EFCE"),
    ]
    for r, label, desc, color in _legend:
        fill = _make_fill(color)
        for col, val in [(12, label), (13, desc)]:
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill = fill
            cell.font = DATA_FONT_BOLD if col == 12 else DATA_FONT
            cell.alignment = DEFAULT_ALIGNMENT
            cell.border = THIN_BORDER

    # ── Row 7: セクション見出し ──────────────────────────────────────
    _write_section_title(ws, 7, "指摘サマリー（チェック項目別）")

    # ── Row 9: TC 集計ヘッダー ─────────────────────────────────────
    # A=項目(濃紺), B=チェック項目(濃紺), C=空, D=重大(赤), E=要注意(黄), F=要確認(緑), G=合計(濃紺)
    _tc_hdr = [
        (1, "項目",        None),
        (2, "チェック項目", None),
        (4, "重大",        "FFC7CE"),
        (5, "要注意",      "FFEB9C"),
        (6, "要確認",      "C6EFCE"),
        (7, "合計",        None),
    ]
    for col, val, color in _tc_hdr:
        _write_sev_header_cell(ws, 9, col, val, color)

    # ── Row 10-16: TC-01〜TC-07 ────────────────────────────────────
    total_red = total_yellow = total_green = total_all = 0
    for tc_row, (tc_key, tc_name) in enumerate(_TC_NAMES, start=10):
        tc_f = [f for f in all_findings if getattr(f, "tc_code", "") == tc_key]
        r = sum(1 for f in tc_f if "🔴" in getattr(f, "severity", ""))
        y = sum(1 for f in tc_f if _is_medium_or_orange(getattr(f, "severity", "")))
        g = sum(1 for f in tc_f if "🟢" in getattr(f, "severity", ""))
        t = len(tc_f)

        for col, val in [(1, tc_key), (2, tc_name), (4, r), (5, y), (6, g), (7, t)]:
            _apply_data_style(ws.cell(row=tc_row, column=col, value=val))

        total_red += r; total_yellow += y; total_green += g; total_all += t

    # ── Row 17: 合計行 ─────────────────────────────────────────────
    for col, val in [(1, "合計"), (4, total_red), (5, total_yellow), (6, total_green), (7, total_all)]:
        cell = ws.cell(row=17, column=col, value=val)
        cell.fill = TOTAL_ROW_FILL
        cell.font = DATA_FONT_BOLD
        cell.alignment = DEFAULT_ALIGNMENT
        cell.border = THIN_BORDER

    # ── Row 20: 下部テーブルヘッダー ─────────────────────────────────
    # A=エリア(濃紺), B=エリア名(濃紺), C=空, D=項目(濃紺), E=チェック項目(濃紺), F=空,
    # G=項目数(濃紺), H=重大 件数(赤), I=要注意 件数(黄), J=要確認 件数(緑),
    # K=合計件数(濃紺), L=影響金額合計(濃紺), M=確認進捗(濃紺)
    _lower_hdr = [
        (1,  "エリア",        None),
        (2,  "エリア名",      None),
        (4,  "項目",          None),
        (5,  "チェック項目",   None),
        (7,  "項目数",        None),
        (8,  "重大 件数",     "FFC7CE"),
        (9,  "要注意 件数",   "FFEB9C"),
        (10, "要確認 件数",   "C6EFCE"),
        (11, "合計件数",      None),
        (12, "影響金額合計",   None),
        (13, "確認進捗",      None),
    ]
    for col, val, color in _lower_hdr:
        _write_sev_header_cell(ws, _SUMMARY_TABLE_HEADER_ROW, col, val, color)

    # 列幅設定
    _col_widths = {
        1: 8, 2: 22, 3: 4, 4: 8, 5: 22, 6: 4,
        7: 8, 8: 12, 9: 12, 10: 12, 11: 10, 12: 16, 13: 35,
    }
    for col_idx, width in _col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ヘッダー固定
    ws.freeze_panes = f"A{_SUMMARY_DATA_START_ROW}"

    # ── Row 21+: データ行 ─────────────────────────────────────────
    groups: dict[tuple[str, str], list] = {}
    for f in all_findings:
        key = (_get_area(f), getattr(f, "tc_code", ""))
        groups.setdefault(key, []).append(f)

    row = _SUMMARY_DATA_START_ROW
    for area in AREA_ORDER:
        area_tcs = sorted({k[1] for k in groups if k[0] == area})
        for tc in area_tcs:
            findings_in_group = groups.get((area, tc), [])
            area_name_raw = AREA_SHEET_MAPPING.get(area, area)
            area_name     = re.sub(r"^A\d+\s+", "", area_name_raw)  # プレフィックス除去
            tc_name       = _TC_DISPLAY.get(tc, tc)                  # TC 表示名

            high_count = sum(1 for f in findings_in_group if "🔴" in getattr(f, "severity", ""))
            med_count  = sum(1 for f in findings_in_group if _is_medium_or_orange(getattr(f, "severity", "")))
            low_count  = sum(1 for f in findings_in_group if "🟢" in getattr(f, "severity", ""))
            total      = len(findings_in_group)
            sub_count  = len({getattr(f, "sub_code", "") for f in findings_in_group})

            row_data = [
                (1, area), (2, area_name),
                (4, tc), (5, tc_name),
                (7, sub_count),
                (8, high_count), (9, med_count), (10, low_count), (11, total),
                (12, "-"), (13, ""),
            ]
            _RIGHT_COLS = {7, 8, 9, 10, 11}
            for col_idx, value in row_data:
                cell = ws.cell(row=row, column=col_idx, value=value)
                _apply_data_style(
                    cell,
                    alignment=Alignment(horizontal="right", vertical="center")
                    if col_idx in _RIGHT_COLS
                    else DEFAULT_ALIGNMENT,
                )
            row += 1


# ─────────────────────────────────────────────────────────────────────
# Finding ソート
# ─────────────────────────────────────────────────────────────────────

def sort_findings_for_sheet(findings: list) -> list:
    """area → tc_code → sort_priority の3階層でソート。"""
    def sort_key(f):
        return (_get_area(f), getattr(f, "tc_code", ""), _get_sort_priority(f))

    return sorted(findings, key=sort_key)
