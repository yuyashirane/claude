"""Excel スタイル定数定義。

仕様書 §11.7 の視覚装飾定数を一元管理する。
openpyxl の PatternFill / Font / Alignment / Border を使用。
"""
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

# ─────────────────────────────────────────────────────────────────────
# 行色分け(severity ベース) §11.7
# ─────────────────────────────────────────────────────────────────────

SEVERITY_ROW_COLORS: dict[str, str] = {
    "🔴 High":    "FFEBEE",  # 薄赤
    "🟡 Medium":  "FFF9E6",  # 薄黄
    "🟠 Warning": "FFE0B2",  # 薄橙
    "🟢 Low":     "E8F5E9",  # 薄緑
}

# 後方互換: emoji のみでも引けるようにするエイリアス
_SEVERITY_EMOJI_ALIAS: dict[str, str] = {
    "🔴": "FFEBEE",
    "🟡": "FFF9E6",
    "🟠": "FFE0B2",
    "🟢": "E8F5E9",
}


def get_row_fill(severity: str) -> PatternFill:
    """severity 文字列から PatternFill を返す。未知は白。"""
    color = SEVERITY_ROW_COLORS.get(severity)
    if color is None:
        # emoji prefix マッチ
        for prefix, c in _SEVERITY_EMOJI_ALIAS.items():
            if severity.startswith(prefix):
                color = c
                break
    if color is None:
        color = "FFFFFF"
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


# ─────────────────────────────────────────────────────────────────────
# ヘッダー装飾
# ─────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=False)

# ─────────────────────────────────────────────────────────────────────
# 罫線
# ─────────────────────────────────────────────────────────────────────

_THIN = Side(style="thin")
_MEDIUM = Side(style="medium")

THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
MEDIUM_BOTTOM_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_MEDIUM)

# ─────────────────────────────────────────────────────────────────────
# 金額列書式
# ─────────────────────────────────────────────────────────────────────

AMOUNT_NUMBER_FORMAT = "#,##0"
AMOUNT_ALIGNMENT = Alignment(horizontal="right")

# ─────────────────────────────────────────────────────────────────────
# 確認状況列の条件付き書式色
# ─────────────────────────────────────────────────────────────────────

CONFIRMATION_OK_COLOR = "C6EFCE"    # 緑
CONFIRMATION_NG_COLOR = "FFC7CE"    # 赤

# ─────────────────────────────────────────────────────────────────────
# 詳細シート列定義(#, ヘッダー名, 幅)
# ─────────────────────────────────────────────────────────────────────

DETAIL_COLUMNS: list[tuple[int, str, int]] = [
    (1,  "優先度",           8),
    (2,  "TC",               10),
    (3,  "観点",             25),
    (4,  "チェック結果",     50),
    (5,  "取引日",           12),
    (6,  "勘定科目",         15),
    (7,  "取引先",           20),
    (8,  "摘要",             30),
    (9,  "借方金額",         12),
    (10, "貸方金額",         12),
    (11, "現在の税区分",     15),
    (12, "推奨税区分",       15),
    (13, "🔗総勘定元帳",     10),
    (14, "🔗仕訳帳",         10),
    (15, "確信度",           8),
    (16, "error_type",       15),
    (17, "確認状況",         15),
    (18, "担当者メモ",       30),
    (19, "walletTxnId",      20),
]

# ─────────────────────────────────────────────────────────────────────
# サマリーシート列定義(#, ヘッダー名, 幅)
# ─────────────────────────────────────────────────────────────────────

SUMMARY_COLUMNS: list[tuple[int, str, int]] = [
    (1,  "エリア",           10),
    (2,  "エリア名",         20),
    (3,  "TC",               10),
    (4,  "サブタイプ数",     12),
    (5,  "🔴 High 件数",     12),
    (6,  "🟡 Medium 件数",   14),
    (7,  "🟢 Low 件数",      12),
    (8,  "合計件数",         10),
    (9,  "影響金額合計",     14),
    (10, "確認進捗",         12),
]
