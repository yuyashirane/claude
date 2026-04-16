"""Excel 出力エントリポイント。

`Finding[]` → `.xlsx` の変換を担う。
仕様書 §11 C案ハイブリッド構成:
    - サマリーシート 1 枚(常に生成)
    - area 別詳細シート(実際に Finding が存在する area のみ生成)

freee リンク生成は Phase 7 以降のスコープ。
本モジュールは判定ロジックを一切含まない(出力専用)。
"""
from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook

from skills.export.excel_report.sheet_builder import (
    AREA_ORDER,
    AREA_SHEET_MAPPING,
    build_detail_sheet,
    build_summary_sheet,
    sort_findings_for_sheet,
)


def export_to_excel(
    findings: list,
    output_path: Path,
    company_name: str = "",
    period: str = "",
) -> Path:
    """Finding 配列を Excel ファイルに変換して保存する。

    Args:
        findings: Finding オブジェクトのリスト(空リストも許容)
        output_path: 出力先 .xlsx のパス
        company_name: 会社名(レポートタイトル用、省略可)
        period: 対象期間の文字列(例: "2026/02" "2026年2月期")

    Returns:
        output_path: 保存された Excel ファイルのパス(引数と同じ)

    Raises:
        ValueError: output_path の親ディレクトリが存在しない場合
        TypeError: findings が list でない場合
    """
    # ── 入力検証 ──
    if not isinstance(findings, list):
        raise TypeError(f"findings must be a list, got {type(findings).__name__}")

    output_path = Path(output_path)
    if not output_path.parent.exists():
        raise ValueError(f"出力先ディレクトリが存在しません: {output_path.parent}")

    # ── ワークブック生成 ──
    wb = Workbook()
    # デフォルトで生成される "Sheet" を削除
    default_sheet = wb.active
    wb.remove(default_sheet)

    # ── area 別に Finding を分類 ──
    area_findings: dict[str, list] = {}
    for f in findings:
        area = getattr(f, "area", "")
        if area:
            area_findings.setdefault(area, []).append(f)

    # ── サマリーシート(常に生成・最初のシート) ──
    ws_summary = wb.create_sheet("サマリー")
    area_sheet_names = {
        area: AREA_SHEET_MAPPING.get(area, area)
        for area in area_findings
    }
    build_summary_sheet(ws_summary, findings, area_sheet_names)

    # ── area 別詳細シート(Finding が存在する area のみ、仕様 §11.2 順) ──
    for area in AREA_ORDER:
        if area not in area_findings:
            continue
        sheet_name = AREA_SHEET_MAPPING.get(area, area)
        ws = wb.create_sheet(sheet_name)

        sorted_findings = sort_findings_for_sheet(area_findings[area])
        build_detail_sheet(ws, sorted_findings)

    # ── 保存 ──
    wb.save(output_path)
    return output_path
