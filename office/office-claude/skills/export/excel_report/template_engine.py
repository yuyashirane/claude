"""テンプレート駆動 Excel エンジン。

TC_template.xlsx を読み込み、Finding データを流し込む。
スタイル（色・フォント・罫線・列幅）はテンプレートから継承する。
Python 側は severity 表示変換と動的データ生成のみ担当する。
"""
from __future__ import annotations

import re
from copy import copy
from datetime import date as _date_cls
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from skills._common.lib.freee_link_generator import (
    build_group_gl_link,
    generate_gl_url,
    generate_jnl_url,
)
from skills._common.lib.finding_grouper import (
    group as _group_findings,
    is_mixing_pattern,
)

from skills.export.excel_report.styles import SEVERITY_DISPLAY
from skills.export.excel_report.sort_priority_map import get_sort_priority

# ─────────────────────────────────────────────────────────────────────
# パス・マスタ定数
# ─────────────────────────────────────────────────────────────────────

DEFAULT_TEMPLATE_PATH = Path("data/reports/template/TC_template.xlsx")

AREA_ORDER: list[str] = ["A4", "A5", "A8", "A10", "A11", "A12"]

_TC_NAMES: list[tuple[str, str]] = [
    ("TC-01", "売上の税区分"),
    ("TC-02", "土地/住宅の非課税"),
    ("TC-03", "給与/人件費"),
    ("TC-04", "非課税/対象外の収益"),
    ("TC-05", "非課税/対象外の費用"),
    ("TC-06", "税金/租税公課"),
    ("TC-07", "福利厚生"),
]
_TC_DISPLAY: dict[str, str] = {code: name for code, name in _TC_NAMES}

# ─────────────────────────────────────────────────────────────────────
# テンプレート内の固定位置（サマリーシート）
# ─────────────────────────────────────────────────────────────────────

_SUM_TITLE_ROW    = 1   # A1: レポートタイトル
_SUM_COMPANY_CELL = (3, 2)   # B3: 対象会社名
_SUM_PERIOD_CELL  = (4, 2)   # B4: 対象月
_SUM_DATE_CELL    = (5, 2)   # B5: チェック実行日
_SUM_TC_START_ROW = 10   # TC-01 データ行
_SUM_TC_TOTAL_ROW = 17   # 合計行
_SUM_TC_COL_CODE  = 1    # A: TC コード
_SUM_TC_COL_HIGH  = 4    # D: 重大件数
_SUM_TC_COL_MED   = 5    # E: 要注意件数
_SUM_TC_COL_LOW   = 6    # F: 要確認件数
_SUM_TC_COL_TOTAL = 7    # G: 合計
_SUM_LOWER_HEADER = 20   # Row 20: 下部テーブルヘッダー
_SUM_LOWER_DATA   = 21   # Row 21+: 下部テーブルデータ

# 下部テーブル列
_LT_AREA     = 1
_LT_AREANAME = 2
_LT_TC       = 4
_LT_TCNAME   = 5
_LT_SUBCNT   = 7
_LT_HIGH     = 8
_LT_MED      = 9
_LT_LOW      = 10
_LT_TOTAL    = 11
_LT_AMOUNT   = 12
_LT_PROGRESS = 13

# ─────────────────────────────────────────────────────────────────────
# テンプレート内の固定位置（詳細シート・23列）
# ─────────────────────────────────────────────────────────────────────

_DET_TITLE_ROW  = 1
_DET_HEADER_ROW = 3
_DET_DATA_START = 4

_D_PRIORITY   = 1    # A: 優先度
_D_SUBCODE    = 2    # B: 項目
_D_TCNAME     = 3    # C: 項目名
_D_VIEWPOINT  = 4    # D: 観点
_D_RESULT     = 5    # E: チェック結果
_D_CURRENT    = 6    # F: 現在の税区分
_D_SUGGESTED  = 7    # G: 推奨税区分
_D_DATE       = 8    # H: 取引日
_D_ACCOUNT    = 9    # I: 勘定科目
_D_PARTNER    = 10   # J: 取引先
_D_ITEM       = 11   # K: 品目
_D_DEPT       = 12   # L: 部門
_D_MEMO       = 13   # M: メモ
_D_DESC       = 14   # N: 摘要
_D_DEBIT      = 15   # O: 借方金額
_D_CREDIT     = 16   # P: 貸方金額
_D_LINK_GL    = 17   # Q: 🔗総勘定元帳
_D_LINK_JNL   = 18   # R: 🔗仕訳帳
_D_CONFIDENCE = 19   # S: 確信度
_D_ERRTYPE    = 20   # T: エラー型
_D_CONFIRM    = 21   # U: 確認状況
_D_STAFFMEMO  = 22   # V: 担当者メモ
_D_WALLETTXN  = 23   # W: walletTxnId

_DET_TOTAL_COLS = 23

# severity 抽出に失敗した場合の最小限フォールバック
_FALLBACK_FILLS: dict[str, PatternFill] = {
    "重大":   PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "要注意": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "要確認": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
}

# ─────────────────────────────────────────────────────────────────────
# Phase 8-B: severity → 親行 Named Style 名の対応表
# ─────────────────────────────────────────────────────────────────────
# add_named_styles.py で TC_template.xlsx に登録済みの 4 つの親行スタイルに
# Finding.severity 文字列をマップする。未知値は low スタイルにフォールバック。
SEVERITY_TO_PARENT_STYLE: dict[str, str] = {
    "🔴 Critical": "parent_row_style_critical",
    "🔴 High":     "parent_row_style_critical",
    "🟠 Warning":  "parent_row_style_warning",
    "🟡 Medium":   "parent_row_style_medium",
    "🟢 Low":      "parent_row_style_low",
}

_PARENT_STYLE_FALLBACK = "parent_row_style_low"
_CHILD_STYLE = "child_row_style"


# ─────────────────────────────────────────────────────────────────────
# Phase 8-B 実務レビュー受 (2026-04-22): 列単位 Alignment 上書き定義
# ─────────────────────────────────────────────────────────────────────
# Named Style (テンプレ) は行全体の共通スタイルのみを担い、
# 列ごとに異なる横位置・折り返しは Python 側でセル単位に上書きする。
# 既存の C 列 indent=2 と同じパターン (テンプレ哲学 + 列単位例外)。
#
# 設計根拠:
#   Phase 8-B 完了後の実務レビュー (悠皓さん確認) で視覚的改善 4 点を検出:
#     ① 子行 F/G/H/I/J/K の中央揃え
#     ③ 親行 C/D/E の wrap_text=True (折り返し)
#     ⑤ O/P (借方/貸方金額) 列の右揃え (親子両方)
#     ⑥ A/B/Q/R の中央揃え (親子両方)
#
# Phase 8-C 再利用性:
#   子行 D/E 復活 (④) / 親行リンク追加 (②) の際も本定義をそのまま拡張可能。

_PARENT_COLUMN_ALIGNMENTS: dict[int, Alignment] = {
    _D_PRIORITY:   Alignment(vertical="center", horizontal="center"),          # A 優先度
    _D_SUBCODE:    Alignment(vertical="center", horizontal="center"),          # B 項目
    _D_TCNAME:     Alignment(vertical="center", horizontal="left",  wrap_text=True),  # C 項目名
    _D_VIEWPOINT:  Alignment(vertical="center", horizontal="left",  wrap_text=True),  # D 観点
    _D_RESULT:     Alignment(vertical="center", horizontal="left",  wrap_text=True),  # E チェック結果
    _D_DEBIT:      Alignment(vertical="center", horizontal="right"),           # O 借方金額
    _D_CREDIT:     Alignment(vertical="center", horizontal="right"),           # P 貸方金額
    _D_LINK_GL:    Alignment(vertical="center", horizontal="center"),          # Q 🔗GL
    _D_LINK_JNL:   Alignment(vertical="center", horizontal="center"),          # R 🔗JNL
    _D_CONFIDENCE: Alignment(vertical="center", horizontal="center"),          # S 確信度 (Phase 8-C ⑦)
    _D_ERRTYPE:    Alignment(vertical="center", horizontal="center"),          # T エラー型 (Phase 8-C ⑦)
}

_CHILD_COLUMN_ALIGNMENTS: dict[int, Alignment] = {
    _D_PRIORITY:   Alignment(vertical="center", horizontal="center"),          # A 優先度
    _D_SUBCODE:    Alignment(vertical="center", horizontal="center"),          # B 項目
    _D_TCNAME:     Alignment(vertical="center", horizontal="left", indent=2),  # C 項目名 (既存 indent 維持)
    # D 列は Phase 8-C Fix v2 で常に空欄化したためエントリを削除 (親行 D がグループ観点を担う)
    _D_RESULT:     Alignment(vertical="center", horizontal="left", wrap_text=True),  # E チェック結果 (Phase 8-C ④)
    _D_CURRENT:    Alignment(vertical="center", horizontal="center"),          # F 現在の税区分
    _D_SUGGESTED:  Alignment(vertical="center", horizontal="center"),          # G 推奨税区分
    _D_DATE:       Alignment(vertical="center", horizontal="center"),          # H 取引日
    _D_ACCOUNT:    Alignment(vertical="center", horizontal="center"),          # I 勘定科目
    _D_PARTNER:    Alignment(vertical="center", horizontal="center"),          # J 取引先
    _D_ITEM:       Alignment(vertical="center", horizontal="center"),          # K 品目
    _D_DEBIT:      Alignment(vertical="center", horizontal="right"),           # O 借方金額
    _D_CREDIT:     Alignment(vertical="center", horizontal="right"),           # P 貸方金額
    _D_LINK_GL:    Alignment(vertical="center", horizontal="center"),          # Q 🔗GL
    _D_LINK_JNL:   Alignment(vertical="center", horizontal="center"),          # R 🔗JNL
    _D_CONFIDENCE: Alignment(vertical="center", horizontal="center"),          # S 確信度 (Phase 8-C ⑦)
    _D_ERRTYPE:    Alignment(vertical="center", horizontal="center"),          # T エラー型 (Phase 8-C ⑦)
}


_HYPERLINK_FONT = Font(
    name="Meiryo UI",
    size=10,
    color="0563C1",
    underline="single",
)


def _apply_parent_row_alignment(ws, row_idx: int) -> None:
    """親行に列単位 Alignment を上書き適用する (Named Style 後に呼ぶ)。"""
    for col_idx, alignment in _PARENT_COLUMN_ALIGNMENTS.items():
        ws.cell(row=row_idx, column=col_idx).alignment = alignment


def _apply_child_row_alignment(ws, row_idx: int) -> None:
    """子行に列単位 Alignment を上書き適用する (Named Style 後に呼ぶ)。"""
    for col_idx, alignment in _CHILD_COLUMN_ALIGNMENTS.items():
        ws.cell(row=row_idx, column=col_idx).alignment = alignment


# ─────────────────────────────────────────────────────────────────────
# 表示整形ヘルパー
# ─────────────────────────────────────────────────────────────────────

def _parse_to_date(val):
    """ISO 文字列 / 'YYYYMM' 文字列 / date オブジェクトを date に正規化する。

    Phase 8-C γ 案の期間表示拡張で使用。パース失敗時は None を返す。
    """
    if val is None:
        return None
    # 既に date/datetime のようなオブジェクト
    if hasattr(val, "year") and hasattr(val, "month"):
        return val
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        # ISO 形式 (YYYY-MM-DD)
        try:
            return _date_cls.fromisoformat(s)
        except ValueError:
            pass
        # 'YYYYMM' 形式
        if len(s) == 6 and s.isdigit():
            try:
                return _date_cls(int(s[:4]), int(s[4:6]), 1)
            except ValueError:
                return None
    return None


def format_target_month(yyyymm_or_start, period_end=None) -> str:
    """対象期間を表示用文字列に変換する（単月 / 累計 両モード対応）。

    単月モード（既存互換、1 引数）:
        format_target_month("202512") → "2025年12月"
        format_target_month("202504") → "2025年4月"
        不正入力は値をそのまま返す（防御的実装）。

    累計モード（Phase 8-C γ 案で追加、2 引数）:
        format_target_month("2025-04-01", "2025-12-31") → "2025年4月〜2025年12月"
        format_target_month(date(2025,4,1), date(2025,12,31)) → 同上
        開始月=終了月の場合は単月表示にフォールバック。
        パース失敗時は生入力を「〜」で連結したフォールバック文字列を返す。

    Args:
        yyyymm_or_start: 'YYYYMM' 文字列（単月）または期間開始（累計: 文字列 / date）
        period_end:      期間終了。指定されれば累計モードで動作

    Returns:
        表示用の期間文字列。
    """
    # ── 累計モード（2 引数） ──
    if period_end is not None:
        start = _parse_to_date(yyyymm_or_start)
        end = _parse_to_date(period_end)
        if start is None or end is None:
            # パース失敗: 生入力を連結してフォールバック
            return f"{yyyymm_or_start}〜{period_end}"
        # 同一月は単月表示にフォールバック（意味的に単月）
        if start.year == end.year and start.month == end.month:
            return f"{start.year}年{start.month}月"
        return (
            f"{start.year}年{start.month}月〜"
            f"{end.year}年{end.month}月"
        )

    # ── 単月モード（既存互換） ──
    yyyymm = yyyymm_or_start
    if isinstance(yyyymm, str) and len(yyyymm) == 6 and yyyymm.isdigit():
        year = yyyymm[:4]
        month = int(yyyymm[4:6])
        return f"{year}年{month}月"
    return yyyymm  # フォールバック: 不正入力はそのまま返す


# ─────────────────────────────────────────────────────────────────────
# Finding アクセスヘルパー
# ─────────────────────────────────────────────────────────────────────

def _severity_label(severity: str) -> str:
    for prefix, label in SEVERITY_DISPLAY.items():
        if severity.startswith(prefix):
            return label
    return severity


def _is_medium_or_orange(severity: str) -> bool:
    return "🟡" in severity or "🟠" in severity


def _sort_key(f) -> tuple:
    sp = getattr(f, "sort_priority", 0)
    if not sp or sp <= 0:
        sp = get_sort_priority(getattr(f, "sub_code", ""))
    return (getattr(f, "tc_code", ""), sp)


def _sort_findings(findings: list) -> list:
    return sorted(findings, key=_sort_key)


def _txn_date(finding) -> str:
    lh = getattr(finding, "link_hints", None)
    if lh is None:
        return ""
    ps = getattr(lh, "period_start", None)
    if ps is None:
        return ""
    return ps.strftime("%Y/%m/%d") if hasattr(ps, "strftime") else str(ps)


def _account_name(finding) -> str:
    lh = getattr(finding, "link_hints", None)
    if lh is None:
        return ""
    return getattr(lh, "account_name", None) or ""


# ─────────────────────────────────────────────────────────────────────
# スタイル抽出・適用
# ─────────────────────────────────────────────────────────────────────

def _copy_fill(cell) -> PatternFill | None:
    """セルの fill を copy して返す。RGB 以外（theme / indexed カラー）は None。"""
    try:
        f = copy(cell.fill)
        if f.fill_type not in ("solid",):
            return None
        # type が "rgb" 以外（"theme", "indexed" 等）はフォールバックに任せる
        if getattr(f.fgColor, "type", None) != "rgb":
            return None
        return f
    except Exception:
        return None


def _extract_severity_fills(wb) -> dict[str, PatternFill]:
    """テンプレートの詳細シート Row 4+ から severity ラベル → fill を抽出する。

    抽出に失敗したラベルはフォールバック値を使用。
    """
    fills: dict[str, PatternFill] = {}
    labels_needed = {"重大", "要注意", "要確認"}

    for sheet_name in wb.sheetnames:
        if sheet_name in ("サマリー", "参考"):
            continue
        ws = wb[sheet_name]
        for row_idx in range(_DET_DATA_START, ws.max_row + 1):
            cell = ws.cell(row_idx, _D_PRIORITY)
            label = cell.value if isinstance(cell.value, str) else ""
            if label in labels_needed and label not in fills:
                f = _copy_fill(cell)
                fills[label] = f if f is not None else _FALLBACK_FILLS[label]
        if fills.keys() >= labels_needed:
            break

    for label in labels_needed:
        if label not in fills:
            fills[label] = _FALLBACK_FILLS[label]
    return fills


def _extract_row_styles(ws, row_idx: int, col_count: int) -> list[dict]:
    """指定行のスタイルを col_count 列分抽出する。"""
    styles = []
    for col in range(1, col_count + 1):
        cell = ws.cell(row_idx, col)
        d: dict = {}
        for attr in ("font", "fill", "border", "alignment"):
            try:
                d[attr] = copy(getattr(cell, attr))
            except Exception:
                pass
        try:
            d["number_format"] = cell.number_format
        except Exception:
            pass
        styles.append(d)
    return styles


def _apply_row_style(cell, style: dict) -> None:
    for attr in ("font", "fill", "border", "alignment"):
        v = style.get(attr)
        if v is not None:
            try:
                setattr(cell, attr, copy(v))
            except Exception:
                pass
    nf = style.get("number_format")
    if nf:
        try:
            cell.number_format = nf
        except Exception:
            pass


# ─────────────────────────────────────────────────────────────────────
# テンプレート情報抽出
# ─────────────────────────────────────────────────────────────────────

def _build_area_sheet_map(wb) -> dict[str, str]:
    """テンプレートのシート名から {area_code: sheet_name} を構築する。"""
    mapping: dict[str, str] = {}
    for name in wb.sheetnames:
        m = re.match(r"^(A\d+)\s+", name)
        if m:
            mapping[m.group(1)] = name
    return mapping


# ─────────────────────────────────────────────────────────────────────
# サマリーシート更新
# ─────────────────────────────────────────────────────────────────────

def _fill_summary(
    ws,
    findings: list,
    company_name: str,
    period: str,
    area_sheet_map: dict[str, str],
    lower_row_styles: list[dict],
) -> None:
    """サマリーシートの動的セルを更新する。固定スタイルはテンプレートから継承。"""

    # Row 1: タイトル（値のみ上書き）
    title = (
        f"{company_name} 消費税区分チェックレポート"
        if company_name
        else "消費税区分チェックレポート"
    )
    ws.cell(_SUM_TITLE_ROW, 1).value = title

    # Row 3-5: メタ情報（値のみ上書き）
    ws.cell(_SUM_COMPANY_CELL[0], _SUM_COMPANY_CELL[1]).value = company_name
    ws.cell(_SUM_PERIOD_CELL[0],  _SUM_PERIOD_CELL[1]).value  = format_target_month(period) if period else ""
    ws.cell(_SUM_DATE_CELL[0],    _SUM_DATE_CELL[1]).value    = _date_cls.today().strftime("%Y/%m/%d")

    # Row 10-16: TC 別件数（D/E/F/G 列を上書き）
    total_r = total_y = total_g = total_all = 0
    for tc_idx, (tc_code, _) in enumerate(_TC_NAMES):
        row = _SUM_TC_START_ROW + tc_idx
        tc_f = [f for f in findings if getattr(f, "tc_code", "") == tc_code]
        r = sum(1 for f in tc_f if "🔴" in getattr(f, "severity", ""))
        y = sum(1 for f in tc_f if _is_medium_or_orange(getattr(f, "severity", "")))
        g = sum(1 for f in tc_f if "🟢" in getattr(f, "severity", ""))
        t = len(tc_f)
        ws.cell(row, _SUM_TC_COL_CODE).value  = tc_code   # 末尾スペース除去
        ws.cell(row, _SUM_TC_COL_HIGH).value  = r
        ws.cell(row, _SUM_TC_COL_MED).value   = y
        ws.cell(row, _SUM_TC_COL_LOW).value   = g
        ws.cell(row, _SUM_TC_COL_TOTAL).value = t
        total_r += r; total_y += y; total_g += g; total_all += t

    # Row 17: 合計（値のみ上書き）
    ws.cell(_SUM_TC_TOTAL_ROW, _SUM_TC_COL_HIGH).value  = total_r
    ws.cell(_SUM_TC_TOTAL_ROW, _SUM_TC_COL_MED).value   = total_y
    ws.cell(_SUM_TC_TOTAL_ROW, _SUM_TC_COL_LOW).value   = total_g
    ws.cell(_SUM_TC_TOTAL_ROW, _SUM_TC_COL_TOTAL).value = total_all

    # Row 21+: 既存データ行を削除して再生成
    last_row = ws.max_row
    if last_row >= _SUM_LOWER_DATA:
        ws.delete_rows(_SUM_LOWER_DATA, last_row - _SUM_LOWER_DATA + 1)

    _fill_lower_table(ws, findings, area_sheet_map, lower_row_styles)


def _fill_lower_table(
    ws,
    findings: list,
    area_sheet_map: dict[str, str],
    row_styles: list[dict],
) -> None:
    """サマリー下部テーブル（Row 21+）にエリア別集計行を生成する。"""
    groups: dict[tuple[str, str], list] = {}
    for f in findings:
        key = (getattr(f, "area", ""), getattr(f, "tc_code", ""))
        groups.setdefault(key, []).append(f)

    row = _SUM_LOWER_DATA
    for area in AREA_ORDER:
        area_tcs = sorted({k[1] for k in groups if k[0] == area})
        for tc in area_tcs:
            tc_f = groups.get((area, tc), [])
            sheet_name = area_sheet_map.get(area, area)
            area_disp  = re.sub(r"^A\d+\s+", "", sheet_name)
            tc_name    = _TC_DISPLAY.get(tc, tc)
            sub_count  = len({getattr(f, "sub_code", "") for f in tc_f})
            r_cnt = sum(1 for f in tc_f if "🔴" in getattr(f, "severity", ""))
            y_cnt = sum(1 for f in tc_f if _is_medium_or_orange(getattr(f, "severity", "")))
            g_cnt = sum(1 for f in tc_f if "🟢" in getattr(f, "severity", ""))
            total = len(tc_f)

            row_data = {
                _LT_AREA:     area,
                _LT_AREANAME: area_disp,
                _LT_TC:       tc,
                _LT_TCNAME:   tc_name,
                _LT_SUBCNT:   sub_count,
                _LT_HIGH:     r_cnt,
                _LT_MED:      y_cnt,
                _LT_LOW:      g_cnt,
                _LT_TOTAL:    total,
                _LT_AMOUNT:   "-",
                _LT_PROGRESS: "",
            }

            for col in range(1, _LT_PROGRESS + 1):
                cell = ws.cell(row, col)
                if col in row_data:
                    cell.value = row_data[col]
                if row_styles and col <= len(row_styles):
                    _apply_row_style(cell, row_styles[col - 1])

            row += 1


# ─────────────────────────────────────────────────────────────────────
# 詳細シート更新
# ─────────────────────────────────────────────────────────────────────

def _fill_detail_sheet(
    ws,
    findings: list,
    severity_fills: dict[str, PatternFill],
) -> None:
    """詳細シートのサンプル行を削除し、Finding データを書き込む。

    Row 1(タイトル)・Row 2(空行)・Row 3(ヘッダー) はテンプレートから継承。
    Row 4+ をデータ行として上書きする。
    """
    # Row 4+ のサンプルデータを削除（スタイルを先に保存してから削除）
    row_styles = _extract_row_styles(ws, _DET_DATA_START, _DET_TOTAL_COLS)
    last_row = ws.max_row
    if last_row >= _DET_DATA_START:
        ws.delete_rows(_DET_DATA_START, last_row - _DET_DATA_START + 1)

    # Finding を書き込む
    prev_sub_code = ""
    for offset, finding in enumerate(findings):
        _write_finding_row(ws, _DET_DATA_START + offset, finding,
                           prev_sub_code, severity_fills, row_styles)
        prev_sub_code = getattr(finding, "sub_code", "")

    # 確認状況列（U=21）にプルダウン
    if findings:
        last_data_row = len(findings) + _DET_DATA_START - 1
        dv = DataValidation(
            type="list",
            formula1='"〇,×,保留"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = f"U{_DET_DATA_START}:U{last_data_row}"
        ws.add_data_validation(dv)


def _write_finding_row(
    ws,
    row: int,
    finding,
    prev_sub_code: str,
    severity_fills: dict[str, PatternFill],
    row_styles: list[dict] | None = None,
) -> None:
    """Finding 1件を詳細シートの指定行に書き込む。"""
    severity  = getattr(finding, "severity", "")
    sub_code  = getattr(finding, "sub_code", "")
    tc_code   = getattr(finding, "tc_code",
                        sub_code[:5] if len(sub_code) >= 5 else "")
    message   = getattr(finding, "message", "")
    same_prev = bool(prev_sub_code) and sub_code == prev_sub_code

    sev_label = _severity_label(severity)
    row_fill  = severity_fills.get(sev_label)

    # 借方・貸方金額（Phase 6.11b）
    debit_val  = getattr(finding, "debit_amount",  None)
    credit_val = getattr(finding, "credit_amount", None)

    values: dict[int, object] = {
        _D_PRIORITY:   sev_label,
        _D_SUBCODE:    sub_code,
        _D_TCNAME:     _TC_DISPLAY.get(tc_code, ""),
        _D_VIEWPOINT:  "同上" if same_prev else (message[:40] if message else ""),
        _D_RESULT:     "同上" if same_prev else message,
        _D_CURRENT:    getattr(finding, "current_value", ""),
        _D_SUGGESTED:  getattr(finding, "suggested_value", ""),
        _D_DATE:       _txn_date(finding),
        _D_ACCOUNT:    _account_name(finding),
        _D_PARTNER:    "",
        _D_ITEM:       "",
        _D_DEPT:       "",
        _D_MEMO:       "",
        _D_DESC:       "",
        _D_DEBIT:      debit_val  if debit_val  is not None else "",
        _D_CREDIT:     credit_val if credit_val is not None else "",
        _D_LINK_GL:    "",
        _D_LINK_JNL:   "",
        _D_CONFIDENCE: getattr(finding, "confidence", 0),
        _D_ERRTYPE:    getattr(finding, "error_type", ""),
        _D_CONFIRM:    "",
        _D_STAFFMEMO:  "",
        _D_WALLETTXN:  getattr(finding, "wallet_txn_id", ""),
    }

    for col in range(1, _DET_TOTAL_COLS + 1):
        cell = ws.cell(row, col, values.get(col, ""))
        if row_styles and col <= len(row_styles):
            _apply_row_style(cell, row_styles[col - 1])
        if row_fill is not None:
            try:
                cell.fill = copy(row_fill)
            except Exception:
                pass

    # O/P 列: 数値が書かれた場合のみ桁区切り書式を設定（テンプレートの "General" を上書き）
    if debit_val is not None:
        ws.cell(row, _D_DEBIT).number_format = "#,##0"
    if credit_val is not None:
        ws.cell(row, _D_CREDIT).number_format = "#,##0"

    # Q/R 列: freee へのハイパーリンクを設定（Phase 7）
    # _apply_row_style() の後に設定することでスタイル（フォント種別・罫線等）を壊さない。
    # Font は明示指定：テンプレートの他セルと揃えるため name/size を継承しつつ色・下線を追加。
    link_hints = getattr(finding, "link_hints", None)

    url_gl = generate_gl_url(link_hints)
    if url_gl:
        cell_gl = ws.cell(row, _D_LINK_GL)
        cell_gl.value = "🔗"
        cell_gl.hyperlink = url_gl
        cell_gl.font = Font(
            name="Meiryo UI",
            size=10,
            color="0563C1",
            underline="single",
        )

    # finding.deal_id を渡す: build_link_hints("general_ledger") は deal_id を含まないため、
    # Finding から直接取得してピンポイント URL を生成する
    url_jnl = generate_jnl_url(link_hints, deal_id=getattr(finding, "deal_id", None))
    if url_jnl:
        cell_jnl = ws.cell(row, _D_LINK_JNL)
        cell_jnl.value = "🔗"
        cell_jnl.hyperlink = url_jnl
        cell_jnl.font = Font(
            name="Meiryo UI",
            size=10,
            color="0563C1",
            underline="single",
        )


# ═════════════════════════════════════════════════════════════════════
# Phase 8-B: 親子行描画（R3 二層責務）
#
# finding_grouper.group() が返す FindingGroup を受け取り、
# 親行 1 + 子行 N を順に描画する。親行は NamedStyle "parent_row_style_*"、
# 子行は "child_row_style" を適用し、仕様案 Z（子行は白地）を保つ。
#
# pure helper の設計原則（戦略 Claude §pure helper §NG/OK パターン 準拠）:
#     - _parent_row_summary / _parent_row_observation / _parent_row_check_result
#       は worksheet/cell に触れず、FindingGroup だけから文字列を組み立てる。
#     - _write_parent_row / _write_child_row は上記 pure helper の結果を
#       ws.cell(...) に書き込む薄い I/O 層。
# ═════════════════════════════════════════════════════════════════════


def _group_account(group) -> str:
    """FindingGroup の代表勘定科目名を取り出す pure helper。

    子 Finding の link_hints.account_name を参照。全子で同じ想定
    （group_key に account が含まれるため）。
    """
    if not group.findings:
        return ""
    first = group.findings[0]
    lh = getattr(first, "link_hints", None)
    if lh is None:
        return ""
    return getattr(lh, "account_name", "") or ""


def _parent_row_summary(group) -> str:
    """親行 C 列用のサマリー文字列 (C-β-3 form)。pure helper。

    Pattern A (単方向):
        "{account} — {count} 件・合計 ¥{total:,}（{current}→{suggested}）"
    Pattern B (混在検知):
        "{account} — {count} 件・合計 ¥{total:,}（税区分混在）"

    total は total_debit + total_credit。単一 Finding は通常どちらか片方なので
    合算で実態と一致する。
    """
    account = _group_account(group) or "（科目名なし）"
    total = group.total_debit + group.total_credit
    if is_mixing_pattern(group):
        tail = "（税区分混在）"
    else:
        first = group.findings[0] if group.findings else None
        cur = getattr(first, "current_value", "") if first else ""
        sug = getattr(first, "suggested_value", "") if first else ""
        tail = f"（{cur}→{sug}）"
    return f"{account} — {group.count} 件・合計 ¥{total:,}{tail}"


def _parent_row_observation(group) -> str:
    """親行 D 列用の観点文字列。pure helper。

    Pattern A: 「{TC名称}の税区分誤り」
    Pattern B: 「同一科目に税区分混在」
    """
    if is_mixing_pattern(group):
        return "同一科目に税区分混在"
    tc_name = _TC_DISPLAY.get(group.tc_code, "")
    return f"{tc_name}の税区分誤り" if tc_name else "税区分誤り"


def _parent_row_check_result(group) -> str:
    """親行 E 列用のチェック結果文字列。pure helper。

    Pattern A: "{count} 件を「{current}」→「{suggested}」へ修正要確認"
    Pattern B: "{variants} 種類の税区分混在 — 勘定科目のルール確認要"
    """
    if is_mixing_pattern(group):
        variants = {
            getattr(f, "current_value", "") for f in group.findings
        }
        variants.discard("")
        n = len(variants) if variants else group.count
        return f"{n} 種類の税区分混在 — 勘定科目のルール確認要"
    first = group.findings[0] if group.findings else None
    cur = (getattr(first, "current_value", "") if first else "") or "-"
    sug = (getattr(first, "suggested_value", "") if first else "") or "-"
    return f"{group.count} 件を「{cur}」→「{sug}」へ修正要確認"


def _write_parent_row(ws, row: int, group, ctx=None) -> None:
    """親行 1 行を詳細シートの指定行に描画する。

    - A 列: severity 表示ラベル（重大/要注意/要確認）
    - B 列: 代表 sub_code
    - C 列: C-β-3 サマリー（件数・合計・論点パターン）
    - D 列: 観点
    - E 列: チェック結果
    - O/P 列: 集計合計（total_debit / total_credit）
    - Q 列: GL リンク（Phase 8-C ②、ctx.period_start/period_end で会計期間全体）
    - それ以外: 空欄
    - 全セルに Named Style "parent_row_style_{severity}" を適用

    Args:
        ctx: Phase 8-C で追加。CheckContext を渡すと GL リンクが会計期間全体の
             範囲で生成される。None の場合は子 Finding[0] の link_hints 期間
             (単月スコープ) にフォールバック。
    """
    style_name = SEVERITY_TO_PARENT_STYLE.get(
        group.severity, _PARENT_STYLE_FALLBACK,
    )
    sev_label = _severity_label(group.severity)

    values: dict[int, object] = {
        _D_PRIORITY:  sev_label,
        _D_SUBCODE:   group.sub_code,
        _D_TCNAME:    _parent_row_summary(group),
        _D_VIEWPOINT: _parent_row_observation(group),
        _D_RESULT:    _parent_row_check_result(group),
        _D_DEBIT:     group.total_debit if group.total_debit else "",
        _D_CREDIT:    group.total_credit if group.total_credit else "",
    }

    for col in range(1, _DET_TOTAL_COLS + 1):
        cell = ws.cell(row, col)
        cell.value = values.get(col, "")
        cell.style = style_name

    if group.total_debit:
        ws.cell(row, _D_DEBIT).number_format = "#,##0"
    if group.total_credit:
        ws.cell(row, _D_CREDIT).number_format = "#,##0"

    # Phase 8-C ②: 親行 Q 列に GL ハイパーリンクを追加
    # Named Style と Alignment の適用後に font を上書きしても style name は保持される。
    gl_url = build_group_gl_link(group, ctx=ctx)
    if gl_url:
        cell_gl = ws.cell(row, _D_LINK_GL)
        cell_gl.value = "GL"  # 判断 4: ラベルは "GL" 固定
        cell_gl.hyperlink = gl_url
        cell_gl.font = _HYPERLINK_FONT

    # 列単位 Alignment 上書き (③⑤⑥⑦ 対応、Named Style の後に適用)
    _apply_parent_row_alignment(ws, row)
    # 折り返し表示のため行高を Excel 自動計算に委ねる (③ 対応)
    ws.row_dimensions[row].height = None


def _child_row_d_e(message: str, prev_message) -> tuple[str, str]:
    """子行 D/E 列の値を決定する pure helper。

    Phase 8-C Fix v2 確定仕様:
        - D 列は常に空欄（親行 D がグループ観点を担うため、子行は冗長）
        - E 列:
            ・message 空 → 空欄
            ・直前子行と同一 message → 「同上」圧縮
            ・それ以外 → Finding.message 全文

    親行 D と子行 D の責務分離:
        親行 D = グループ観点（"給与/人件費の税区分誤り" など、Phase 8-B 仕様）
        子行 D = 常に空欄（本ヘルパーが保証）
        子行 E = Finding.message 全文（詳細理由）

    Args:
        message:      今回の Finding.message（空文字列可）
        prev_message: 直前の子行 message。None または未設定なら最初の子行扱い。

    Returns:
        (d_value, e_value) タプル。d_value は常に空文字列。
    """
    if not message:
        return ("", "")
    # 同一 message の 2 件目以降は E 列のみ "同上" で圧縮（D は空欄のまま）
    if prev_message is not None and prev_message != "" and message == prev_message:
        return ("", "同上")
    return ("", message)


def _write_child_row(ws, row: int, finding, prev_message=None) -> None:
    """子行 1 行を詳細シートの指定行に描画する。

    - A/B/C 列: 空欄（親行に集約済み）
    - D/E 列: Phase 8-C ④ で復活。Finding.message を表示。
              同一グループ内同一 message の 2 件目以降は「同上」で圧縮。
    - F/G 列: 個別の現在/推奨税区分
    - H 列: 取引日
    - I 列: 勘定科目
    - J〜N 列: 空欄（将来拡張予定: 取引先・品目・部門・メモ・摘要）
    - O/P 列: 個別の借方/貸方金額
    - Q/R 列: freee ハイパーリンク
    - S 列: 確信度
    - T 列: エラー型
    - U/V 列: 空欄（スタッフ入力欄）
    - W 列: walletTxnId
    - 全セルに Named Style "child_row_style"（白地）を適用

    Args:
        prev_message: 直前子行の Finding.message。「同上」圧縮の判定に使う。
                      None なら圧縮しない（最初の子行扱い）。
    """
    debit_val  = getattr(finding, "debit_amount",  None)
    credit_val = getattr(finding, "credit_amount", None)

    # Phase 8-C ④: 子行 D/E に Finding.message 由来の文言を復活
    message = getattr(finding, "message", "") or ""
    d_val, e_val = _child_row_d_e(message, prev_message)

    values: dict[int, object] = {
        _D_PRIORITY:   "",
        _D_SUBCODE:    "",
        _D_TCNAME:     "",
        _D_VIEWPOINT:  d_val,
        _D_RESULT:     e_val,
        _D_CURRENT:    getattr(finding, "current_value", ""),
        _D_SUGGESTED:  getattr(finding, "suggested_value", ""),
        _D_DATE:       _txn_date(finding),
        _D_ACCOUNT:    _account_name(finding),
        _D_PARTNER:    "",
        _D_ITEM:       "",
        _D_DEPT:       "",
        _D_MEMO:       "",
        _D_DESC:       "",
        _D_DEBIT:      debit_val  if debit_val  is not None else "",
        _D_CREDIT:     credit_val if credit_val is not None else "",
        _D_LINK_GL:    "",
        _D_LINK_JNL:   "",
        _D_CONFIDENCE: getattr(finding, "confidence", 0),
        _D_ERRTYPE:    getattr(finding, "error_type", ""),
        _D_CONFIRM:    "",
        _D_STAFFMEMO:  "",
        _D_WALLETTXN:  getattr(finding, "wallet_txn_id", ""),
    }

    for col in range(1, _DET_TOTAL_COLS + 1):
        cell = ws.cell(row, col)
        cell.value = values.get(col, "")
        cell.style = _CHILD_STYLE

    if debit_val is not None:
        ws.cell(row, _D_DEBIT).number_format = "#,##0"
    if credit_val is not None:
        ws.cell(row, _D_CREDIT).number_format = "#,##0"

    # Q/R 列: freee ハイパーリンク（Phase 7 ロジック継承）
    link_hints = getattr(finding, "link_hints", None)

    url_gl = generate_gl_url(link_hints)
    if url_gl:
        cell_gl = ws.cell(row, _D_LINK_GL)
        cell_gl.value = "🔗"
        cell_gl.hyperlink = url_gl
        cell_gl.font = Font(
            name="Meiryo UI", size=10, color="0563C1", underline="single",
        )

    url_jnl = generate_jnl_url(
        link_hints, deal_id=getattr(finding, "deal_id", None),
    )
    if url_jnl:
        cell_jnl = ws.cell(row, _D_LINK_JNL)
        cell_jnl.value = "🔗"
        cell_jnl.hyperlink = url_jnl
        cell_jnl.font = Font(
            name="Meiryo UI", size=10, color="0563C1", underline="single",
        )

    # 列単位 Alignment 上書き (①⑤⑥⑦ 対応、Named Style の後に適用)
    _apply_child_row_alignment(ws, row)
    # E の wrap_text 対応: 行高を Excel 自動計算に委ねる
    ws.row_dimensions[row].height = None


def _fill_detail_sheet_grouped(ws, findings: list, ctx=None) -> None:
    """Phase 8-B: Finding をグループ化し、親行→子行の順で描画する。

    事前に `_sort_findings` でソート済みの findings を受け取る想定。
    finding_grouper.group() は挿入順を保持するため、ソート結果がそのまま
    グループ順序になる。

    Phase 8-C: ctx (CheckContext) を受け取り、親行 GL リンクに会計期間全体
    を渡せるよう拡張。既存の子行 GL リンクは link_hints (単月) ベースのまま。
    """
    # Row 4+ のサンプルデータを削除
    last_row = ws.max_row
    if last_row >= _DET_DATA_START:
        ws.delete_rows(
            _DET_DATA_START, last_row - _DET_DATA_START + 1,
        )

    groups = _group_findings(findings)

    row = _DET_DATA_START
    for g in groups:
        _write_parent_row(ws, row, g, ctx=ctx)
        row += 1
        prev_message = None
        for child in g.findings:
            _write_child_row(ws, row, child, prev_message=prev_message)
            prev_message = getattr(child, "message", None) or ""
            row += 1

    # U 列（確認状況）にプルダウン。親行 U は空欄だが範囲に含めても害なし。
    if findings:
        last_data_row = row - 1
        dv = DataValidation(
            type="list",
            formula1='"〇,×,保留"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = f"U{_DET_DATA_START}:U{last_data_row}"
        ws.add_data_validation(dv)


# ─────────────────────────────────────────────────────────────────────
# メインエントリポイント
# ─────────────────────────────────────────────────────────────────────

def build_output(
    findings: list,
    output_path: Path,
    company_name: str = "",
    period: str = "",
    template_path: Path | None = None,
    ctx=None,
) -> Path:
    """テンプレートを読み込み、Finding データを流し込んで Excel を生成する。

    Args:
        findings:      Finding オブジェクトのリスト（空リストも許容）
        output_path:   出力先 .xlsx のパス
        company_name:  会社名（タイトル・メタ情報に反映）
        period:        対象期間文字列（例: "2026/02"）
        template_path: テンプレートファイルのパス。None の場合はデフォルト使用
        ctx:           CheckContext（Phase 8-C 追加）。渡されると親行 GL リンク
                      が会計期間全体（ctx.period_start/period_end）で生成される。

    Returns:
        output_path: 保存されたファイルのパス

    Raises:
        FileNotFoundError: テンプレートファイルが存在しない場合
        ValueError:        output_path の親ディレクトリが存在しない場合
    """
    tpl = Path(template_path) if template_path else DEFAULT_TEMPLATE_PATH
    if not tpl.exists():
        raise FileNotFoundError(f"テンプレートが見つかりません: {tpl}")

    output_path = Path(output_path)
    if not output_path.parent.exists():
        raise ValueError(f"出力先ディレクトリが存在しません: {output_path.parent}")

    # ── テンプレート読み込み ──
    wb = load_workbook(tpl)
    area_sheet_map = _build_area_sheet_map(wb)

    # ── severity fill を抽出（詳細シートの Row 4+ から） ──
    severity_fills = _extract_severity_fills(wb)

    # ── area 別 Finding を分類 ──
    area_findings: dict[str, list] = {}
    for f in findings:
        area = getattr(f, "area", "")
        if area:
            area_findings.setdefault(area, []).append(f)

    # ── サマリーシート更新 ──
    ws_sum = wb["サマリー"]
    lower_styles = _extract_row_styles(ws_sum, _SUM_LOWER_DATA,
                                       _LT_PROGRESS)
    _fill_summary(ws_sum, findings, company_name, period,
                  area_sheet_map, lower_styles)

    # ── 詳細シート処理 ──
    for area in AREA_ORDER:
        sheet_name = area_sheet_map.get(area)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        if area not in area_findings:
            wb.remove(ws)
        else:
            sorted_f = _sort_findings(area_findings[area])
            # Phase 8-B: 親子行レイアウト。旧 _fill_detail_sheet は dead code として
            # 残置（Phase 8-C 累計モデル化で再利用される可能性があるため）。
            # Phase 8-C: ctx を渡すと親行 Q 列に会計期間全体の GL リンクが付く。
            _fill_detail_sheet_grouped(ws, sorted_f, ctx=ctx)

    # ── 参考シートはそのまま（何もしない） ──

    wb.save(output_path)
    return output_path
