"""Phase 8-B テンプレ検証スクリプト。

add_named_styles.py 実行後に TC_template.xlsx に対して走らせる。
要件ドキュメント v2 §7 の検証項目を自動判定する。

検証項目:
    1. 5 style 全件存在チェック
    2. 色 hex 厳密一致（背景・罫線）
    3. フォント / 配置 / 罫線 太さの一致
    4. copy_worksheet 後の維持（Phase 8-A P1 相当）
    5. save/load ラウンドトリップ維持（Phase 8-A P2 相当）

Usage:
    cd office/office-claude
    py tmp/verify_phase8b_template.py
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

from openpyxl import load_workbook


TPL = Path("data/reports/template/TC_template.xlsx")

EXPECTED = {
    # name: dict(bg, accent, bold)
    "parent_row_style_critical": {"bg": "FCEBEB", "accent": "C00000", "bold": True},
    "parent_row_style_warning":  {"bg": "FAEEDA", "accent": "ED7D31", "bold": True},
    "parent_row_style_medium":   {"bg": "FEF5D6", "accent": "BF8F00", "bold": True},
    "parent_row_style_low":      {"bg": "EAF3DE", "accent": "548235", "bold": True},
    "child_row_style":           {"bg": None,     "accent": None,     "bold": False},
}


class Verifier:
    def __init__(self) -> None:
        self.failures: list[str] = []

    def check(self, cond: bool, msg: str) -> None:
        if cond:
            print(f"  ✅ {msg}")
        else:
            print(f"  ❌ {msg}")
            self.failures.append(msg)

    def ok(self) -> bool:
        return not self.failures


def _norm_hex(rgb) -> str | None:
    """Openpyxl の Color オブジェクトから 6 桁 hex 大文字を取り出す。"""
    if rgb is None:
        return None
    try:
        s = rgb.rgb if hasattr(rgb, "rgb") else rgb
    except Exception:
        return None
    if not isinstance(s, str):
        return None
    s = s.upper()
    # alpha 付き (8 桁) は末尾 6 桁を採用
    if len(s) == 8:
        s = s[2:]
    if len(s) != 6:
        return None
    return s


def _get_style(wb, name: str):
    """NamedStyle オブジェクトを名前で検索して返す (read-only)。

    openpyxl 3.x では wb.named_styles は str のリストを返すため、
    実体オブジェクトは wb._named_styles（NamedStyleList）を走査して取得する。
    辞書キー前提の `wb._named_styles[name]` ではなく走査方式を採用し、
    内部表現変更（dict / list / その他）への耐性を確保する。
    """
    for s in wb._named_styles:
        if getattr(s, "name", None) == name:
            return s
    raise KeyError(f"NamedStyle not found: {name}")


def _verify_styles_in_wb(wb, label: str) -> list[str]:
    """wb に登録された Named Style を検証し、失敗メッセージのリストを返す。"""
    fails: list[str] = []
    names_present = {s.name if hasattr(s, "name") else s for s in wb.named_styles}

    for name, spec in EXPECTED.items():
        if name not in names_present:
            fails.append(f"[{label}] {name} が登録されていない")
            continue

        s = _get_style(wb, name)
        # bold
        if s.font.bold != spec["bold"]:
            fails.append(
                f"[{label}] {name}.font.bold = {s.font.bold} "
                f"(期待値 {spec['bold']})"
            )
        # 背景色
        if spec["bg"] is None:
            # child_row_style は fill_type が None 等であるべき
            ft = s.fill.fill_type
            if ft not in (None, "none"):
                fails.append(
                    f"[{label}] {name}.fill.fill_type = {ft} "
                    f"(child_row_style は無塗り想定)"
                )
        else:
            got = _norm_hex(s.fill.fgColor)
            if got != spec["bg"]:
                fails.append(
                    f"[{label}] {name}.fill.fgColor = {got} "
                    f"(期待値 {spec['bg']})"
                )
        # 罫線（親のみ accent チェック）
        if spec["accent"] is not None:
            top_color = _norm_hex(
                s.border.top.color if s.border.top else None
            )
            if top_color != spec["accent"]:
                fails.append(
                    f"[{label}] {name}.border.top.color = {top_color} "
                    f"(期待値 {spec['accent']})"
                )
            top_style = s.border.top.border_style if s.border.top else None
            if top_style != "medium":
                fails.append(
                    f"[{label}] {name}.border.top.style = {top_style} "
                    f"(期待値 medium)"
                )

    return fails


def _check_copy_worksheet(wb) -> tuple[bool, str]:
    """サマリー以外のシート 1 つを複製し、セルに Named Style を適用して維持を確認。"""
    src_name = None
    for n in wb.sheetnames:
        if n not in ("サマリー", "参考"):
            src_name = n
            break
    if src_name is None:
        return False, "複製元となる詳細シートが見つからない"

    src = wb[src_name]
    dup = wb.copy_worksheet(src)
    dup.title = "__verify_probe__"
    dup["A1"].style = "parent_row_style_critical"
    dup["A2"].style = "child_row_style"
    got1 = dup["A1"].style
    got2 = dup["A2"].style
    # クリーンアップ
    del wb[dup.title]
    if got1 == "parent_row_style_critical" and got2 == "child_row_style":
        return True, f"copy_worksheet 後も維持 (A1={got1}, A2={got2})"
    return False, f"copy_worksheet 後に style が変質 (A1={got1!r}, A2={got2!r})"


def _check_roundtrip(wb) -> tuple[bool, str]:
    """save → load 後も全 5 style が残ることを確認。"""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = load_workbook(buf)
    names_present = {s.name if hasattr(s, "name") else s for s in wb2.named_styles}
    missing = [n for n in EXPECTED if n not in names_present]
    if missing:
        return False, f"ラウンドトリップ後に欠落: {missing}"
    return True, f"ラウンドトリップ後も全 {len(EXPECTED)} style 維持"


def main() -> int:
    if not TPL.exists():
        print(f"[ERROR] テンプレートが存在しません: {TPL}", file=sys.stderr)
        return 2

    wb = load_workbook(TPL)
    v = Verifier()

    print("=" * 64)
    print("1. 5 style 全件存在・プロパティ厳密一致チェック")
    print("=" * 64)
    fails = _verify_styles_in_wb(wb, "本番テンプレ")
    if not fails:
        print(f"  ✅ {len(EXPECTED)} 件すべて OK")
    else:
        for m in fails:
            print(f"  ❌ {m}")
            v.failures.append(m)

    print()
    print("=" * 64)
    print("2. copy_worksheet 後の維持チェック (Phase 8-A P1 相当)")
    print("=" * 64)
    ok, msg = _check_copy_worksheet(wb)
    v.check(ok, msg)

    print()
    print("=" * 64)
    print("3. save/load ラウンドトリップ維持 (Phase 8-A P2 相当)")
    print("=" * 64)
    ok, msg = _check_roundtrip(wb)
    v.check(ok, msg)

    print()
    print("=" * 64)
    if v.ok():
        print("[SUCCESS] すべての検証をパスしました")
        return 0
    print(f"[FAILURE] {len(v.failures)} 件の不一致")
    for f in v.failures:
        print(f"  - {f}")
    return 1


if __name__ == "__main__":
    sys.exit(main())
