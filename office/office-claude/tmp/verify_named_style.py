"""Phase 8-A 事前検証: Named Style の動的伝播可否を確認する。

3 パターンを検証:
  P1. 空 Workbook で add_named_style → copy_worksheet → スタイル付き cell に反映されるか
  P2. save → load ラウンドトリップ後も Named Style が保持されるか
  P3. テンプレートに事前登録した Named Style が copy_worksheet 先で利用可能か
"""
from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, NamedStyle, Alignment
from openpyxl.styles.borders import Border, Side


def _make_style(name: str, color: str) -> NamedStyle:
    ns = NamedStyle(name=name)
    ns.font = Font(bold=True, color="000000")
    ns.fill = PatternFill("solid", fgColor=color)
    ns.alignment = Alignment(vertical="center")
    thin = Side(border_style="thin", color="888888")
    ns.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return ns


def pattern1_copy_worksheet() -> tuple[bool, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = "src"
    style = _make_style("ParentRow", "FFE699")
    wb.add_named_style(style)
    ws["A1"] = "Parent"
    ws["A1"].style = "ParentRow"
    copied = wb.copy_worksheet(ws)
    copied.title = "dst"
    try:
        ok_src = ws["A1"].style == "ParentRow"
        ok_dst = copied["A1"].style == "ParentRow"
        return (ok_src and ok_dst), f"src.style={ws['A1'].style!r}, dst.style={copied['A1'].style!r}"
    except Exception as e:
        return False, f"exception: {e!r}"


def pattern2_roundtrip() -> tuple[bool, str]:
    wb = Workbook()
    ws = wb.active
    style = _make_style("ChildRow", "FFF2CC")
    wb.add_named_style(style)
    ws["A1"] = "Child"
    ws["A1"].style = "ChildRow"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = load_workbook(buf)
    ws2 = wb2.active
    try:
        registered = "ChildRow" in [s.name if hasattr(s, "name") else s for s in wb2.named_styles]
        applied = ws2["A1"].style == "ChildRow"
        return (registered and applied), f"registered={registered}, applied={applied}, style={ws2['A1'].style!r}"
    except Exception as e:
        return False, f"exception: {e!r}"


def pattern3_template_preregistered() -> tuple[bool, str]:
    tpl = Path("data/reports/template/TC_template.xlsx")
    if not tpl.exists():
        return False, f"template not found at {tpl}"
    wb = load_workbook(tpl)
    names = [s.name if hasattr(s, "name") else s for s in wb.named_styles]
    # 現在のテンプレートに ParentRow/ChildRow がないことを確認
    has_parent = "ParentRow" in names
    has_child = "ChildRow" in names
    # そこに後付けで追加できるかも確認
    added_ok = True
    added_msg = ""
    try:
        if not has_parent:
            wb.add_named_style(_make_style("ParentRow", "FFE699"))
        if not has_child:
            wb.add_named_style(_make_style("ChildRow", "FFF2CC"))
        # copy_worksheet 後の適用
        src = wb["detail"] if "detail" in wb.sheetnames else wb.active
        dup = wb.copy_worksheet(src)
        dup.title = "dup_probe"
        dup["A1"].style = "ParentRow"
        dup["A2"].style = "ChildRow"
        added_msg = f"apply dup[A1]={dup['A1'].style!r} dup[A2]={dup['A2'].style!r}"
    except Exception as e:
        added_ok = False
        added_msg = f"exception: {e!r}"
    msg = (
        f"template named_styles={names} | "
        f"has_parent={has_parent}, has_child={has_child} | "
        f"post-add {added_msg}"
    )
    return added_ok, msg


def main() -> None:
    results = {
        "P1 copy_worksheet":          pattern1_copy_worksheet(),
        "P2 save/load roundtrip":     pattern2_roundtrip(),
        "P3 template pre-registered": pattern3_template_preregistered(),
    }
    print("=" * 60)
    print("Phase 8-A Named Style 事前検証")
    print("=" * 60)
    for label, (ok, msg) in results.items():
        mark = "OK" if ok else "NG"
        print(f"[{mark}] {label}")
        print(f"     {msg}")
    print("=" * 60)
    all_ok = all(ok for ok, _ in results.values())
    print(f"総合: {'ALL OK' if all_ok else 'FAIL あり'}")


if __name__ == "__main__":
    main()
