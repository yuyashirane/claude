"""Step 3-C 検証③: 3 社の Excel から検証観点を一括抽出。

抽出項目:
    - 全シート名・親行/子行行数
    - 親行 GL ハイパーリンク URL → start_date/end_date/tax_group_codes
    - 親行 C 列(summary)・E 列(check_result) 文言
    - 子行 J/K/L/M/N 列(取引先・品目・部門・メモ・摘要)
    - 子行 H 列(取引日)昇順チェック
    - 子行 GL/JR ハイパーリンクの存在
"""
from __future__ import annotations
import sys
import urllib.parse
from pathlib import Path
import openpyxl


def parse_gl_url(url):
    if not url:
        return {}
    return urllib.parse.parse_qs(urllib.parse.urlparse(url).query)


def inspect(xlsx: Path) -> None:
    print(f"\n{'='*80}\nFILE: {xlsx.name}\n{'='*80}")
    wb = openpyxl.load_workbook(xlsx, data_only=False)  # hyperlink を読むため data_only=False
    print(f"SHEETS: {wb.sheetnames}\n")

    for sname in wb.sheetnames:
        ws = wb[sname]
        # 詳細シート判定: 行 4 以降にデータあり、列 W (23) まで存在
        if ws.max_row < 4:
            continue
        # parent_row のスタイル名で親行判定
        parent_rows = []
        child_rows = []
        for r in range(4, ws.max_row + 1):
            try:
                style = ws.cell(r, 1).style
            except Exception:
                style = None
            if style and "parent_row_style" in str(style):
                parent_rows.append(r)
            elif style == "child_row_style":
                child_rows.append(r)
        if not parent_rows:
            continue

        print(f"\n--- {sname} (parents={len(parent_rows)} children={len(child_rows)}) ---")

        # 親行サマリ
        for pr in parent_rows[:8]:  # 最大 8 親行
            sub_code = ws.cell(pr, 2).value
            c_summary = ws.cell(pr, 3).value
            e_check = ws.cell(pr, 5).value
            link_cell = ws.cell(pr, 17)
            url = link_cell.hyperlink.target if link_cell.hyperlink else None
            q = parse_gl_url(url)
            sd = q.get("start_date", [None])[0]
            ed = q.get("end_date", [None])[0]
            tgc = q.get("tax_group_codes", [])
            print(f"  P r{pr} {sub_code}")
            print(f"    C: {c_summary}")
            print(f"    E: {e_check}")
            print(f"    GL: start={sd} end={ed} tax_group_codes={tgc}")

        # 子行サンプル(最初の 5 件)+ 取引日昇順チェック
        if child_rows:
            print(f"  Child sample (first 5 of {len(child_rows)}):")
            for cr in child_rows[:5]:
                date_h = ws.cell(cr, 8).value
                acct = ws.cell(cr, 9).value
                j = ws.cell(cr, 10).value  # 取引先
                k = ws.cell(cr, 11).value  # 品目
                l = ws.cell(cr, 12).value  # 部門
                m = ws.cell(cr, 13).value  # メモ
                n = ws.cell(cr, 14).value  # 摘要
                gl = ws.cell(cr, 17).hyperlink.target if ws.cell(cr, 17).hyperlink else None
                jr = ws.cell(cr, 18).hyperlink.target if ws.cell(cr, 18).hyperlink else None
                print(f"    r{cr} date={date_h} acct={acct}")
                print(f"      J={j!r} K={k!r} L={l!r} M={m!r} N={n!r}")
                print(f"      GL={'YES' if gl else 'NO'} JR={'YES' if jr else 'NO'}")

            # 取引日昇順チェック (per group: 親と次の親の間の子行で)
            for i, pr in enumerate(parent_rows):
                next_pr = parent_rows[i + 1] if i + 1 < len(parent_rows) else ws.max_row + 1
                group_children = [r for r in child_rows if pr < r < next_pr]
                dates = []
                for r in group_children:
                    v = ws.cell(r, 8).value
                    if v:
                        dates.append((r, str(v)))
                if dates:
                    sorted_dates = sorted(dates, key=lambda x: x[1])
                    is_sorted = [d[1] for d in dates] == [d[1] for d in sorted_dates]
                    if not is_sorted:
                        print(f"    [SORT] group@P{pr}: NOT ASCENDING -> {[d[1] for d in dates]}")
                    elif len(dates) > 1:
                        print(f"    [SORT] group@P{pr}: ascending OK ({len(dates)} children)")


def main():
    files = [Path(p) for p in sys.argv[1:]]
    for f in files:
        if not f.exists():
            print(f"NOT FOUND: {f}")
            continue
        inspect(f)


if __name__ == "__main__":
    main()
