"""Phase 8-B E2E 検証スクリプト（仕様案 Z + Named Style + 親子行構造）。

Phase 3-5 のゴール:
    antread 2025/12 の実データで export_to_excel を回し、
    生成された e2e_report_phase8b_*.xlsx が以下 13 項目を満たすかを判定する。

検証項目（要件ドキュメント v2 §7 + 戦略 Claude Step 3 GO 指示 準拠）:
    1. テンプレ本体の 5 NamedStyle 存在（verify_phase8b_template.py 相当）
    2. 生成ファイルにも 5 NamedStyle が保持されていること
    3. 少なくとも 1 枚の詳細シートで親行が parent_row_style_* を使っている
    4. 同シートの子行が child_row_style を使っている
    5. 親行 C 列が C-β-3 形式（"— X 件・合計 ¥Y"）を含む
    6. 親行 D/E 列が空文字ではない（観点・チェック結果が入る）
    7. Pattern B（TC-06）が存在すれば "税区分混在" 文言を含む親行が 1 件以上ある
    8. 子行の Q/R 列ハイパーリンクが維持されている
    9. 子行の M 列 severity が空白である（仕様案 Z: 子行は severity 列出力なし）
    10. 子行 fill_type が None/"none" である（仕様案 Z: 白地維持）
    11. サマリーシート B3/B4 に会社名・対象月が入っている
    12. 参考シートが存在する
    13. save/load ラウンドトリップで全 5 style が維持される

Usage:
    cd office/office-claude
    PYTHONIOENCODING=utf-8 py tmp/verify_phase8b.py
"""
from __future__ import annotations

import importlib.util
import io
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, ".")

from openpyxl import load_workbook

from scripts.e2e.freee_to_context import build_check_context
from skills.export.excel_report.exporter import export_to_excel


TPL = Path("data/reports/template/TC_template.xlsx")
BASE = Path("data/e2e/3525430/202512")

EXPECTED_STYLES = {
    "parent_row_style_critical": "FCEBEB",
    "parent_row_style_warning":  "FAEEDA",
    "parent_row_style_medium":   "FEF5D6",
    "parent_row_style_low":      "EAF3DE",
    "child_row_style":           None,
}


class Verifier:
    def __init__(self) -> None:
        self.passes: list[str] = []
        self.failures: list[str] = []

    def check(self, cond: bool, msg: str) -> None:
        if cond:
            print(f"  ✅ {msg}")
            self.passes.append(msg)
        else:
            print(f"  ❌ {msg}")
            self.failures.append(msg)

    def ok(self) -> bool:
        return not self.failures


def _norm_hex(rgb) -> str | None:
    if rgb is None:
        return None
    try:
        s = rgb.rgb if hasattr(rgb, "rgb") else rgb
    except Exception:
        return None
    if not isinstance(s, str):
        return None
    s = s.upper()
    if len(s) == 8:
        s = s[2:]
    if len(s) != 6:
        return None
    return s


def _all_styles_present(wb) -> tuple[bool, list[str]]:
    names = {s.name if hasattr(s, "name") else s for s in wb.named_styles}
    missing = [n for n in EXPECTED_STYLES if n not in names]
    return (not missing, missing)


def _run_e2e() -> Path:
    """antread 2025/12 の実データで Excel 出力を生成する。"""
    skill_dir = Path("skills/verify/V1-3-rule/check-tax-classification")
    sys.path.insert(0, str(skill_dir))
    spec = importlib.util.spec_from_file_location("checker", skill_dir / "checker.py")
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)

    ctx = build_check_context(
        deals_path=BASE / "deals_202512.json",
        partners_path=BASE / "partners_all.json",
        account_items_path=BASE / "account_items_all.json",
        company_info_path=BASE / "company_info.json",
        taxes_codes_path=BASE / "taxes_codes.json",
    )
    findings = m.run(ctx)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = BASE / f"e2e_report_phase8b_{ts}.xlsx"
    export_to_excel(findings, output, company_name=ctx.company_name, period=BASE.name)
    print(f"  → Generated: {output.name} ({output.stat().st_size:,} bytes, "
          f"{len(findings)} findings)")
    return output


def _find_parent_child_rows(ws):
    """Find first parent/child pair in a detail sheet.

    Returns (parent_row, child_row) or (None, None) if not found.
    Heuristic: parent row style name starts with 'parent_row_style_',
    child row style equals 'child_row_style'.
    """
    parent = None
    child = None
    for row in range(4, ws.max_row + 1):
        style_name = ws.cell(row=row, column=3).style  # C 列
        if parent is None and style_name.startswith("parent_row_style_"):
            parent = row
            continue
        if parent is not None and style_name == "child_row_style":
            child = row
            break
    return parent, child


def main() -> int:
    v = Verifier()

    # ─── 項目 1: テンプレ本体の style 存在 ───
    print("=" * 64)
    print("[1] テンプレ本体の 5 NamedStyle 存在チェック")
    print("=" * 64)
    if not TPL.exists():
        print(f"[ERROR] テンプレートが存在しません: {TPL}", file=sys.stderr)
        return 2
    wb_tpl = load_workbook(TPL)
    ok, missing = _all_styles_present(wb_tpl)
    v.check(ok, f"テンプレに 5 style 全件登録 (missing={missing if missing else 'なし'})")

    # ─── E2E 実行 ───
    print()
    print("=" * 64)
    print("E2E 実行（antread 2025/12）")
    print("=" * 64)
    output = _run_e2e()
    wb = load_workbook(output)

    # ─── 項目 2: 生成ファイルの style 存在 ───
    print()
    print("=" * 64)
    print("[2] 生成ファイルに 5 NamedStyle 保持")
    print("=" * 64)
    ok, missing = _all_styles_present(wb)
    v.check(ok, f"生成ファイルに 5 style 保持 (missing={missing if missing else 'なし'})")

    # ─── 詳細シートを走査して検証項目 3〜10 を判定 ───
    detail_sheets = [
        n for n in wb.sheetnames
        if n not in ("サマリー", "参考") and n.startswith("A")
    ]
    print()
    print(f"詳細シート {len(detail_sheets)} 枚を走査: {detail_sheets}")

    found_parent_style = False
    found_child_style = False
    found_summary_form = False   # C-β-3 "— X 件・合計 ¥Y"
    found_observation = False    # D 列非空
    found_check_result = False   # E 列非空
    found_mixing_tag = False     # Pattern B "税区分混在"
    found_hyperlink = False      # 子行 Q/R
    found_m_blank_child = False  # 子行 M 列空白
    found_child_no_fill = False  # 子行 fill_type None
    has_tc06 = False

    for sheet_name in detail_sheets:
        ws = wb[sheet_name]
        for row in range(4, ws.max_row + 1):
            style_name = ws.cell(row=row, column=3).style
            c_val = ws.cell(row=row, column=3).value or ""
            d_val = ws.cell(row=row, column=4).value or ""
            e_val = ws.cell(row=row, column=5).value or ""
            m_val = ws.cell(row=row, column=13).value
            q_cell = ws.cell(row=row, column=17)
            r_cell = ws.cell(row=row, column=18)
            tc_val = ws.cell(row=row, column=2).value or ""

            if style_name.startswith("parent_row_style_"):
                found_parent_style = True
                if "件" in c_val and "合計" in c_val:
                    found_summary_form = True
                if d_val.strip():
                    found_observation = True
                if e_val.strip():
                    found_check_result = True
                if "TC-06" in tc_val:
                    has_tc06 = True
                    if "税区分混在" in c_val or "税区分混在" in d_val or "税区分混在" in e_val:
                        found_mixing_tag = True

            elif style_name == "child_row_style":
                found_child_style = True
                if q_cell.hyperlink or r_cell.hyperlink:
                    found_hyperlink = True
                if m_val is None or m_val == "":
                    found_m_blank_child = True
                fill_type = ws.cell(row=row, column=3).fill.fill_type
                if fill_type in (None, "none"):
                    found_child_no_fill = True

    print()
    print("=" * 64)
    print("[3〜10] 詳細シート走査結果")
    print("=" * 64)
    v.check(found_parent_style, "親行で parent_row_style_* が使われている")
    v.check(found_child_style,  "子行で child_row_style が使われている")
    v.check(found_summary_form, "親行 C 列に C-β-3 形式（件・合計 ¥）")
    v.check(found_observation,  "親行 D 列が非空（観点）")
    v.check(found_check_result, "親行 E 列が非空（チェック結果）")
    if has_tc06:
        v.check(found_mixing_tag, "Pattern B (TC-06) に '税区分混在' 文言")
    else:
        print("  ⏭  TC-06 が検出 Finding になかったため Pattern B チェックはスキップ")
    v.check(found_hyperlink,    "子行 Q/R 列ハイパーリンク保持")
    v.check(found_m_blank_child, "子行 M 列（severity）が空白（仕様案 Z）")
    v.check(found_child_no_fill, "子行 fill_type = None/none（仕様案 Z 白地）")

    # ─── 項目 11: サマリー ───
    print()
    print("=" * 64)
    print("[11] サマリーシート会社名・対象月")
    print("=" * 64)
    ws_sum = wb["サマリー"]
    b3 = ws_sum.cell(3, 2).value
    b4 = ws_sum.cell(4, 2).value
    v.check(bool(b3), f"サマリー B3 会社名 = {b3!r}")
    v.check(bool(b4), f"サマリー B4 対象月 = {b4!r}")

    # ─── 項目 12: 参考シート ───
    print()
    print("=" * 64)
    print("[12] 参考シート存在")
    print("=" * 64)
    v.check("参考" in wb.sheetnames, f"'参考' シート存在 (sheets={wb.sheetnames})")

    # ─── 項目 13: ラウンドトリップ ───
    print()
    print("=" * 64)
    print("[13] save/load ラウンドトリップで 5 style 維持")
    print("=" * 64)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = load_workbook(buf)
    ok, missing = _all_styles_present(wb2)
    v.check(ok, f"ラウンドトリップ後も 5 style 維持 (missing={missing if missing else 'なし'})")

    # ─── サマリー ───
    print()
    print("=" * 64)
    total = len(v.passes) + len(v.failures)
    if v.ok():
        print(f"[SUCCESS] Phase 8-B E2E 検証 {total}/{total} 項目 PASS")
        print(f"生成ファイル: {output.resolve()}")
        return 0
    print(f"[FAILURE] {len(v.failures)}/{total} 項目が不一致")
    for m in v.failures:
        print(f"  - {m}")
    return 1


if __name__ == "__main__":
    sys.exit(main())
