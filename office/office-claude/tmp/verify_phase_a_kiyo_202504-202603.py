"""Phase A E2E 検証スクリプト (株式会社輝陽 2025/04-2026/03)。

verify_phase8c.py を参考に、輝陽の会計期間 (2025-04-01〜2026-03-31) で
V1-3-10 checker を実行し、Phase 8-C の Excel exporter でレポートを出力する。

Usage:
    cd office/office-claude
    PYTHONIOENCODING=utf-8 py tmp/verify_phase_a_kiyo_202504-202603.py
"""
from __future__ import annotations

import importlib.util
import io
import sys
import urllib.parse
from datetime import datetime
from pathlib import Path

sys.path.insert(0, ".")

from openpyxl import load_workbook

from scripts.e2e.freee_to_context import build_check_context
from skills.export.excel_report.exporter import export_to_excel


TPL = Path("data/reports/template/TC_template.xlsx")
BASE = Path("data/e2e/1062190_株式会社輝陽/202504-202603")
COMPANY_SHORT = "kiyo"
PERIOD_DISPLAY = "2025年4月〜2026年3月"

EXPECTED_STYLES = {
    "parent_row_style_critical": "FCEBEB",
    "parent_row_style_warning":  "FAEEDA",
    "parent_row_style_medium":   "FEF5D6",
    "parent_row_style_low":      "EAF3DE",
    "child_row_style":           None,
}


class Verifier:
    def __init__(self) -> None:
        self.passes: list[str] = []
        self.failures: list[str] = []

    def check(self, cond: bool, msg: str) -> None:
        if cond:
            print(f"  ✅ {msg}")
            self.passes.append(msg)
        else:
            print(f"  ❌ {msg}")
            self.failures.append(msg)

    def ok(self) -> bool:
        return not self.failures


def _all_styles_present(wb) -> tuple[bool, list[str]]:
    names = {s.name if hasattr(s, "name") else s for s in wb.named_styles}
    missing = [n for n in EXPECTED_STYLES if n not in names]
    return (not missing, missing)


def _run_e2e() -> tuple[Path, object, int]:
    skill_dir = Path("skills/verify/V1-3-rule/check-tax-classification")
    sys.path.insert(0, str(skill_dir))
    spec = importlib.util.spec_from_file_location("checker", skill_dir / "checker.py")
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)

    ctx = build_check_context(
        deals_path=BASE / "deals_202504-202603.json",
        partners_path=BASE / "partners_all.json",
        account_items_path=BASE / "account_items_all.json",
        company_info_path=BASE / "company_info.json",
        taxes_codes_path=BASE / "taxes_codes.json",
    )
    findings = m.run(ctx)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = BASE / f"e2e_report_phase_a_{COMPANY_SHORT}_{ts}.xlsx"
    export_to_excel(
        findings, output,
        company_name=ctx.company_name,
        period=PERIOD_DISPLAY,
        ctx=ctx,
    )
    print(f"  → Generated: {output.name} ({output.stat().st_size:,} bytes, "
          f"{len(findings)} findings)")
    print(f"  → ctx.period_start = {ctx.period_start}, period_end = {ctx.period_end}")
    return output, ctx, len(findings)


def main() -> int:
    v = Verifier()

    print("=" * 64)
    print("[1] テンプレ本体の 5 NamedStyle 存在チェック")
    print("=" * 64)
    if not TPL.exists():
        print(f"[ERROR] テンプレートが存在しません: {TPL}", file=sys.stderr)
        return 2
    wb_tpl = load_workbook(TPL)
    ok, missing = _all_styles_present(wb_tpl)
    v.check(ok, f"テンプレに 5 style 全件登録 (missing={missing if missing else 'なし'})")

    print()
    print("=" * 64)
    print("E2E 実行（株式会社輝陽 2025/04-2026/03、Phase 8-C: ctx 連携）")
    print("=" * 64)
    output, ctx, n_findings = _run_e2e()
    wb = load_workbook(output)

    print()
    print("=" * 64)
    print("[2] 生成ファイルに 5 NamedStyle 保持")
    print("=" * 64)
    ok, missing = _all_styles_present(wb)
    v.check(ok, f"生成ファイルに 5 style 保持 (missing={missing if missing else 'なし'})")

    detail_sheets = [
        n for n in wb.sheetnames
        if n not in ("サマリー", "参考") and n.startswith("A")
    ]
    print()
    print(f"詳細シート {len(detail_sheets)} 枚を走査: {detail_sheets}")

    found_parent_style = False
    found_child_style = False
    found_summary_form = False
    found_observation = False
    found_check_result = False
    found_mixing_tag = False
    found_hyperlink = False
    found_m_blank_child = False
    found_child_no_fill = False
    has_tc06 = False
    found_parent_gl_link = False
    found_parent_gl_monthly = False
    found_parent_gl_tax_codes = False
    found_child_e_nonempty = False
    found_child_d_empty = True

    for sheet_name in detail_sheets:
        ws = wb[sheet_name]
        for row in range(4, ws.max_row + 1):
            style_name = ws.cell(row=row, column=3).style
            c_val = ws.cell(row=row, column=3).value or ""
            d_val = ws.cell(row=row, column=4).value or ""
            e_val = ws.cell(row=row, column=5).value or ""
            m_val = ws.cell(row=row, column=13).value
            q_cell = ws.cell(row=row, column=17)
            r_cell = ws.cell(row=row, column=18)
            tc_val = ws.cell(row=row, column=2).value or ""

            if style_name.startswith("parent_row_style_"):
                found_parent_style = True
                if "件" in c_val and "合計" in c_val:
                    found_summary_form = True
                if isinstance(d_val, str) and d_val.strip():
                    found_observation = True
                if isinstance(e_val, str) and e_val.strip():
                    found_check_result = True
                if "TC-06" in tc_val:
                    has_tc06 = True
                    if ("税区分混在" in str(c_val)
                            or "税区分混在" in str(d_val)
                            or "税区分混在" in str(e_val)):
                        found_mixing_tag = True

                if q_cell.value == "GL" and q_cell.hyperlink is not None:
                    found_parent_gl_link = True
                    url = q_cell.hyperlink.target or ""
                    parsed = urllib.parse.urlparse(url)
                    qs = urllib.parse.parse_qs(parsed.query)
                    sd = qs.get("start_date", [""])[0]
                    ed = qs.get("end_date", [""])[0]
                    if len(sd) == 10 and len(ed) == 10 and sd[:7] == ed[:7] and sd[8:10] == "01":
                        if ed[8:10] in ("28", "29", "30", "31"):
                            found_parent_gl_monthly = True
                    if qs.get("tax_group_codes"):
                        found_parent_gl_tax_codes = True

            elif style_name == "child_row_style":
                found_child_style = True
                if q_cell.hyperlink or r_cell.hyperlink:
                    found_hyperlink = True
                if m_val is None or m_val == "":
                    found_m_blank_child = True
                fill_type = ws.cell(row=row, column=3).fill.fill_type
                if fill_type in (None, "none"):
                    found_child_no_fill = True
                if isinstance(e_val, str) and e_val.strip():
                    found_child_e_nonempty = True
                if isinstance(d_val, str) and d_val.strip():
                    found_child_d_empty = False

    print()
    print("=" * 64)
    print("[3〜10] 詳細シート走査結果（Phase 8-B 継承）")
    print("=" * 64)
    v.check(found_parent_style, "親行で parent_row_style_* が使われている")
    v.check(found_child_style,  "子行で child_row_style が使われている")
    v.check(found_summary_form, "親行 C 列に C-β-3 形式（件・合計 ¥）")
    v.check(found_observation,  "親行 D 列が非空（観点）")
    v.check(found_check_result, "親行 E 列が非空（チェック結果）")
    if has_tc06:
        v.check(found_mixing_tag, "Pattern B (TC-06) に '税区分混在' 文言")
    else:
        print("  ⏭  TC-06 が検出 Finding になかったため Pattern B チェックはスキップ")
    v.check(found_hyperlink,    "子行 Q/R 列ハイパーリンク保持")
    v.check(found_m_blank_child, "子行 M 列（severity）が空白")
    v.check(found_child_no_fill, "子行 fill_type = None/none")

    print()
    print("=" * 64)
    print("[11] サマリーシート会社名・対象月")
    print("=" * 64)
    ws_sum = wb["サマリー"]
    b3 = ws_sum.cell(3, 2).value
    b4 = ws_sum.cell(4, 2).value
    v.check(bool(b3), f"サマリー B3 会社名 = {b3!r}")
    v.check(bool(b4), f"サマリー B4 対象月 = {b4!r}")

    print()
    print("=" * 64)
    print("[12] 参考シート存在")
    print("=" * 64)
    v.check("参考" in wb.sheetnames, f"'参考' シート存在 (sheets={wb.sheetnames})")

    print()
    print("=" * 64)
    print("[13] save/load ラウンドトリップで 5 style 維持")
    print("=" * 64)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = load_workbook(buf)
    ok, missing = _all_styles_present(wb2)
    v.check(ok, f"ラウンドトリップ後も 5 style 維持 (missing={missing if missing else 'なし'})")

    print()
    print("=" * 64)
    print("[14] 親行 Q 列に GL ハイパーリンク")
    print("=" * 64)
    v.check(found_parent_gl_link, "親行 Q 列 = 'GL' + ハイパーリンクあり")

    print()
    print("=" * 64)
    print("[14.1] 親行 GL URL の期間が単月")
    print("=" * 64)
    v.check(found_parent_gl_monthly,
            "親行 GL URL の start_date/end_date が同一月かつ月初〜月末形式")

    print()
    print("=" * 64)
    print("[14.2] 親行 GL URL に tax_group_codes が付与")
    print("=" * 64)
    v.check(found_parent_gl_tax_codes,
            "親行 GL URL に tax_group_codes=... が 1 つ以上含まれる")

    print()
    print("=" * 64)
    print("[15] 子行 E 列 message 由来、D 列は常に空欄")
    print("=" * 64)
    v.check(found_child_e_nonempty, "子行 E 列が非空（Finding.message 由来）")
    v.check(found_child_d_empty, "子行 D 列は全行空欄")

    print()
    print("=" * 64)
    print("[16] サマリー B4 対象月が format_target_month 出力形式")
    print("=" * 64)
    b4_str = str(b4 or "")
    is_monthly_form = ("年" in b4_str and "月" in b4_str)
    v.check(is_monthly_form,
            f"サマリー B4 対象月表示: {b4_str!r}")

    print()
    print("=" * 64)
    total = len(v.passes) + len(v.failures)
    if v.ok():
        print(f"[SUCCESS] E2E 検証 {total}/{total} 項目 PASS")
        print(f"Findings: {n_findings}")
        print(f"生成ファイル: {output.resolve()}")
        return 0
    print(f"[FAILURE] {len(v.failures)}/{total} 項目が不一致")
    for msg in v.failures:
        print(f"  - {msg}")
    print(f"Findings: {n_findings}")
    print(f"生成ファイル: {output.resolve()}")
    return 1


if __name__ == "__main__":
    sys.exit(main())
