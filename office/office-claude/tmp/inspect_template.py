"""Phase 8-B 事前調査: TC_template.xlsx の現状構造を洗い出す。

シート一覧 / 行高 / Row 4 の既存スタイル / Named Style 登録状況を収集し、
docs/phase8b_template_requirements.md の下地にする。
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


TPL = Path("data/reports/template/TC_template.xlsx")


def main() -> None:
    wb = load_workbook(TPL)
    print("=" * 70)
    print("Template:", TPL)
    print("=" * 70)
    print("sheetnames:", wb.sheetnames)

    print()
    print("Named Styles 登録:", [s.name if hasattr(s, "name") else s for s in wb.named_styles])

    # 代表的な詳細シートを 1 枚選んで Row 4 の様子を確認
    target = None
    for name in wb.sheetnames:
        if name.startswith("A5"):
            target = name
            break
    if target is None:
        for name in wb.sheetnames:
            if name not in ("サマリー", "参考"):
                target = name
                break
    print()
    print("Inspect target sheet:", target)
    ws = wb[target]
    print("max_row:", ws.max_row, "max_col:", ws.max_col if hasattr(ws, "max_col") else ws.max_column)

    # Row 高
    print()
    print("Row heights (first 10):")
    for r in range(1, 11):
        h = ws.row_dimensions[r].height
        print(f"  row {r}: height={h}")

    # Row 3/4 の各セルの font / fill / border / alignment
    print()
    for r in (3, 4):
        print(f"Row {r} cell styles:")
        for col in range(1, 8):
            cell = ws.cell(r, col)
            font = cell.font
            fill = cell.fill
            al = cell.alignment
            print(
                f"  col {col}: val={cell.value!r} "
                f"font(name={font.name!r}, size={font.size}, bold={font.bold}, color={font.color.rgb if font.color else None}) "
                f"fill(type={fill.fill_type}, fg={fill.fgColor.rgb if fill.fgColor else None}) "
                f"align(h={al.horizontal}, v={al.vertical}, indent={al.indent})"
            )

    # Column widths
    print()
    print("Column widths (A-W):")
    for col_letter in "ABCDEFGHIJKLMNOPQRSTUVW":
        d = ws.column_dimensions.get(col_letter)
        if d is not None:
            print(f"  {col_letter}: width={d.width}")


if __name__ == "__main__":
    main()
