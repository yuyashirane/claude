"""Phase 8-B プレビュー Named Style サンプル生成 (v3 = 仕様案 Z)。

要件ドキュメント v2 §3.1-3.5 の 5 style を定義し、
4 severity ごとに親行 + 子行 2 本のサンプル 4 セットを配置する。

v3 差分（仕様案 Z 確定）:
    子行の severity fill 重ね塗りを廃止。子行は child_row_style（白地）のまま。
    個別 severity は子行 A 列の絵文字（🔴/🟠/🟡/🟢）でのみ表現する。
    親行の 4 色だけが severity を視覚的に伝える帯として機能する。

    v2 で行っていた `cell.fill = copy(sev_fill)` を子行に対して一切実行しない。
    それ以外（絵文字・文言・リンク列・インデント・number_format）は v2 と同一。

親行は A〜W 列すべてに Named Style を適用（要件 §4「1 行の帯」ルール）。
子行は C 列のみ Alignment(indent=2) を Python 側で上書き（方式 α）。

出力: tmp/preview_named_styles_sample.xlsx
本番テンプレ (TC_template.xlsx) には触れない。
"""
from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.styles.borders import Border, Side


# ─────────────────────────────────────────────────────────────
# Named Style 色定義（要件 v2 §3）
# ─────────────────────────────────────────────────────────────
CRITICAL_BG     = "FCEBEB"  # 淡い赤
CRITICAL_BORDER = "C00000"
WARNING_BG      = "FAEEDA"  # 淡いオレンジ
WARNING_BORDER  = "ED7D31"
MEDIUM_BG       = "FEF5D6"  # 淡い黄
MEDIUM_BORDER   = "BF8F00"
LOW_BG          = "EAF3DE"  # 淡い緑
LOW_BORDER      = "548235"

PARENT_STYLES = [
    ("parent_row_style_critical", CRITICAL_BG, CRITICAL_BORDER),
    ("parent_row_style_warning",  WARNING_BG,  WARNING_BORDER),
    ("parent_row_style_medium",   MEDIUM_BG,   MEDIUM_BORDER),
    ("parent_row_style_low",      LOW_BG,      LOW_BORDER),
]

# ─────────────────────────────────────────────────────────────
# v3（仕様案 Z）: 子行は白地維持。severity fill は親行のみ。
# よって SEVERITY_FILL_HEX 定義は不要（v2 から削除）。
# ─────────────────────────────────────────────────────────────

MAX_COL = 23  # A〜W
HEADER_ROW = 3
DATA_START = 4
ROW_HEIGHT = 42.75  # 要件 v2 §3.6（Q3 確定）


def _side(style: str, color: str) -> Side:
    return Side(border_style=style, color=f"FF{color}")


def make_parent_style(name: str, bg_hex: str, accent_hex: str) -> NamedStyle:
    ns = NamedStyle(name=name)
    ns.font = Font(name="Meiryo UI", size=10, bold=True, color="FF000000")
    ns.fill = PatternFill("solid", fgColor=f"FF{bg_hex}")
    ns.alignment = Alignment(vertical="center", horizontal="left")
    thin_gray = _side("thin", "D9D9D9")
    ns.border = Border(
        left=thin_gray,
        right=thin_gray,
        top=_side("medium", accent_hex),
        bottom=_side("thin", accent_hex),
    )
    return ns


def make_child_style() -> NamedStyle:
    ns = NamedStyle(name="child_row_style")
    ns.font = Font(name="Meiryo UI", size=10, bold=False, color="FF000000")
    # 背景色は設定しない（severity_fills 重ね塗りを壊さない）
    ns.alignment = Alignment(vertical="center", horizontal="left")
    thin = _side("thin", "D9D9D9")
    ns.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return ns


# ─────────────────────────────────────────────────────────────
# ヘッダー定義（本番テンプレの 23 列と同じ順序）
# ─────────────────────────────────────────────────────────────
HEADERS = [
    "優先度", "項目", "項目名", "観点", "チェック結果",
    "現在の税区分", "推奨税区分", "取引日", "勘定科目", "取引先",
    "品目", "部門", "メモ", "摘要", "借方金額",
    "貸方金額", "🔗総勘定元帳", "🔗仕訳帳", "確信度", "エラー型",
    "確認状況", "担当者メモ", "walletTxnId",
]

# 列幅（本番テンプレから抜粋・一部既定）
COL_WIDTHS = {
    1: 8.0, 2: 10.0, 3: 16.875, 4: 25.0, 5: 40.625, 6: 10.625,
    7: 10.625, 8: 12.0, 9: 15.0, 10: 10.625, 11: 12.0, 12: 10.625,
    13: 12.0, 14: 20.625, 15: 12.0, 16: 12.0, 17: 10.0, 18: 10.0,
    19: 8.0, 20: 15.0, 21: 15.0, 22: 30.0, 23: 20.0,
}


# ─────────────────────────────────────────────────────────────
# サンプルデータ（4 severity × 1 親行 + 2 子行 = 12 データ行）
# ─────────────────────────────────────────────────────────────
SAMPLE_SETS = [
    {
        "severity_label": "重大",
        "style": "parent_row_style_critical",
        "parent": {
            3: "給与手当 — 2 件・合計 ¥600,000",
            4: "（親行・集約）",
            5: "給与系勘定に課税仕入が混入（TC-03a）",
            9: "給与手当",
            15: 600000,
        },
        "children": [
            {1: "重大", 2: "TC-03a", 3: "給与/人件費",
             4: "給与系に課税区分", 5: "給与は対象外。課税仕入は誤り。",
             6: "課対仕入10%", 7: "対象外", 8: "2026/01/25",
             9: "給与手当", 10: "", 15: 300000, 19: 90, 20: "direct_error"},
            {1: "重大", 2: "TC-03a", 3: "給与/人件費",
             4: "同上", 5: "同上",
             6: "課対仕入10%", 7: "対象外", 8: "2026/02/25",
             9: "給与手当", 10: "", 15: 300000, 19: 90, 20: "direct_error"},
        ],
    },
    {
        "severity_label": "要注意",
        "style": "parent_row_style_warning",
        "parent": {
            3: "支払手数料 — 2 件・合計 ¥44,000",
            4: "（親行・混在）",
            5: "同一科目に複数税区分が混在（TC-06a）",
            9: "支払手数料",
            15: 44000,
        },
        "children": [
            {1: "要注意", 2: "TC-06a", 3: "税金/租税公課",
             4: "課税区分が混在", 5: "租税公課は対象外のはず。",
             6: "課対仕入10%", 7: "対象外", 8: "2026/01/10",
             9: "支払手数料", 15: 22000, 19: 75, 20: "gray_review"},
            {1: "要注意", 2: "TC-06a", 3: "税金/租税公課",
             4: "同上", 5: "同上",
             6: "非課仕入", 7: "対象外", 8: "2026/02/10",
             9: "支払手数料", 15: 22000, 19: 75, 20: "gray_review"},
        ],
    },
    {
        "severity_label": "要確認",
        "style": "parent_row_style_medium",
        "parent": {
            3: "消耗品費 — 2 件・合計 ¥15,400",
            4: "（親行・判断）",
            5: "軽油関連キーワード検出（TC-06d）",
            9: "消耗品費",
            15: 15400,
        },
        "children": [
            {1: "要確認", 2: "TC-06d", 3: "税金/租税公課",
             4: "軽油引取税の判定要確認", 5: "軽油分の分離が必要か確認。",
             6: "課対仕入10%", 7: "要判断", 8: "2026/03/05",
             9: "消耗品費", 14: "軽油代", 15: 7700, 19: 70, 20: "gray_review"},
            {1: "要確認", 2: "TC-06d", 3: "税金/租税公課",
             4: "同上", 5: "同上",
             6: "課対仕入10%", 7: "要判断", 8: "2026/03/18",
             9: "消耗品費", 14: "軽油代", 15: 7700, 19: 70, 20: "gray_review"},
        ],
    },
    {
        "severity_label": "参考",
        "style": "parent_row_style_low",
        "parent": {
            3: "印紙税 — 2 件・合計 ¥1,200",
            4: "（親行・許容）",
            5: "非課仕入も許容だが対象外が推奨（TC-06c）",
            9: "印紙税",
            15: 1200,
        },
        "children": [
            {1: "参考", 2: "TC-06c", 3: "税金/租税公課",
             4: "非課仕入=許容", 5: "より正確には対象外が適切。",
             6: "非課仕入", 7: "対象外", 8: "2026/01/15",
             9: "印紙税", 15: 600, 19: 60, 20: "mild_warning"},
            {1: "参考", 2: "TC-06c", 3: "税金/租税公課",
             4: "同上", 5: "同上",
             6: "非課仕入", 7: "対象外", 8: "2026/02/15",
             9: "印紙税", 15: 600, 19: 60, 20: "mild_warning"},
        ],
    },
]


# ─────────────────────────────────────────────────────────────
# 書き込みヘルパー
# ─────────────────────────────────────────────────────────────

def _write_header(ws) -> None:
    # Row 1: タイトル
    ws.cell(1, 1).value = "Phase 8-B プレビュー v3：仕様案 Z（親行色分け + 子行白地）"
    ws.cell(1, 1).font = Font(name="Meiryo UI", size=12, bold=True)
    ws.row_dimensions[1].height = 21.0

    # Row 2: 空行
    ws.row_dimensions[2].height = 14.25

    # Row 3: ヘッダー（本番のヘッダースタイルに近づける）
    header_fill = PatternFill("solid", fgColor="FF2F5496")
    header_font = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = _side("thin", "808080")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, label in enumerate(HEADERS, start=1):
        cell = ws.cell(HEADER_ROW, col_idx)
        cell.value = label
        cell.font = copy(header_font)
        cell.fill = copy(header_fill)
        cell.alignment = copy(header_align)
        cell.border = copy(header_border)
    ws.row_dimensions[HEADER_ROW].height = 30.0


def _apply_parent_row(ws, row_idx: int, style_name: str,
                      values: dict[int, object]) -> None:
    """親行: A〜W 全セルに style 適用（要件 §4 遵守）。"""
    for col in range(1, MAX_COL + 1):
        cell = ws.cell(row_idx, col)
        cell.value = values.get(col, "")
        cell.style = style_name
    # 金額列の number_format（Python 側で設定、要件 Q5 確定）
    if values.get(15) is not None:
        ws.cell(row_idx, 15).number_format = "#,##0"
    if values.get(16) is not None:
        ws.cell(row_idx, 16).number_format = "#,##0"
    ws.row_dimensions[row_idx].height = ROW_HEIGHT


def _apply_child_row(ws, row_idx: int,
                     values: dict[int, object],
                     severity_label: str) -> None:
    """子行: A〜W 全セルに child_row_style、C 列のみ Alignment(indent=2)（方式 α）。

    v3（仕様案 Z）: severity fill の重ね塗りを行わない。子行は白地維持。
    個別 severity は A 列の絵文字（🔴/🟠/🟡/🟢）でのみ表現する。

    引数 severity_label は互換のため残すが v3 では未使用。
    """
    _ = severity_label  # v3 では未使用（v2 からのシグネチャ互換）

    for col in range(1, MAX_COL + 1):
        cell = ws.cell(row_idx, col)
        cell.value = values.get(col, "")
        cell.style = "child_row_style"
        # ── C 列のみインデント ──
        if col == 3:
            cell.alignment = Alignment(
                vertical="center", horizontal="left", indent=2,
            )
    # 金額列の number_format
    if values.get(15) is not None:
        ws.cell(row_idx, 15).number_format = "#,##0"
    if values.get(16) is not None:
        ws.cell(row_idx, 16).number_format = "#,##0"
    ws.row_dimensions[row_idx].height = ROW_HEIGHT


def _set_col_widths(ws) -> None:
    from openpyxl.utils import get_column_letter
    for idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


# ─────────────────────────────────────────────────────────────
# メイン
# ─────────────────────────────────────────────────────────────

def main() -> None:
    wb = Workbook()

    # Named Style 5 つを登録
    for name, bg, accent in PARENT_STYLES:
        wb.add_named_style(make_parent_style(name, bg, accent))
    wb.add_named_style(make_child_style())

    ws = wb.active
    ws.title = "A5 プレビュー"

    _write_header(ws)
    _set_col_widths(ws)

    # サンプル 4 セット配置
    row = DATA_START
    for s in SAMPLE_SETS:
        _apply_parent_row(ws, row, s["style"], s["parent"])
        row += 1
        for child in s["children"]:
            _apply_child_row(ws, row, child, s["severity_label"])
            row += 1

    # 参考情報シート（悠皓さん向け視覚確認ポイント）
    ws_info = wb.create_sheet("📋 視覚確認ポイント")
    ws_info.column_dimensions["A"].width = 5
    ws_info.column_dimensions["B"].width = 90
    notes = [
        ("①", "親行 4 色がちゃんと見分けられるか（Critical / Warning / Medium / Low が一目で区別できるか）"),
        ("②", "親行の帯が行全体に自然につながっているか（A〜W 列が 1 本の帯として見えるか）"),
        ("③", "子行 C 列のインデントが強すぎないか（indent=2 が適切かどうか）"),
        ("④", "子行白地で情報欠落がないか — 子行 A 列の severity 絵文字（🔴/🟠/🟡/🟢）だけで個別 severity が十分伝わるか。「やっぱり子行にも色が欲しい」と感じる場面がないか（白地に絵文字のみで severity が読み取れるか検証）"),
        ("", ""),
        ("親行 色見本", "Critical=#FCEBEB / Warning=#FAEEDA / Medium=#FEF5D6 / Low=#EAF3DE"),
        ("親行 罫線色", "Critical=#C00000 / Warning=#ED7D31 / Medium=#BF8F00 / Low=#548235"),
        ("子行", "白地（塗りなし）。severity は A 列絵文字のみで表現。"),
        ("", ""),
        ("v3 差分", "v2 では子行にも severity fill（#FFC7CE / #FFEB9C / #C6EFCE）を重ね塗りしていたが、v3 = 仕様案 Z では子行を白地のまま維持。親行 4 色だけで severity の帯を示し、子行は絵文字依存。"),
        ("v3 実装差分", "_apply_child_row 内の `cell.fill = copy(sev_fill)` を削除のみ。その他（絵文字・文言・インデント・罫線・リンク列）は v2 と同一。"),
        ("", ""),
        ("調整希望の場合", "色 hex または罫線太さを伝えてください → Claude Code がスクリプト側で調整し再生成します"),
    ]
    for i, (mark, text) in enumerate(notes, start=1):
        ws_info.cell(i, 1).value = mark
        ws_info.cell(i, 2).value = text
        ws_info.cell(i, 1).font = Font(name="Meiryo UI", size=11, bold=True)
        ws_info.cell(i, 2).font = Font(name="Meiryo UI", size=11)

    out = Path("tmp/preview_named_styles_sample.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(out)
    except PermissionError:
        # v1/v2 が Excel で開かれている場合のフォールバック
        out = Path("tmp/preview_named_styles_sample_v3.xlsx")
        wb.save(out)
        print("[WARN] 既存ファイルが開かれていたため v3 を別名保存しました")
    print(f"saved: {out}")
    print("Named styles:",
          [s.name if hasattr(s, "name") else s for s in wb.named_styles])


if __name__ == "__main__":
    main()
