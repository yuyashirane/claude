"""Phase 8-B 本番テンプレ反映スクリプト。

TC_template.xlsx に 5 つの Named Style を追加する。
- 実行前に自動バックアップ（TC_template_backup_YYYYMMDD_HHMMSS.xlsx）
- 既存衝突はエラー終了（上書きしない）
- 保存後の即時検証

悠皓さんがプレビュー (tmp/preview_named_styles_sample.xlsx) を視覚確認し
OK を出した後にのみ実行する。

Usage:
    cd office/office-claude
    py tmp/add_named_styles.py
"""
from __future__ import annotations

import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.styles.borders import Border, Side


TPL = Path("data/reports/template/TC_template.xlsx")

# 要件ドキュメント v2 §3 の色定義
CRITICAL_BG,     CRITICAL_BORDER = "FCEBEB", "C00000"
WARNING_BG,      WARNING_BORDER  = "FAEEDA", "ED7D31"
MEDIUM_BG,       MEDIUM_BORDER   = "FEF5D6", "BF8F00"
LOW_BG,          LOW_BORDER      = "EAF3DE", "548235"

PARENT_STYLES = [
    ("parent_row_style_critical", CRITICAL_BG, CRITICAL_BORDER),
    ("parent_row_style_warning",  WARNING_BG,  WARNING_BORDER),
    ("parent_row_style_medium",   MEDIUM_BG,   MEDIUM_BORDER),
    ("parent_row_style_low",      LOW_BG,      LOW_BORDER),
]

CHILD_STYLE = "child_row_style"

ALL_TARGET_NAMES = [name for name, _, _ in PARENT_STYLES] + [CHILD_STYLE]


def _side(style: str, color: str) -> Side:
    return Side(border_style=style, color=f"FF{color}")


def make_parent_style(name: str, bg_hex: str, accent_hex: str) -> NamedStyle:
    ns = NamedStyle(name=name)
    ns.font = Font(name="Meiryo UI", size=10, bold=True, color="FF000000")
    ns.fill = PatternFill("solid", fgColor=f"FF{bg_hex}")
    ns.alignment = Alignment(vertical="center", horizontal="left")
    thin_gray = _side("thin", "D9D9D9")
    ns.border = Border(
        left=thin_gray,
        right=thin_gray,
        top=_side("medium", accent_hex),
        bottom=_side("thin", accent_hex),
    )
    return ns


def make_child_style() -> NamedStyle:
    ns = NamedStyle(name=CHILD_STYLE)
    ns.font = Font(name="Meiryo UI", size=10, bold=False, color="FF000000")
    # 背景色は設定しない（severity_fills 重ね塗りを壊さない）
    ns.alignment = Alignment(vertical="center", horizontal="left")
    thin = _side("thin", "D9D9D9")
    ns.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return ns


def _existing_style_names(wb) -> set[str]:
    return {s.name if hasattr(s, "name") else s for s in wb.named_styles}


def _make_backup(src: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dst = src.with_name(f"{src.stem}_backup_{ts}{src.suffix}")
    shutil.copy2(src, dst)
    return dst


def main() -> int:
    if not TPL.exists():
        print(f"[ERROR] テンプレートが存在しません: {TPL}", file=sys.stderr)
        print("カレントディレクトリが office-claude かを確認してください。",
              file=sys.stderr)
        return 2

    # ── 事前読み込みで衝突チェック ──
    wb = load_workbook(TPL)
    existing = _existing_style_names(wb)
    conflicts = [n for n in ALL_TARGET_NAMES if n in existing]
    if conflicts:
        print(f"[ERROR] 既存の Named Style と衝突: {conflicts}", file=sys.stderr)
        print("上書きは行いません。既存を削除してから再実行してください。",
              file=sys.stderr)
        print(f"既存登録全体: {sorted(existing)}", file=sys.stderr)
        return 3

    # ── バックアップ ──
    backup = _make_backup(TPL)
    print(f"[INFO] バックアップ作成: {backup.name}")

    # ── 追加 ──
    print("[INFO] Named Style 追加中...")
    for name, bg, accent in PARENT_STYLES:
        wb.add_named_style(make_parent_style(name, bg, accent))
        print(f"  ✅ {name}")
    wb.add_named_style(make_child_style())
    print(f"  ✅ {CHILD_STYLE}")

    # ── 保存 ──
    wb.save(TPL)
    print(f"[INFO] 保存完了: {TPL.name}")

    # ── 即時検証 ──
    print("[INFO] 検証中...")
    wb2 = load_workbook(TPL)
    after = _existing_style_names(wb2)
    missing = [n for n in ALL_TARGET_NAMES if n not in after]
    if missing:
        print(f"[ERROR] 検証失敗: 登録されていない style があります: {missing}",
              file=sys.stderr)
        print(f"バックアップから復元してください:")
        print(f"    cp {backup} {TPL}")
        return 4

    print(f"    {len(ALL_TARGET_NAMES)}/{len(ALL_TARGET_NAMES)} style が正しく登録されています")
    for n in ALL_TARGET_NAMES:
        print(f"      - {n}")
    print("[SUCCESS] 完了")
    return 0


if __name__ == "__main__":
    sys.exit(main())
