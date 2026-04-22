"""Phase 8-C E2E 検証スクリプト（Phase 8-B 14 項目継承 + Phase 8-C 3 項目追加）。

Phase 4 のゴール:
    antread 2025/12 の実データで export_to_excel(ctx=ctx) を回し、
    生成された e2e_report_phase8c_*.xlsx が以下 17 項目を満たすかを判定する。

検証項目:
    Phase 8-B 継承 (14 項目):
        1. テンプレ本体の 5 NamedStyle 存在
        2. 生成ファイルにも 5 NamedStyle が保持
        3. 親行で parent_row_style_* が使われている
        4. 子行で child_row_style が使われている
        5. 親行 C 列が C-β-3 形式（"— X 件・合計 ¥Y"）を含む
        6. 親行 D/E 列が非空
        7. Pattern B（TC-06）が存在すれば "税区分混在" 文言
        8. 子行の Q/R 列ハイパーリンクが維持
        9. 子行の M 列 severity が空白
        10. 子行 fill_type が None/"none"
        11. サマリーシート B3/B4 に会社名・対象月
        12. 参考シートが存在
        13. save/load ラウンドトリップで全 5 style 維持

    Phase 8-C 追加 (Fix v2 改訂版):
        14. 親行 Q 列 = "GL" + ハイパーリンク
        14.1. 親行 GL URL の期間が単月 (start_date == 取引月初, end_date == 取引月末)
        14.2. 親行 GL URL に tax_group_codes が 1 つ以上含まれる (ctx.tax_code_master 解決可能時)
        15. 子行 E 列が非空 (Finding.message 由来、D 列は常に空欄)
        16. サマリー B4 対象月表示が format_target_month の出力形式

Usage:
    cd office/office-claude
    PYTHONIOENCODING=utf-8 py tmp/verify_phase8c.py
"""
from __future__ import annotations

import importlib.util
import io
import sys
import urllib.parse
from datetime import datetime
from pathlib import Path

sys.path.insert(0, ".")

from openpyxl import load_workbook

from scripts.e2e.freee_to_context import build_check_context
from skills.export.excel_report.exporter import export_to_excel


TPL = Path("data/reports/template/TC_template.xlsx")
BASE = Path("data/e2e/3525430/202512")

EXPECTED_STYLES = {
    "parent_row_style_critical": "FCEBEB",
    "parent_row_style_warning":  "FAEEDA",
    "parent_row_style_medium":   "FEF5D6",
    "parent_row_style_low":      "EAF3DE",
    "child_row_style":           None,
}


class Verifier:
    def __init__(self) -> None:
        self.passes: list[str] = []
        self.failures: list[str] = []

    def check(self, cond: bool, msg: str) -> None:
        if cond:
            print(f"  ✅ {msg}")
            self.passes.append(msg)
        else:
            print(f"  ❌ {msg}")
            self.failures.append(msg)

    def ok(self) -> bool:
        return not self.failures


def _all_styles_present(wb) -> tuple[bool, list[str]]:
    names = {s.name if hasattr(s, "name") else s for s in wb.named_styles}
    missing = [n for n in EXPECTED_STYLES if n not in names]
    return (not missing, missing)


def _run_e2e() -> tuple[Path, object]:
    """antread 2025/12 の実データで Excel 出力を生成する。

    Phase 8-C: ctx を export_to_excel に渡すことで、親行 Q 列の GL リンクが
    会計期間全体 (ctx.period_start〜ctx.period_end) で生成される。
    """
    skill_dir = Path("skills/verify/V1-3-rule/check-tax-classification")
    sys.path.insert(0, str(skill_dir))
    spec = importlib.util.spec_from_file_location("checker", skill_dir / "checker.py")
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)

    ctx = build_check_context(
        deals_path=BASE / "deals_202512.json",
        partners_path=BASE / "partners_all.json",
        account_items_path=BASE / "account_items_all.json",
        company_info_path=BASE / "company_info.json",
        taxes_codes_path=BASE / "taxes_codes.json",
    )
    findings = m.run(ctx)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output = BASE / f"e2e_report_phase8c_{ts}.xlsx"
    # Phase 8-C: ctx を渡す
    export_to_excel(
        findings, output,
        company_name=ctx.company_name,
        period=BASE.name,
        ctx=ctx,
    )
    print(f"  → Generated: {output.name} ({output.stat().st_size:,} bytes, "
          f"{len(findings)} findings)")
    print(f"  → ctx.period_start = {ctx.period_start}, period_end = {ctx.period_end}")
    return output, ctx


def main() -> int:
    v = Verifier()

    # ─── 項目 1: テンプレ本体の style 存在 ───
    print("=" * 64)
    print("[1] テンプレ本体の 5 NamedStyle 存在チェック")
    print("=" * 64)
    if not TPL.exists():
        print(f"[ERROR] テンプレートが存在しません: {TPL}", file=sys.stderr)
        return 2
    wb_tpl = load_workbook(TPL)
    ok, missing = _all_styles_present(wb_tpl)
    v.check(ok, f"テンプレに 5 style 全件登録 (missing={missing if missing else 'なし'})")

    # ─── E2E 実行 ───
    print()
    print("=" * 64)
    print("E2E 実行（antread 2025/12、Phase 8-C: ctx 連携）")
    print("=" * 64)
    output, ctx = _run_e2e()
    wb = load_workbook(output)

    # ─── 項目 2: 生成ファイルの style 存在 ───
    print()
    print("=" * 64)
    print("[2] 生成ファイルに 5 NamedStyle 保持")
    print("=" * 64)
    ok, missing = _all_styles_present(wb)
    v.check(ok, f"生成ファイルに 5 style 保持 (missing={missing if missing else 'なし'})")

    # ─── 詳細シートを走査して検証項目 3〜10 + Phase 8-C 追加項目を判定 ───
    detail_sheets = [
        n for n in wb.sheetnames
        if n not in ("サマリー", "参考") and n.startswith("A")
    ]
    print()
    print(f"詳細シート {len(detail_sheets)} 枚を走査: {detail_sheets}")

    found_parent_style = False
    found_child_style = False
    found_summary_form = False   # C-β-3 "— X 件・合計 ¥Y"
    found_observation = False    # D 列非空
    found_check_result = False   # E 列非空
    found_mixing_tag = False     # Pattern B "税区分混在"
    found_hyperlink = False      # 子行 Q/R
    found_m_blank_child = False  # 子行 M 列空白
    found_child_no_fill = False  # 子行 fill_type None
    has_tc06 = False

    # Phase 8-C 追加 (Fix v2 改訂版)
    found_parent_gl_link = False        # 項目 14: 親行 Q に "GL" + URL
    found_parent_gl_monthly = False     # 項目 14.1: URL 期間が単月 (start/end が同月かつ 1日〜月末)
    found_parent_gl_tax_codes = False   # 項目 14.2: URL に tax_group_codes が 1 つ以上
    found_child_e_nonempty = False      # 項目 15: 子行 E 列が非空
    found_child_d_empty = True          # 項目 15: 子行 D 列は常に空欄（反例検出で False）

    for sheet_name in detail_sheets:
        ws = wb[sheet_name]
        for row in range(4, ws.max_row + 1):
            style_name = ws.cell(row=row, column=3).style
            c_val = ws.cell(row=row, column=3).value or ""
            d_val = ws.cell(row=row, column=4).value or ""
            e_val = ws.cell(row=row, column=5).value or ""
            m_val = ws.cell(row=row, column=13).value
            q_cell = ws.cell(row=row, column=17)
            r_cell = ws.cell(row=row, column=18)
            tc_val = ws.cell(row=row, column=2).value or ""

            if style_name.startswith("parent_row_style_"):
                found_parent_style = True
                if "件" in c_val and "合計" in c_val:
                    found_summary_form = True
                if isinstance(d_val, str) and d_val.strip():
                    found_observation = True
                if isinstance(e_val, str) and e_val.strip():
                    found_check_result = True
                if "TC-06" in tc_val:
                    has_tc06 = True
                    if ("税区分混在" in str(c_val)
                            or "税区分混在" in str(d_val)
                            or "税区分混在" in str(e_val)):
                        found_mixing_tag = True

                # Phase 8-C ②: 親行 Q に GL リンク
                if q_cell.value == "GL" and q_cell.hyperlink is not None:
                    found_parent_gl_link = True
                    url = q_cell.hyperlink.target or ""
                    # URL をパースして期間と税区分コードを検査
                    parsed = urllib.parse.urlparse(url)
                    qs = urllib.parse.parse_qs(parsed.query)
                    sd = qs.get("start_date", [""])[0]
                    ed = qs.get("end_date", [""])[0]
                    # 単月判定: start/end が "YYYY-MM-DD" 形式で、年月一致・start は 01 日
                    if len(sd) == 10 and len(ed) == 10 and sd[:7] == ed[:7] and sd[8:10] == "01":
                        # end_date が月末（28〜31）かを緩く判定
                        if ed[8:10] in ("28", "29", "30", "31"):
                            found_parent_gl_monthly = True
                    # tax_group_codes が 1 つ以上含まれる
                    if qs.get("tax_group_codes"):
                        found_parent_gl_tax_codes = True

            elif style_name == "child_row_style":
                found_child_style = True
                if q_cell.hyperlink or r_cell.hyperlink:
                    found_hyperlink = True
                if m_val is None or m_val == "":
                    found_m_blank_child = True
                fill_type = ws.cell(row=row, column=3).fill.fill_type
                if fill_type in (None, "none"):
                    found_child_no_fill = True
                # Phase 8-C Fix v2 ④: 子行 E 列は非空、D 列は常に空欄
                if isinstance(e_val, str) and e_val.strip():
                    found_child_e_nonempty = True
                # D 列に非空文字列があれば反例（常に空欄仕様を破っている）
                if isinstance(d_val, str) and d_val.strip():
                    found_child_d_empty = False

    print()
    print("=" * 64)
    print("[3〜10] 詳細シート走査結果（Phase 8-B 継承）")
    print("=" * 64)
    v.check(found_parent_style, "親行で parent_row_style_* が使われている")
    v.check(found_child_style,  "子行で child_row_style が使われている")
    v.check(found_summary_form, "親行 C 列に C-β-3 形式（件・合計 ¥）")
    v.check(found_observation,  "親行 D 列が非空（観点）")
    v.check(found_check_result, "親行 E 列が非空（チェック結果）")
    if has_tc06:
        v.check(found_mixing_tag, "Pattern B (TC-06) に '税区分混在' 文言")
    else:
        print("  ⏭  TC-06 が検出 Finding になかったため Pattern B チェックはスキップ")
    v.check(found_hyperlink,    "子行 Q/R 列ハイパーリンク保持")
    v.check(found_m_blank_child, "子行 M 列（severity）が空白（仕様案 Z）")
    v.check(found_child_no_fill, "子行 fill_type = None/none（仕様案 Z 白地）")

    # ─── 項目 11: サマリー ───
    print()
    print("=" * 64)
    print("[11] サマリーシート会社名・対象月")
    print("=" * 64)
    ws_sum = wb["サマリー"]
    b3 = ws_sum.cell(3, 2).value
    b4 = ws_sum.cell(4, 2).value
    v.check(bool(b3), f"サマリー B3 会社名 = {b3!r}")
    v.check(bool(b4), f"サマリー B4 対象月 = {b4!r}")

    # ─── 項目 12: 参考シート ───
    print()
    print("=" * 64)
    print("[12] 参考シート存在")
    print("=" * 64)
    v.check("参考" in wb.sheetnames, f"'参考' シート存在 (sheets={wb.sheetnames})")

    # ─── 項目 13: ラウンドトリップ ───
    print()
    print("=" * 64)
    print("[13] save/load ラウンドトリップで 5 style 維持")
    print("=" * 64)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = load_workbook(buf)
    ok, missing = _all_styles_present(wb2)
    v.check(ok, f"ラウンドトリップ後も 5 style 維持 (missing={missing if missing else 'なし'})")

    # ─── Phase 8-C 追加項目 (Fix v2) ───
    print()
    print("=" * 64)
    print("[14] Phase 8-C ②: 親行 Q 列に GL ハイパーリンク")
    print("=" * 64)
    v.check(found_parent_gl_link, "親行 Q 列 = 'GL' + ハイパーリンクあり")

    print()
    print("=" * 64)
    print("[14.1] Phase 8-C Fix v2: 親行 GL URL の期間が単月（取引月初〜月末）")
    print("=" * 64)
    v.check(found_parent_gl_monthly,
            "親行 GL URL の start_date/end_date が同一月かつ月初〜月末形式")

    print()
    print("=" * 64)
    print("[14.2] Phase 8-C Fix v2: 親行 GL URL に tax_group_codes が付与")
    print("=" * 64)
    v.check(found_parent_gl_tax_codes,
            "親行 GL URL に tax_group_codes=... が 1 つ以上含まれる")

    print()
    print("=" * 64)
    print("[15] Phase 8-C Fix v2 ④: 子行 E 列 message 由来、D 列は常に空欄")
    print("=" * 64)
    v.check(found_child_e_nonempty, "子行 E 列が非空（Finding.message 由来）")
    v.check(found_child_d_empty, "子行 D 列は全行空欄（Fix v2 仕様）")

    print()
    print("=" * 64)
    print("[16] Phase 8-C γ: サマリー B4 対象月が format_target_month 出力形式")
    print("=" * 64)
    # 単月 "YYYY年M月" または累計 "YYYY年M月〜YYYY年M月" のいずれか
    b4_str = str(b4 or "")
    is_monthly_form = ("年" in b4_str and "月" in b4_str)
    v.check(is_monthly_form,
            f"サマリー B4 対象月が format_target_month 出力形式: {b4_str!r}")

    # ─── サマリー ───
    print()
    print("=" * 64)
    total = len(v.passes) + len(v.failures)
    if v.ok():
        print(f"[SUCCESS] Phase 8-C E2E 検証 {total}/{total} 項目 PASS")
        print(f"生成ファイル: {output.resolve()}")
        return 0
    print(f"[FAILURE] {len(v.failures)}/{total} 項目が不一致")
    for m in v.failures:
        print(f"  - {m}")
    return 1


if __name__ == "__main__":
    sys.exit(main())
