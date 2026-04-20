"""Phase 8-B 事前調査: インデント 3 案のサンプル Excel を生成する。

悠皓さんが視覚判断するための素材。tmp/ 配下に生成し、
コミットせず破棄可能。
"""
from __future__ import annotations

from pathlib import Path
from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, NamedStyle, Alignment
from openpyxl.styles.borders import Border, Side


TPL = Path("data/reports/template/TC_template.xlsx")
OUT_A = Path("tmp/indent_sample_a_alignment.xlsx")
OUT_B = Path("tmp/indent_sample_b_zenkaku.xlsx")
OUT_C = Path("tmp/indent_sample_c_c_only.xlsx")


def _make_parent_style() -> NamedStyle:
    ns = NamedStyle(name="parent_row_style")
    ns.font = Font(name="Meiryo UI", size=10, bold=True, color="FF1F4E78")
    ns.fill = PatternFill("solid", fgColor="FFDEEBF7")  # 淡い青
    ns.alignment = Alignment(vertical="center", horizontal="left")
    thin = Side(border_style="thin", color="FF8EA9DB")
    top = Side(border_style="medium", color="FF2F5496")
    ns.border = Border(left=thin, right=thin, top=top, bottom=thin)
    return ns


def _make_child_style_a() -> NamedStyle:
    """案 a: Alignment.indent=1 でテキスト左マージン。"""
    ns = NamedStyle(name="child_row_style")
    ns.font = Font(name="Meiryo UI", size=10, color="FF595959")
    ns.fill = PatternFill("solid", fgColor="FFF8F8F8")
    ns.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    thin = Side(border_style="thin", color="FFD9D9D9")
    ns.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return ns


def _make_child_style_plain() -> NamedStyle:
    """案 b/c で共通使用する子行スタイル(インデントなし)。"""
    ns = NamedStyle(name="child_row_style_plain")
    ns.font = Font(name="Meiryo UI", size=10, color="FF595959")
    ns.fill = PatternFill("solid", fgColor="FFF8F8F8")
    ns.alignment = Alignment(vertical="center", horizontal="left")
    thin = Side(border_style="thin", color="FFD9D9D9")
    ns.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return ns


def _find_area_sheet(wb, prefix: str = "A5"):
    for name in wb.sheetnames:
        if name.startswith(prefix):
            return wb[name]
    raise RuntimeError(f"area sheet {prefix} not found")


def _write_header_and_sample(ws, parent_style: NamedStyle, child_style: NamedStyle,
                             indent_col_only: bool, zenkaku_prefix: bool) -> None:
    # Row 4: parent row
    parent_vals = {
        1: "重大",
        2: "TC-03a",
        3: f"給与手当 — 3 件・合計 ¥900,000",
        4: "（親行）",
        5: "集約: 給与系勘定に課税仕入が 3 件混入",
        9: "給与手当",
        15: 900000,
    }
    for col in range(1, 24):
        c = ws.cell(4, col)
        c.value = parent_vals.get(col, "")
        c.style = "parent_row_style"
    ws.cell(4, 15).number_format = "#,##0"

    # Row 5-7: child rows
    child_data = [
        ("重大", "TC-03a", "課税仕入の給与", "給与手当", "課対仕入10%", "対象外", "2026/01/25", 300000),
        ("重大", "TC-03a", "課税仕入の給与", "給与手当", "課対仕入10%", "対象外", "2026/02/25", 300000),
        ("重大", "TC-03a", "課税仕入の給与", "給与手当", "課対仕入10%", "対象外", "2026/03/25", 300000),
    ]
    for i, row in enumerate(child_data):
        rnum = 5 + i
        values = {
            1: row[0], 2: row[1], 3: row[2], 9: row[3],
            6: row[4], 7: row[5], 8: row[6], 15: row[7],
        }
        if zenkaku_prefix:
            values[3] = "　" + str(values[3])  # 全角スペース
        for col in range(1, 24):
            c = ws.cell(rnum, col)
            c.value = values.get(col, "")
            c.style = "child_row_style"
            if indent_col_only and col == 3:
                # 案 c: C 列だけ indent を強制
                al = copy(c.alignment)
                c.alignment = Alignment(
                    vertical=al.vertical, horizontal=al.horizontal, indent=2,
                )
        ws.cell(rnum, 15).number_format = "#,##0"


def build(out_path: Path, indent_mode: str) -> None:
    wb = load_workbook(TPL)

    parent = _make_parent_style()
    wb.add_named_style(parent)

    if indent_mode == "a":
        child = _make_child_style_a()
        wb.add_named_style(child)
        indent_col_only = False
        zenkaku_prefix = False
    elif indent_mode == "b":
        child = _make_child_style_plain()
        child.name = "child_row_style"
        wb.add_named_style(child)
        indent_col_only = False
        zenkaku_prefix = True
    elif indent_mode == "c":
        child = _make_child_style_plain()
        child.name = "child_row_style"
        wb.add_named_style(child)
        indent_col_only = True
        zenkaku_prefix = False
    else:
        raise ValueError(indent_mode)

    ws = _find_area_sheet(wb, "A5")
    # 既存 Row 4 以降をクリア(値だけ)
    last = ws.max_row
    for r in range(4, max(last + 1, 9)):
        for c in range(1, 24):
            ws.cell(r, c).value = None

    _write_header_and_sample(ws, parent, child,
                             indent_col_only=indent_col_only,
                             zenkaku_prefix=zenkaku_prefix)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"saved: {out_path}")


def main() -> None:
    build(OUT_A, "a")
    build(OUT_B, "b")
    build(OUT_C, "c")


if __name__ == "__main__":
    main()
