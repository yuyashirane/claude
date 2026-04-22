"""Phase 6 Excel 出力テスト（Phase 6.12 テンプレート駆動版）。

構造テスト（スナップショット不使用）。
生成された .xlsx を openpyxl で再オープンしてピンポイント検証。

テスト分類:
    A. シート構造    5件
    B. 列構成        4件
    C. データ配置    5件
    D. ソート        3件
    E. 視覚装飾      4件
    F. エッジケース  5件  (+1: テンプレートエラー)
    G. Phase 6.11a 追加テスト  5件
    H. Phase 6.11a v2 追加テスト  6件
    I. Phase 6.12 テンプレート駆動テスト  3件
    J. Phase 6.12a スタイル保持テスト  2件
    K. Phase 6.11b 金額列O/Pテスト  3件
    L. 対象月フォーマットテスト  2件
    M. Phase 7 Q/R 列ハイパーリンクテスト  5件
合計: 52件

詳細シートレイアウト（テンプレート準拠、23列）:
    Row 1: シートタイトル
    Row 2: 空行
    Row 3: ヘッダー行（23列）
    Row 4+: データ行

列マッピング（23列）:
    A=優先度, B=項目, C=項目名, D=観点, E=チェック結果,
    F=現在の税区分, G=推奨税区分, H=取引日, I=勘定科目, J=取引先,
    K=品目, L=部門, M=メモ, N=摘要, O=借方金額, P=貸方金額,
    Q=🔗総勘定元帳, R=🔗仕訳帳, S=確信度, T=エラー型,
    U=確認状況, V=担当者メモ, W=walletTxnId
"""
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────────
# Finding ビルダー
# ─────────────────────────────────────────────────────────────────────

def _make_finding(
    sub_code: str = "TC-07a",
    area: str = "A10",
    severity: str = "🔴 High",
    show_by_default: bool = True,
    sort_priority: int = 12,
    error_type: str = "direct_error",
    review_level: str = "🔴必修",
    current_value: str = "課対仕入10%",
    suggested_value: str = "対象外",
    confidence: int = 80,
    message: str = "テストメッセージ:慶弔見舞金が課税仕入になっています。",
    wallet_txn_id: str = "test-txn-001",
    link_hints=None,
    debit_amount=None,
    credit_amount=None,
    deal_id=None,
):
    """テスト用 Finding ビルダー（実際の schema.py の Finding 定義に完全準拠）。

    tc_code は sub_code[:5] から自動導出（TC-07a → TC-07）。
    """
    import importlib.util
    if "schema" not in sys.modules:
        _root = Path(__file__).parent.parent.parent
        _schema_path = (
            _root / "skills" / "verify" / "V1-3-rule"
            / "check-tax-classification" / "schema.py"
        )
        spec = importlib.util.spec_from_file_location("schema", _schema_path)
        mod = importlib.util.module_from_spec(spec)
        sys.modules["schema"] = mod
        spec.loader.exec_module(mod)

    schema = sys.modules["schema"]
    tc_code = sub_code[:5]

    return schema.Finding(
        tc_code=tc_code,
        sub_code=sub_code,
        severity=severity,
        error_type=error_type,
        review_level=review_level,
        area=area,
        sort_priority=sort_priority,
        wallet_txn_id=wallet_txn_id,
        current_value=current_value,
        suggested_value=suggested_value,
        confidence=confidence,
        message=message,
        show_by_default=show_by_default,
        link_hints=link_hints,
        debit_amount=debit_amount,
        credit_amount=credit_amount,
        deal_id=deal_id,
    )


def _make_finding_with_hints(sub_code="TC-07a", area="A10", **kwargs):
    """link_hints 付き Finding。取引日・勘定科目の列テスト用。"""
    import importlib.util
    if "schema" not in sys.modules:
        _root = Path(__file__).parent.parent.parent
        _schema_path = (
            _root / "skills" / "verify" / "V1-3-rule"
            / "check-tax-classification" / "schema.py"
        )
        spec = importlib.util.spec_from_file_location("schema", _schema_path)
        mod = importlib.util.module_from_spec(spec)
        sys.modules["schema"] = mod
        spec.loader.exec_module(mod)

    schema = sys.modules["schema"]
    lh = schema.LinkHints(
        target="general_ledger",
        account_name="福利厚生費",
        period_start=date(2026, 2, 1),
        period_end=date(2026, 2, 28),
    )
    return _make_finding(sub_code=sub_code, area=area, link_hints=lh, **kwargs)


# ─────────────────────────────────────────────────────────────────────
# A. シート構造（5件）
# ─────────────────────────────────────────────────────────────────────

def test_summary_sheet_exists(tmp_path):
    """空 findings でもサマリーシートが生成される。"""
    from skills.export.excel_report.exporter import export_to_excel
    output = tmp_path / "out.xlsx"
    export_to_excel([], output)
    wb = load_workbook(output)
    assert "サマリー" in wb.sheetnames


def test_area_sheet_created_when_findings_exist(tmp_path):
    """TC-02 の Finding があれば A4 シートが生成される。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-02c", area="A4", sort_priority=8)]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    assert "A4 家賃・地代" in wb.sheetnames


def test_area_sheet_not_created_when_no_findings(tmp_path):
    """TC-02 のみの場合、TC-03 用の A5 シートは生成されない。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-02c", area="A4", sort_priority=8)]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    assert "A5 人件費" not in wb.sheetnames


def test_multi_area_multi_sheet(tmp_path):
    """複数 area の Finding があれば複数シートが生成される。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [
        _make_finding(sub_code="TC-02c", area="A4", sort_priority=8),
        _make_finding(sub_code="TC-07a", area="A10", sort_priority=12),
    ]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    assert "A4 家賃・地代" in wb.sheetnames
    assert "A10 その他経費" in wb.sheetnames


def test_empty_findings_summary_and_sankou(tmp_path):
    """空 findings 時はサマリーシートと参考シートが残る（area シートは削除）。"""
    from skills.export.excel_report.exporter import export_to_excel
    output = tmp_path / "out.xlsx"
    export_to_excel([], output)
    wb = load_workbook(output)
    assert "サマリー" in wb.sheetnames
    assert "参考" in wb.sheetnames
    # area シートは全削除
    for area_sheet in ["A4 家賃・地代", "A5 人件費", "A8 売上",
                       "A10 その他経費", "A11 営業外・特別損益", "A12 税金"]:
        assert area_sheet not in wb.sheetnames


# ─────────────────────────────────────────────────────────────────────
# B. 列構成（4件）
# ─────────────────────────────────────────────────────────────────────

def test_detail_sheet_has_23_columns(tmp_path):
    """詳細シートのヘッダー行（Row 3）に非空セルが 23 個ある。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    header_cells = [c for c in ws[3] if c.value is not None]
    assert len(header_cells) == 23


def test_summary_lower_table_has_11_columns(tmp_path):
    """サマリーシートの下部テーブルヘッダー行（Row 20）に非空セルが 11 個ある。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["サマリー"]
    header_cells = [c for c in ws[20] if c.value is not None]
    assert len(header_cells) == 11


def test_detail_sheet_header_values(tmp_path):
    """詳細シートのヘッダー行（Row 3）の主要ヘッダー文字列を検証。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    assert ws.cell(3, 1).value == "優先度"     # A
    assert ws.cell(3, 2).value == "項目"       # B
    assert ws.cell(3, 21).value == "確認状況"  # U（23列構成）


def test_column_widths_set(tmp_path):
    """詳細シートの列幅がテンプレートから継承される。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    assert ws.column_dimensions["A"].width == 8       # 優先度（テンプレート値）
    assert ws.column_dimensions["E"].width == 40.625  # チェック結果（テンプレート値）


# ─────────────────────────────────────────────────────────────────────
# C. データ配置（5件）
# ─────────────────────────────────────────────────────────────────────

def test_tc02_finding_placed_in_a4_sheet(tmp_path):
    """TC-02c の Finding が A4 シートの親行 (Row 4) B 列に配置される。

    Phase 8-B: 親行 B 列には代表 sub_code が入る（判断 3 確定仕様）。
    """
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-02c", area="A4", sort_priority=8)]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A4 家賃・地代"]
    assert ws.cell(4, 2).value == "TC-02c"   # 親行 B 列に sub_code


def test_tc07_finding_placed_in_a10_sheet(tmp_path):
    """TC-07f の Finding が A10 シートの親行 (Row 4) B 列に配置される。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07f", area="A10", show_by_default=False,
                               sort_priority=28)]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    assert ws.cell(4, 2).value == "TC-07f"   # 親行 B 列に sub_code


def test_current_and_suggested_value_placed(tmp_path):
    """現在の税区分 (F 列)・推奨税区分 (G 列) は子行に配置される。親行は空欄。

    Phase 8-B: 親行は集約サマリー専用のため F/G 列は空欄。個別 Finding の
    F/G 値は子行 (Row 5) に入る。
    """
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(
        sub_code="TC-07a", area="A10",
        current_value="課対仕入10%", suggested_value="対象外",
    )]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    # 親行 (Row 4): F/G 列は空欄
    assert ws.cell(4, 6).value in (None, ""), \
        f"親行 F4 は空欄: got {ws.cell(4, 6).value!r}"
    assert ws.cell(4, 7).value in (None, ""), \
        f"親行 G4 は空欄: got {ws.cell(4, 7).value!r}"
    # 子行 (Row 5): 個別の税区分
    assert ws.cell(5, 6).value == "課対仕入10%"   # 子行 F 列
    assert ws.cell(5, 7).value == "対象外"          # 子行 G 列


def test_wallet_txn_id_placed(tmp_path):
    """walletTxnId は子行 (Row 5) の W 列に配置される。親行 W は空欄。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10", wallet_txn_id="txn-xyz-999")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    # 親行 W 列は空欄
    assert ws.cell(4, 23).value in (None, ""), \
        f"親行 W4 は空欄: got {ws.cell(4, 23).value!r}"
    # 子行 W 列に txn-id
    assert ws.cell(5, 23).value == "txn-xyz-999"


def test_link_hints_account_name_in_column9(tmp_path):
    """link_hints.account_name は子行 (Row 5) の I 列 (勘定科目) に反映される。

    Phase 8-B: 親行の C 列サマリーにも勘定科目が含まれるが、個別の勘定科目値は
    子行 I 列に入る。親行 I 列は空欄。
    """
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding_with_hints(sub_code="TC-07a", area="A10")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    # 親行 I 列は空欄
    assert ws.cell(4, 9).value in (None, ""), \
        f"親行 I4 は空欄: got {ws.cell(4, 9).value!r}"
    # 子行 I 列に勘定科目名
    assert ws.cell(5, 9).value == "福利厚生費"


# ─────────────────────────────────────────────────────────────────────
# D. ソート（3件）
# ─────────────────────────────────────────────────────────────────────

def test_rows_sorted_by_sort_priority(tmp_path):
    """同一 area 内のグループが sort_priority 昇順にソートされる。

    Phase 8-B: 3 件の異なる sub_code → 3 グループ → 親子交互に 6 行展開。
    親行は Row 4 (TC-02a), Row 6 (TC-02c), Row 8 (TC-02f) に配置される
    （親4/子5/親6/子7/親8/子9）。
    """
    from skills.export.excel_report.exporter import export_to_excel
    findings = [
        _make_finding(sub_code="TC-02f", area="A4", sort_priority=18),
        _make_finding(sub_code="TC-02a", area="A4", sort_priority=6),
        _make_finding(sub_code="TC-02c", area="A4", sort_priority=8),
    ]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A4 家賃・地代"]
    # 親行（各グループ先頭）が sort_priority 昇順
    assert ws.cell(4, 2).value == "TC-02a"
    assert ws.cell(6, 2).value == "TC-02c"
    assert ws.cell(8, 2).value == "TC-02f"
    # 親行スタイルが適用されている
    assert ws.cell(4, 1).style.startswith("parent_row_style_")
    assert ws.cell(6, 1).style.startswith("parent_row_style_")
    assert ws.cell(8, 1).style.startswith("parent_row_style_")
    # 子行は child_row_style
    assert ws.cell(5, 1).style == "child_row_style"
    assert ws.cell(7, 1).style == "child_row_style"
    assert ws.cell(9, 1).style == "child_row_style"


def test_sort_priority_from_map_when_zero(tmp_path):
    """Finding.sort_priority が 0 の場合、SORT_PRIORITY_MAP から解決される。

    Phase 8-B: 2 件異 sub_code → 2 グループ → 親子交互に 4 行（親4/子5/親6/子7）。
    """
    from skills.export.excel_report.exporter import export_to_excel
    f_b = _make_finding(sub_code="TC-07b", area="A10", sort_priority=0)
    f_a = _make_finding(sub_code="TC-07a", area="A10", sort_priority=0)
    output = tmp_path / "out.xlsx"
    export_to_excel([f_b, f_a], output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    # 親行が sort_priority 解決で並ぶ
    assert ws.cell(4, 2).value == "TC-07a"
    assert ws.cell(6, 2).value == "TC-07b"


def test_summary_sheet_sorted_by_area_order(tmp_path):
    """サマリーシートは area 順（A4→A10）にソートされる。データは Row 21 から始まる。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [
        _make_finding(sub_code="TC-07a", area="A10", sort_priority=12),
        _make_finding(sub_code="TC-02c", area="A4", sort_priority=8),
    ]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["サマリー"]
    assert ws.cell(21, 1).value == "A4"
    assert ws.cell(22, 1).value == "A10"


# ─────────────────────────────────────────────────────────────────────
# E. 視覚装飾（4件）
# ─────────────────────────────────────────────────────────────────────

def test_high_severity_row_has_red_fill(tmp_path):
    """🔴 High/Critical の親行が重大色 (FCEBEB) で塗られる。

    Phase 8-B: 親行スタイル parent_row_style_critical の背景色は FCEBEB
    （Named Style テンプレ反映時の §3 確定値）。
    """
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10", severity="🔴 High")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    parent_cell = ws.cell(4, 1)
    # 親行スタイルが適用されている
    assert parent_cell.style == "parent_row_style_critical", \
        f"親行 A4 スタイル: got {parent_cell.style!r}"
    # 背景色が FCEBEB
    rgb = str(parent_cell.fill.start_color.rgb).upper()
    assert "FCEBEB" in rgb, f"親行 A4 fill: got {rgb!r}"


def test_show_by_default_false_row_visible(tmp_path):
    """show_by_default=False でも初期は全行表示（hidden=False）。

    Phase 8-B: 2 件異 sub_code → 2 グループ → 4 行（rows 4-7）が全て visible。
    """
    from skills.export.excel_report.exporter import export_to_excel
    findings = [
        _make_finding(sub_code="TC-07a", area="A10", show_by_default=True,
                       sort_priority=12),
        _make_finding(sub_code="TC-07f", area="A10", show_by_default=False,
                       sort_priority=28),
    ]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    # 親子 4 行すべて visible
    for r in range(4, 8):
        assert ws.row_dimensions[r].hidden is False, \
            f"Row {r} should be visible"


def test_header_row_frozen(tmp_path):
    """詳細シートのヘッダーが A4 で固定される。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    assert ws.freeze_panes == "A4"


def test_confirmation_column_has_dropdown(tmp_path):
    """確認状況列（U=21列目）にプルダウンが設定される。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    has_dv = any(
        "U" in str(dv.sqref)
        for dv in ws.data_validations.dataValidation
    )
    assert has_dv


# ─────────────────────────────────────────────────────────────────────
# F. エッジケース（5件）
# ─────────────────────────────────────────────────────────────────────

def test_url_none_leaves_link_cells_empty(tmp_path):
    """freee URL が None の場合、親行・子行とも Q/R 列が空欄になる。

    Phase 8-B: 親行 Q/R は常に空欄。子行 Q/R は link_hints が None または URL
    生成不能のときに空欄。
    """
    from skills.export.excel_report.exporter import export_to_excel
    f = _make_finding(sub_code="TC-07a", area="A10")
    output = tmp_path / "out.xlsx"
    export_to_excel([f], output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    # 親行
    assert ws.cell(4, 17).value in (None, "", "-")
    assert ws.cell(4, 18).value in (None, "", "-")
    # 子行
    assert ws.cell(5, 17).value in (None, "", "-")
    assert ws.cell(5, 18).value in (None, "", "-")


def test_single_finding_generates_valid_file(tmp_path):
    """Finding 1件で有効な .xlsx が生成される。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10")]
    output = tmp_path / "out.xlsx"
    result = export_to_excel(findings, output)
    assert result == output
    assert output.exists()
    assert output.stat().st_size > 0


def test_output_path_parent_must_exist(tmp_path):
    """親ディレクトリが存在しない場合は ValueError が発生する。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10")]
    invalid = tmp_path / "nonexistent_dir" / "out.xlsx"
    with pytest.raises(ValueError):
        export_to_excel(findings, invalid)


def test_type_error_when_findings_not_list(tmp_path):
    """findings が list でない場合は TypeError が発生する。"""
    from skills.export.excel_report.exporter import export_to_excel
    output = tmp_path / "out.xlsx"
    with pytest.raises(TypeError):
        export_to_excel("not a list", output)  # type: ignore


def test_template_not_found_raises_error(tmp_path):
    """存在しないテンプレートパスを渡すと FileNotFoundError が発生する。"""
    from skills.export.excel_report.exporter import export_to_excel
    output = tmp_path / "out.xlsx"
    with pytest.raises(FileNotFoundError):
        export_to_excel([], output, template_path=tmp_path / "no_such_template.xlsx")


# ─────────────────────────────────────────────────────────────────────
# G. Phase 6.11a 追加テスト（5件）
# ─────────────────────────────────────────────────────────────────────

def test_summary_header_title_in_row1(tmp_path):
    """サマリー Row 1 にレポートタイトルが入る（company_name なし）。"""
    from skills.export.excel_report.exporter import export_to_excel
    output = tmp_path / "out.xlsx"
    export_to_excel([], output)
    wb = load_workbook(output)
    ws = wb["サマリー"]
    assert ws.cell(1, 1).value == "消費税区分チェックレポート"


def test_summary_header_title_with_company_name(tmp_path):
    """company_name を渡すと Row 1 に '{名前} 消費税区分チェックレポート' が入る。"""
    from skills.export.excel_report.exporter import export_to_excel
    output = tmp_path / "out.xlsx"
    export_to_excel([], output, company_name="テスト株式会社")
    wb = load_workbook(output)
    ws = wb["サマリー"]
    assert "テスト株式会社" in ws.cell(1, 1).value
    assert "消費税区分チェックレポート" in ws.cell(1, 1).value


def test_summary_tc_matrix_tc07_row16(tmp_path):
    """サマリー Row 16（TC-07）に severity 別件数が入る。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [
        _make_finding(sub_code="TC-07a", area="A10", severity="🔴 High"),
        _make_finding(sub_code="TC-07b", area="A10", severity="🟢 Low"),
        _make_finding(sub_code="TC-07c", area="A10", severity="🟢 Low"),
    ]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["サマリー"]
    assert ws.cell(16, 1).value == "TC-07"   # A列: TC コード
    assert ws.cell(16, 4).value == 1          # D列: 重大
    assert ws.cell(16, 5).value == 0          # E列: 要注意
    assert ws.cell(16, 6).value == 2          # F列: 要確認
    assert ws.cell(16, 7).value == 3          # G列: 合計


def test_summary_tc_matrix_header_row9(tmp_path):
    """サマリー Row 9 に TC 別集計ヘッダーが入る。"""
    from skills.export.excel_report.exporter import export_to_excel
    output = tmp_path / "out.xlsx"
    export_to_excel([], output)
    wb = load_workbook(output)
    ws = wb["サマリー"]
    assert ws.cell(9, 1).value == "項目"
    assert ws.cell(9, 7).value == "合計"


def test_header_fill_is_dark_blue(tmp_path):
    """詳細シートのヘッダー行（Row 3）が濃紺（2F5496）で塗られる。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    rgb = str(ws.cell(3, 1).fill.start_color.rgb).upper()
    assert "2F5496" in rgb


# ─────────────────────────────────────────────────────────────────────
# H. Phase 6.11a v2 追加テスト（6件）
# ─────────────────────────────────────────────────────────────────────

def test_detail_sheet_title_in_row1(tmp_path):
    """詳細シート Row 1 にシートタイトルが入る。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    assert ws.cell(1, 1).value == "A10 その他経費"


def test_detail_sheet_item_name_column(tmp_path):
    """親行 C 列にグループサマリー文字列が入る。子行 C 列は空欄。

    Phase 8-B: C 列は集約サマリー専用（C-β-3 form）。TC 名称は C 列の
    サマリー内ではなく、D 列「観点」に「{TC 名称}の税区分誤り」として出現。
    """
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding_with_hints(sub_code="TC-03c", area="A5")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A5 人件費"]
    # 親行 C 列: 勘定科目 — N 件・合計 ¥…（current→suggested） 形式
    parent_c = ws.cell(4, 3).value
    assert parent_c and "福利厚生費" in parent_c, \
        f"親行 C4 はサマリー: got {parent_c!r}"
    assert "件" in parent_c, f"親行 C4 に件数表示なし: {parent_c!r}"
    # 親行 D 列: TC 名称を含む観点
    parent_d = ws.cell(4, 4).value
    assert parent_d and "給与/人件費" in parent_d, \
        f"親行 D4 に TC 名称なし: {parent_d!r}"
    # 子行 C 列: 空欄
    assert ws.cell(5, 3).value in (None, ""), \
        f"子行 C5 は空欄: got {ws.cell(5, 3).value!r}"


def test_parent_row_d_e_has_group_observation_children_have_message(tmp_path):
    """親行 D/E = グループ観点、子行 D = 常に空欄、子行 E = Finding.message。

    Phase 8-C Fix v2 確定仕様:
      親行 D/E: グループ代表のカテゴリ + 総論（既存維持）
      子行 D:   常に空欄（親行 D が観点を担うため子行は冗長）
      子行 E:   Finding.message 全文
                同一 message の 2 件目以降は "同上" 圧縮（E 列のみ）
    """
    from skills.export.excel_report.exporter import export_to_excel
    findings = [
        _make_finding_with_hints(sub_code="TC-03c", area="A5",
                                 message="給与が課税仕入になっています"),
        _make_finding_with_hints(sub_code="TC-03c", area="A5",
                                 message="給与が課税仕入になっています"),
    ]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A5 人件費"]
    # 親行 (Row 4): D/E 列にグループ共通の観点・チェック結果
    parent_d = ws.cell(4, 4).value
    parent_e = ws.cell(4, 5).value
    assert parent_d and parent_d != "同上", \
        f"親行 D4 には観点テキストが入るはず: got {parent_d!r}"
    assert parent_e and parent_e != "同上", \
        f"親行 E4 にはチェック結果テキストが入るはず: got {parent_e!r}"
    # 子行 (Row 5): D は空欄、E に message 全文
    assert ws.cell(5, 4).value in (None, ""), \
        f"子行 D5 は常に空欄: got {ws.cell(5, 4).value!r}"
    assert ws.cell(5, 5).value == "給与が課税仕入になっています", \
        f"子行 E5 は message 全文: got {ws.cell(5, 5).value!r}"
    # 子行 (Row 6): 同一 message の 2 件目 → D 空欄 / E = "同上"
    assert ws.cell(6, 4).value in (None, ""), \
        f"子行 D6 は空欄 (Phase 8-C Fix v2): got {ws.cell(6, 4).value!r}"
    assert ws.cell(6, 5).value == "同上", \
        f"子行 E6 は '同上' で圧縮: got {ws.cell(6, 5).value!r}"


def test_multiple_groups_each_parent_has_own_observation(tmp_path):
    """異なる sub_code で複数グループが生成され、各親行が独自の観点を持つ。

    Phase 8-C ④: 子行 D/E には個別の message が入る（グループ境界を跨いだ
    「同上」圧縮は発生しない: prev_message はグループ内で閉じる）。
    親子交互配置で rows 4 (親1), 5 (子1), 6 (親2), 7 (子2) となる。
    """
    from skills.export.excel_report.exporter import export_to_excel
    findings = [
        _make_finding_with_hints(sub_code="TC-03a", area="A5", message="メッセージA"),
        _make_finding_with_hints(sub_code="TC-03b", area="A5", message="メッセージB"),
    ]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A5 人件費"]
    # 親行 1 (Row 4) と親行 2 (Row 6) がそれぞれ独自の観点
    parent1_d = ws.cell(4, 4).value
    parent2_d = ws.cell(6, 4).value
    assert parent1_d, f"親行 1 D4 に観点: got {parent1_d!r}"
    assert parent2_d, f"親行 2 D6 に観点: got {parent2_d!r}"
    # 子行 D/E: 各 Finding の message が入る（"同上" ではない、空欄でもない）
    assert ws.cell(5, 5).value == "メッセージA", "子行 E5 = メッセージA"
    assert ws.cell(7, 5).value == "メッセージB", "子行 E7 = メッセージB"


def test_summary_legend_in_l_m_columns(tmp_path):
    """サマリーシートの凡例が L 列（12）に配置される（テンプレート構造準拠）。"""
    from skills.export.excel_report.exporter import export_to_excel
    output = tmp_path / "out.xlsx"
    export_to_excel([], output, company_name="テスト会社")
    wb = load_workbook(output)
    ws = wb["サマリー"]
    # Row 2: L="【判定凡例】"（テンプレートでは Row 2 に配置）
    assert ws.cell(2, 12).value == "【判定凡例】"
    # Row 3: L="重大", Row 4: L="要注意", Row 5: L="要確認"
    assert ws.cell(3, 12).value == "重大"
    assert ws.cell(4, 12).value == "要注意"
    assert ws.cell(5, 12).value == "要確認"


def test_severity_display_low_is_youkakunin(tmp_path):
    """🟢 Low の親行 A 列表示が「要確認」になる。子行 A 列は空欄。

    Phase 8-B: 親行 A 列に severity ラベル表示、子行 A 列は空欄。
    """
    from skills.export.excel_report.exporter import export_to_excel
    findings = [_make_finding(sub_code="TC-07a", area="A10", severity="🟢 Low")]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    # 親行 A 列: 要確認
    assert ws.cell(4, 1).value == "要確認"
    # 子行 A 列: 空欄
    assert ws.cell(5, 1).value in (None, ""), \
        f"子行 A5 は空欄: got {ws.cell(5, 1).value!r}"


# ─────────────────────────────────────────────────────────────────────
# I. Phase 6.12 テンプレート駆動テスト（3件）
# ─────────────────────────────────────────────────────────────────────

def test_sankou_sheet_preserved(tmp_path):
    """参考シートがテンプレートからそのままコピーされる。"""
    from skills.export.excel_report.exporter import export_to_excel
    output = tmp_path / "out.xlsx"
    export_to_excel([], output)
    wb = load_workbook(output)
    assert "参考" in wb.sheetnames
    ws = wb["参考"]
    # 参考シートに TC-01 などの参照データが残っていること
    values = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)]
    assert any(v and "TC-01" in str(v) for v in values)


def test_custom_template_path_works(tmp_path):
    """template_path 引数でカスタムテンプレートを指定できる。"""
    from skills.export.excel_report.exporter import export_to_excel
    from skills.export.excel_report.template_engine import DEFAULT_TEMPLATE_PATH
    output = tmp_path / "out.xlsx"
    # デフォルトテンプレートを明示的に渡しても同じ結果
    result = export_to_excel([], output, template_path=DEFAULT_TEMPLATE_PATH)
    assert result == output
    wb = load_workbook(output)
    assert "サマリー" in wb.sheetnames


def test_tc_code_no_trailing_spaces(tmp_path):
    """サマリーシートの TC コード列（A列）に末尾スペースがない。"""
    from skills.export.excel_report.exporter import export_to_excel
    output = tmp_path / "out.xlsx"
    export_to_excel([], output)
    wb = load_workbook(output)
    ws = wb["サマリー"]
    for row in range(10, 17):   # TC-01〜TC-07 の行
        val = ws.cell(row, 1).value
        if val:
            assert val == val.strip(), f"Row {row}: '{val}' に末尾スペースあり"


# ─────────────────────────────────────────────────────────────────────
# J. Phase 6.12a スタイル保持テスト（2件）
# ─────────────────────────────────────────────────────────────────────

def test_area_sheet_preserves_template_font(tmp_path):
    """エリアシート全行（親行・子行）で Named Style 由来のフォントが維持される。

    Phase 8-B: 同 sub_code 3 Finding → 1 group → 4 rows（parent + 3 children）。
    Named Style parent_row_style_* / child_row_style が Meiryo UI 10pt で定義されている
    ため、全行でフォント継承を検証する。
    """
    from skills.export.excel_report.exporter import export_to_excel
    # 同 sub_code 3 件 → 1 group → 親 1 + 子 3 = rows 4-7
    findings = [
        _make_finding_with_hints(sub_code="TC-07a", area="A10", sort_priority=12),
        _make_finding_with_hints(sub_code="TC-07a", area="A10", sort_priority=12),
        _make_finding_with_hints(sub_code="TC-07a", area="A10", sort_priority=12),
    ]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    for data_row in range(4, 8):   # rows 4-7（親 1 + 子 3）
        for col in (1, 5, 14, 23):   # 代表列：優先度/チェック結果/摘要/walletTxnId
            cell = ws.cell(data_row, col)
            assert cell.font.name == "Meiryo UI", \
                f"row {data_row} col {col}: font={cell.font.name!r} (expected 'Meiryo UI')"
            assert cell.font.size == 10.0, \
                f"row {data_row} col {col}: size={cell.font.size} (expected 10.0)"
            assert cell.alignment.vertical == "center", \
                f"row {data_row} col {col}: v_align={cell.alignment.vertical!r} (expected 'center')"


def test_summary_sheet_unaffected_by_area_fix(tmp_path):
    """エリアシート修正後もサマリーシートが正常動作する（回帰確認）。"""
    from skills.export.excel_report.exporter import export_to_excel
    findings = [
        _make_finding(sub_code="TC-07a", area="A10", severity="🔴 High"),
        _make_finding(sub_code="TC-03c", area="A5", severity="🟢 Low"),
    ]
    output = tmp_path / "out.xlsx"
    export_to_excel(findings, output, company_name="回帰テスト株式会社", period="202512")
    wb = load_workbook(output)
    ws = wb["サマリー"]
    # タイトルに会社名が含まれる
    assert "回帰テスト株式会社" in ws.cell(1, 1).value
    # TC-07 行（Row 16）: 重大=1
    assert ws.cell(16, 4).value == 1
    # TC-03 行（Row 12）: 要確認=1
    assert ws.cell(12, 6).value == 1
    # 両エリアシートが存在する
    assert "A5 人件費" in wb.sheetnames
    assert "A10 その他経費" in wb.sheetnames


# ─────────────────────────────────────────────────────────────────────
# K. Phase 6.11b 金額列O/Pテスト（3件）
# ─────────────────────────────────────────────────────────────────────

def test_debit_amount_written_to_col_O(tmp_path):
    """借方金額が親行 O 列に集計値、子行 O 列に個別値として書き込まれる。

    Phase 8-B: Finding 1 件 → count=1 グループ → 親行 (Row 4) O 列 = 集計合計、
    子行 (Row 5) O 列 = 個別値。P 列はどちらも空。
    """
    from skills.export.excel_report.exporter import export_to_excel
    finding = _make_finding(sub_code="TC-03a", area="A5", debit_amount=150000)
    output = tmp_path / "out.xlsx"
    export_to_excel([finding], output)
    wb = load_workbook(output)
    ws = wb["A5 人件費"]
    # 親行 (Row 4): グループ集計合計
    parent_o = ws.cell(4, 15)
    parent_p = ws.cell(4, 16)
    assert parent_o.value == 150000, \
        f"親行 O4 (合計): expected 150000, got {parent_o.value!r}"
    assert parent_p.value in (None, ""), \
        f"親行 P4: expected empty, got {parent_p.value!r}"
    assert parent_o.number_format == "#,##0"
    # 子行 (Row 5): 個別 Finding の金額
    child_o = ws.cell(5, 15)
    child_p = ws.cell(5, 16)
    assert child_o.value == 150000, \
        f"子行 O5 (個別): expected 150000, got {child_o.value!r}"
    assert child_p.value in (None, ""), \
        f"子行 P5: expected empty, got {child_p.value!r}"
    assert child_o.number_format == "#,##0"


def test_credit_amount_written_to_col_P(tmp_path):
    """貸方金額が親行 P 列に集計値、子行 P 列に個別値として書き込まれる。

    Phase 8-B: 親子構造での貸方集計を検証。O 列はどちらも空。
    """
    from skills.export.excel_report.exporter import export_to_excel
    finding = _make_finding(sub_code="TC-05a", area="A11", credit_amount=50000)
    output = tmp_path / "out.xlsx"
    export_to_excel([finding], output)
    wb = load_workbook(output)
    sheet_name = next(n for n in wb.sheetnames if n.startswith("A11"))
    ws = wb[sheet_name]
    # 親行: 集計
    parent_o = ws.cell(4, 15)
    parent_p = ws.cell(4, 16)
    assert parent_o.value in (None, ""), f"親行 O4: expected empty, got {parent_o.value!r}"
    assert parent_p.value == 50000, f"親行 P4: expected 50000, got {parent_p.value!r}"
    assert parent_p.number_format == "#,##0"
    # 子行: 個別
    child_o = ws.cell(5, 15)
    child_p = ws.cell(5, 16)
    assert child_o.value in (None, ""), f"子行 O5: expected empty, got {child_o.value!r}"
    assert child_p.value == 50000, f"子行 P5: expected 50000, got {child_p.value!r}"
    assert child_p.number_format == "#,##0"


def test_amount_none_writes_empty_cells(tmp_path):
    """debit/credit が None のとき、親行・子行とも O/P 列は空欄。"""
    from skills.export.excel_report.exporter import export_to_excel
    finding = _make_finding(sub_code="TC-07a", area="A10")  # debit/credit=None
    output = tmp_path / "out.xlsx"
    export_to_excel([finding], output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    # 親行
    assert ws.cell(4, 15).value in (None, ""), "親行 O4 は空欄"
    assert ws.cell(4, 16).value in (None, ""), "親行 P4 は空欄"
    # 子行
    assert ws.cell(5, 15).value in (None, ""), "子行 O5 は空欄"
    assert ws.cell(5, 16).value in (None, ""), "子行 P5 は空欄"


# ─────────────────────────────────────────────────────────────────────
# L. 対象月フォーマットテスト（2件）
# ─────────────────────────────────────────────────────────────────────

def test_format_target_month_standard():
    """'YYYYMM' 文字列を '年月' 表記に正しく変換する。"""
    from skills.export.excel_report.template_engine import format_target_month
    assert format_target_month("202512") == "2025年12月"
    assert format_target_month("202504") == "2025年4月"   # 月の先頭ゼロは除く
    assert format_target_month("202601") == "2026年1月"


def test_format_target_month_fallback():
    """不正な入力はそのまま返す（防御的実装の確認）。"""
    from skills.export.excel_report.template_engine import format_target_month
    assert format_target_month("invalid") == "invalid"
    assert format_target_month("2025-12") == "2025-12"   # ハイフン区切りは変換しない
    assert format_target_month("") == ""


# ─────────────────────────────────────────────────────────────────────
# Section G. Phase 8-C γ 案: 累計期間表示テスト（2件）
# ─────────────────────────────────────────────────────────────────────
# 判断2 確定: 集計ロジックは Phase 8-D、表示層のみ先行拡張。
# 累計モード発動は 2 引数呼び出し（period_end 指定）。単月モードは既存互換。

def test_format_target_month_cumulative():
    """2 引数呼び出しで累計期間表示「YYYY年M月〜YYYY年M月」を生成する。"""
    from datetime import date as _d
    from skills.export.excel_report.template_engine import format_target_month
    # ISO 文字列
    assert format_target_month("2025-04-01", "2025-12-31") == "2025年4月〜2025年12月"
    # date オブジェクト
    assert format_target_month(_d(2025, 4, 1), _d(2025, 12, 31)) == "2025年4月〜2025年12月"
    # YYYYMM 文字列（start のみ 'YYYYMM' 形式）
    assert format_target_month("202504", "2025-12-31") == "2025年4月〜2025年12月"
    # 年跨ぎ
    assert format_target_month("2025-04-01", "2026-03-31") == "2025年4月〜2026年3月"


def test_format_target_month_cumulative_same_month():
    """期首=期末が同一月の場合は単月表示にフォールバックする。"""
    from datetime import date as _d
    from skills.export.excel_report.template_engine import format_target_month
    assert format_target_month("2025-12-01", "2025-12-31") == "2025年12月"
    assert format_target_month(_d(2025, 12, 1), _d(2025, 12, 31)) == "2025年12月"


# ─────────────────────────────────────────────────────────────────────
# M. Phase 7 Q/R 列ハイパーリンクテスト（4件）
# ─────────────────────────────────────────────────────────────────────

def _make_link_hints_for_test(
    target="general_ledger",
    account_name="支払手数料",
    period_start=date(2025, 12, 1),
    period_end=date(2025, 12, 31),
    fiscal_year_id="9842248",
    company_id="3525430",
    deal_id=None,
):
    """Section M 用 LinkHints ファクトリ。"""
    import importlib.util
    if "schema" not in sys.modules:
        _root = Path(__file__).parent.parent.parent
        _schema_path = (
            _root / "skills" / "verify" / "V1-3-rule"
            / "check-tax-classification" / "schema.py"
        )
        spec = importlib.util.spec_from_file_location("schema", _schema_path)
        mod = importlib.util.module_from_spec(spec)
        sys.modules["schema"] = mod
        spec.loader.exec_module(mod)
    schema = sys.modules["schema"]
    return schema.LinkHints(
        target=target,
        account_name=account_name,
        period_start=period_start,
        period_end=period_end,
        fiscal_year_id=fiscal_year_id,
        company_id=company_id,
        deal_id=deal_id,
    )


def test_q_column_has_gl_hyperlink(tmp_path):
    """親行・子行それぞれ Q 列に総勘定元帳ハイパーリンクが設定される。

    Phase 8-B: 子行のみ。
    Phase 8-C ② 以降: 親行にもグループ全体の GL リンクを付与する。
        - 親行 Q: 値 "GL"、URL は link_hints ベース期間 (ctx 未指定時のフォールバック)
        - 子行 Q: 値 "🔗"、URL は個別 Finding の link_hints 期間
    """
    import urllib.parse
    from skills.export.excel_report.exporter import export_to_excel

    lh = _make_link_hints_for_test()
    finding = _make_finding(sub_code="TC-03a", area="A5", link_hints=lh)
    output = tmp_path / "out.xlsx"
    export_to_excel([finding], output)
    wb = load_workbook(output)
    ws = wb["A5 人件費"]
    # 親行 Q 列: "GL" ラベル + ハイパーリンク (Phase 8-C ②)
    parent_q = ws.cell(4, 17)
    assert parent_q.value == "GL", f"親行 Q4: expected 'GL', got {parent_q.value!r}"
    assert parent_q.hyperlink is not None, "親行 Q4 に GL ハイパーリンクが付く"
    parent_url = parent_q.hyperlink.target
    assert "general_ledgers/show" in parent_url
    parent_parsed = urllib.parse.parse_qs(
        urllib.parse.urlparse(parent_url).query
    )
    assert parent_parsed["name"] == ["支払手数料"]
    # 子行 Q 列に GL ハイパーリンク (Phase 7 以前から維持)
    child_q = ws.cell(5, 17)
    assert child_q.value == "🔗", f"子行 Q5: expected '🔗', got {child_q.value!r}"
    assert child_q.hyperlink is not None, "子行 Q5 should have a hyperlink"
    url = child_q.hyperlink.target
    assert "general_ledgers/show" in url, \
        f"子行 Q5 URL should contain 'general_ledgers/show': {url}"
    parsed = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
    assert parsed["name"] == ["支払手数料"]
    assert parsed["start_date"] == ["2025-12-01"]
    assert parsed["company_id"] == ["3525430"]


def test_r_column_has_jnl_hyperlink_with_deal_id(tmp_path):
    """子行 R 列に deal_id ピンポイント仕訳帳 URL が設定される。親行 R 列は空欄。"""
    import urllib.parse
    from skills.export.excel_report.exporter import export_to_excel

    lh = _make_link_hints_for_test(deal_id="2730330344")
    finding = _make_finding(sub_code="TC-03a", area="A5", link_hints=lh)
    output = tmp_path / "out.xlsx"
    export_to_excel([finding], output)
    wb = load_workbook(output)
    ws = wb["A5 人件費"]
    # 親行 R 列は空欄
    parent_r = ws.cell(4, 18)
    assert parent_r.value in (None, ""), \
        f"親行 R4 は空欄: got {parent_r.value!r}"
    assert parent_r.hyperlink is None, "親行 R4 にハイパーリンクは付かない"
    # 子行 R 列に jnl ピンポイント URL
    child_r = ws.cell(5, 18)
    assert child_r.value == "🔗", f"子行 R5: expected '🔗', got {child_r.value!r}"
    assert child_r.hyperlink is not None, "子行 R5 should have a hyperlink"
    url = child_r.hyperlink.target
    assert "/reports/journals" in url
    parsed = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
    assert parsed["deal_id"] == ["2730330344"]
    assert "start_date" not in parsed
    assert "name" not in parsed


def test_q_r_no_hyperlink_when_link_hints_none(tmp_path):
    """link_hints=None の Finding では親行・子行とも Q/R 列にハイパーリンクなし。"""
    from skills.export.excel_report.exporter import export_to_excel

    finding = _make_finding(sub_code="TC-07a", area="A10", link_hints=None)
    output = tmp_path / "out.xlsx"
    export_to_excel([finding], output)
    wb = load_workbook(output)
    ws = wb["A10 その他経費"]
    # 親行
    assert ws.cell(4, 17).hyperlink is None
    assert ws.cell(4, 18).hyperlink is None
    # 子行
    assert ws.cell(5, 17).hyperlink is None
    assert ws.cell(5, 18).hyperlink is None


def test_r_column_fallback_url_when_no_deal_id(tmp_path):
    """deal_id なしの場合、子行 R 列が期間ベース URL にフォールバックする。"""
    import urllib.parse
    from skills.export.excel_report.exporter import export_to_excel

    lh = _make_link_hints_for_test(deal_id=None)
    finding = _make_finding(sub_code="TC-03a", area="A5", link_hints=lh)
    output = tmp_path / "out.xlsx"
    export_to_excel([finding], output)
    wb = load_workbook(output)
    ws = wb["A5 人件費"]
    # 子行 R 列
    child_r = ws.cell(5, 18)
    assert child_r.hyperlink is not None, \
        "子行 R5: deal_id なしでも期間ベース URL でリンク設定"
    url = child_r.hyperlink.target
    parsed = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
    assert "deal_id" not in parsed
    assert parsed["start_date"] == ["2025-12-01"]
    assert parsed["end_date"] == ["2025-12-31"]


def test_r_column_uses_finding_deal_id_for_pinpoint(tmp_path):
    """finding.deal_id が設定されていれば、子行 R 列にピンポイント URL が生成される。

    背景: build_link_hints("general_ledger") は link_hints.deal_id を設定しないため、
    Finding.deal_id を template_engine が直接参照してピンポイント URL を生成する必要がある。
    Phase 7 の主要 UX 目的「ワンクリックで該当取引を開ける」の実現に不可欠。
    """
    import urllib.parse
    from skills.export.excel_report.exporter import export_to_excel

    lh = _make_link_hints_for_test(deal_id=None)
    finding = _make_finding(
        sub_code="TC-03a", area="A5",
        link_hints=lh,
        deal_id="3237332503",
    )
    output = tmp_path / "out.xlsx"
    export_to_excel([finding], output)
    wb = load_workbook(output)
    ws = wb["A5 人件費"]
    child_r = ws.cell(5, 18)
    assert child_r.value == "🔗"
    assert child_r.hyperlink is not None
    url = child_r.hyperlink.target
    parsed = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
    assert parsed.get("deal_id") == ["3237332503"], \
        f"finding.deal_id should produce pinpoint URL, got: {url}"
    assert "start_date" not in parsed
