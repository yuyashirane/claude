"""Step 3-C: exporter 改修 (C-1〜C-3) の検証テスト。

§7.1 の 7 グループに対応する新規テスト群:

    Group 1. C-1: J〜N 列が ctx.transactions から逆引きで埋まる
    Group 2. C-1: ctx=None / wallet_txn_id 未一致時の安全フォールバック
    Group 3. C-2: ctx.period_* が GL リンクに反映される (新ふるまい)
    Group 4. C-2: _sort_key の安定化 (transaction_date + deal_id)
    Group 5. C-3: _is_master_value ヘルパーの判定
    Group 6. C-3: _parent_row_check_result の 4 分岐
    Group 7. C-3: _parent_row_summary の 3 分岐

配置: tests/unit/test_step3c_exporter.py
依存: test_template_engine_phase8b.py の _mk_finding / _MockCtx / _load_schema を再利用。
"""
from __future__ import annotations

import sys
from datetime import date
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

# 既存テストファイルから共通ビルダーを取り込む
from tests.unit.test_template_engine_phase8b import (
    _mk_finding,
    _MockCtx,
    _load_schema,
)


def _mk_txn(
    wallet_txn_id: str,
    *,
    partner: str = "",
    item: str | None = None,
    memo_tag: str | None = None,
    description: str = "",
    account: str = "給与手当",
    tax_label: str = "課対仕入10%",
    transaction_date: date = date(2025, 12, 1),
    deal_id: str | None = None,
):
    schema = _load_schema()
    return schema.TransactionRow(
        wallet_txn_id=wallet_txn_id,
        deal_id=deal_id,
        transaction_date=transaction_date,
        account=account,
        tax_label=tax_label,
        partner=partner,
        description=description,
        debit_amount=Decimal("100000"),
        credit_amount=Decimal("0"),
        item=item,
        memo_tag=memo_tag,
    )


def _mk_ctx(transactions, *, period_start=date(2025, 4, 1), period_end=date(2026, 3, 31)):
    schema = _load_schema()
    return schema.CheckContext(
        company_id="3525430",
        fiscal_year_id="9842248",
        period_start=period_start,
        period_end=period_end,
        transactions=transactions,
        tax_code_master={
            "課対仕入10%": "136",
            "非課仕入": "21",
            "対象外": "2",
        },
    )


# ═════════════════════════════════════════════════════════════════════
# Group 1. C-1: J〜N 列が txn_index から埋まる
# ═════════════════════════════════════════════════════════════════════

class TestC1ChildRowJtoNFromTxnIndex:
    """C-1: 子行 J/K/L/M/N が ctx.transactions から逆引きで埋まる。

    L (DEPT) は schema に部門フィールドがないため常に空欄。
    """

    def test_partner_filled_from_txn(self, tmp_path):
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(wallet_txn_id="W1")
        ctx = _mk_ctx([_mk_txn("W1", partner="株式会社サンプル")])
        out = tmp_path / "out.xlsx"
        export_to_excel([f], out, ctx=ctx)
        ws = load_workbook(out)["A5 人件費"]
        # J 列 (取引先)
        assert ws.cell(5, 10).value == "株式会社サンプル"

    def test_item_memo_desc_filled_from_txn(self, tmp_path):
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(wallet_txn_id="W2")
        ctx = _mk_ctx([
            _mk_txn(
                "W2",
                partner="A社",
                item="品目X",
                memo_tag="メモY",
                description="摘要Z",
            ),
        ])
        out = tmp_path / "out.xlsx"
        export_to_excel([f], out, ctx=ctx)
        ws = load_workbook(out)["A5 人件費"]
        assert ws.cell(5, 10).value == "A社"     # J 取引先
        assert ws.cell(5, 11).value == "品目X"    # K 品目
        assert ws.cell(5, 12).value in (None, "")  # L 部門 (schema 未定義 → 空)
        assert ws.cell(5, 13).value == "メモY"    # M メモ
        assert ws.cell(5, 14).value == "摘要Z"    # N 摘要

    def test_dept_always_empty(self, tmp_path):
        """schema に部門フィールドが存在しないため L 列は常に空欄。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(wallet_txn_id="W3")
        ctx = _mk_ctx([_mk_txn("W3", partner="X", item="Y", description="Z")])
        out = tmp_path / "out.xlsx"
        export_to_excel([f], out, ctx=ctx)
        ws = load_workbook(out)["A5 人件費"]
        assert ws.cell(5, 12).value in (None, "")


# ═════════════════════════════════════════════════════════════════════
# Group 2. C-1: ctx=None / wallet_txn_id 未一致時のフォールバック
# ═════════════════════════════════════════════════════════════════════

class TestC1ChildRowFallback:
    """C-1: txn_index が空 / 該当なしの場合は J〜N が空欄のまま。"""

    def test_no_ctx_keeps_jn_empty(self, tmp_path):
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(wallet_txn_id="W")
        out = tmp_path / "out.xlsx"
        export_to_excel([f], out)  # ctx 省略
        ws = load_workbook(out)["A5 人件費"]
        for col in (10, 11, 12, 13, 14):
            assert ws.cell(5, col).value in (None, ""), f"col={col} should be empty"

    def test_unmatched_wallet_txn_id_keeps_jn_empty(self, tmp_path):
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(wallet_txn_id="MISSING")
        ctx = _mk_ctx([_mk_txn("OTHER", partner="無関係", description="無関係")])
        out = tmp_path / "out.xlsx"
        export_to_excel([f], out, ctx=ctx)
        ws = load_workbook(out)["A5 人件費"]
        for col in (10, 11, 12, 13, 14):
            assert ws.cell(5, col).value in (None, "")


# ═════════════════════════════════════════════════════════════════════
# Group 3. C-2: GL リンク期間が ctx 期間に変わる
# (詳細は test_template_engine_phase8b.py の TestParentRowGlHyperlink で書き換え済)
# ═════════════════════════════════════════════════════════════════════

class TestC2GlLinkPeriod:

    def test_parent_uses_ctx_period_child_uses_link_hints(self, tmp_path):
        """親行は ctx.period_* (会計期間)、子行は link_hints (単月) を使う。"""
        import urllib.parse
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(
            period_start=date(2025, 12, 1),
            period_end=date(2025, 12, 31),
        )
        ctx = _MockCtx(
            period_start=date(2025, 4, 1),
            period_end=date(2026, 3, 31),
        )
        out = tmp_path / "out.xlsx"
        export_to_excel([f], out, ctx=ctx)
        ws = load_workbook(out)["A5 人件費"]
        parent_q = urllib.parse.parse_qs(
            urllib.parse.urlparse(ws.cell(4, 17).hyperlink.target).query
        )
        child_q = urllib.parse.parse_qs(
            urllib.parse.urlparse(ws.cell(5, 17).hyperlink.target).query
        )
        assert parent_q["start_date"] == ["2025-04-01"]
        assert parent_q["end_date"]   == ["2026-03-31"]
        assert child_q["start_date"]  == ["2025-12-01"]
        assert child_q["end_date"]    == ["2025-12-31"]


# ═════════════════════════════════════════════════════════════════════
# Group 4. C-2: _sort_key の安定化
# ═════════════════════════════════════════════════════════════════════

class TestC2SortStability:
    """C-2: 同一 (tc_code, sort_priority) 内では transaction_date → deal_id 順。"""

    def test_sort_by_period_start_then_deal_id(self):
        from skills.export.excel_report.template_engine import _sort_findings
        # 同じ tc_code/sort_priority だが period_start と deal_id が異なる Finding
        f_late_a = _mk_finding(
            wallet_txn_id="A",
            period_start=date(2026, 1, 15), period_end=date(2026, 1, 31),
            deal_id="d100",
        )
        f_early_b = _mk_finding(
            wallet_txn_id="B",
            period_start=date(2025, 6, 1), period_end=date(2025, 6, 30),
            deal_id="d999",
        )
        f_early_a = _mk_finding(
            wallet_txn_id="C",
            period_start=date(2025, 6, 1), period_end=date(2025, 6, 30),
            deal_id="d100",
        )
        sorted_list = _sort_findings([f_late_a, f_early_b, f_early_a])
        # 期待順: 早い日付 (deal_id d100) → 早い日付 (deal_id d999) → 遅い日付
        assert [getattr(f, "wallet_txn_id") for f in sorted_list] == ["C", "B", "A"]

    def test_sort_key_handles_missing_period(self):
        """link_hints.period_start が None でも比較が壊れないこと。"""
        from skills.export.excel_report.template_engine import _sort_key
        f1 = _mk_finding(wallet_txn_id="A", period_start=None, period_end=None)
        f2 = _mk_finding(wallet_txn_id="B")
        # tuple 比較で例外が起きない
        _sort_key(f1)
        _sort_key(f2)


# ═════════════════════════════════════════════════════════════════════
# Group 5. C-3: _is_master_value ヘルパー
# ═════════════════════════════════════════════════════════════════════

class TestC3IsMasterValue:

    @pytest.mark.parametrize("v", [
        "課税売上10%", "課税売上8%(軽)", "輸出売上", "非課売上",
        "課対仕入10%", "課対仕入8%(軽)", "非課仕入", "対象外",
    ])
    def test_master_values_recognized(self, v):
        from skills.export.excel_report.template_engine import _is_master_value
        assert _is_master_value(v) is True

    @pytest.mark.parametrize("v", [
        "", None, "要判断", "課税仕入の可能性", "対象外の可能性",
    ])
    def test_non_master_values_rejected(self, v):
        from skills.export.excel_report.template_engine import _is_master_value
        assert _is_master_value(v) is False


# ═════════════════════════════════════════════════════════════════════
# Group 6. C-3: _parent_row_check_result の 4 分岐
# ═════════════════════════════════════════════════════════════════════

def _group_from(findings):
    """FindingGroup を 1 件分組み立てる（テスト用）。"""
    from skills._common.lib.finding_grouper import group as group_fn
    groups = group_fn(findings)
    assert len(groups) == 1
    return groups[0]


class TestC3ParentRowCheckResult:

    def test_pattern_b_variants_ge2_uses_mixing_message(self):
        """Pattern B かつ variants≥2 → "N 種類の税区分混在 — 勘定科目のルール確認要"。"""
        from skills.export.excel_report.template_engine import _parent_row_check_result
        findings = [
            _mk_finding(sub_code="TC-06a", area="A12", account_name="租税公課",
                        current_value="課対仕入10%", suggested_value="対象外", wallet_txn_id="A"),
            _mk_finding(sub_code="TC-06a", area="A12", account_name="租税公課",
                        current_value="非課仕入", suggested_value="対象外", wallet_txn_id="B"),
        ]
        result = _parent_row_check_result(_group_from(findings))
        assert "2 種類の税区分混在" in result
        assert "勘定科目のルール確認要" in result

    def test_pattern_b_variants_eq1_falls_back_to_current_only(self):
        """Pattern B でも variants=1 (同一 current) → 「現状「X」— 勘定科目のルール確認要」。"""
        from skills.export.excel_report.template_engine import _parent_row_check_result
        findings = [
            _mk_finding(sub_code="TC-06a", area="A12", account_name="租税公課",
                        current_value="課対仕入10%", suggested_value="対象外", wallet_txn_id="A"),
            _mk_finding(sub_code="TC-06a", area="A12", account_name="租税公課",
                        current_value="課対仕入10%", suggested_value="対象外", wallet_txn_id="B"),
        ]
        result = _parent_row_check_result(_group_from(findings))
        assert "種類の税区分混在" not in result, f"variants=1 で混在表現は禁止: {result!r}"
        assert "現状「課対仕入10%」" in result
        assert "勘定科目のルール確認要" in result

    def test_pattern_a_master_suggested_uses_arrow_message(self):
        """Pattern A かつ sug がマスタ値 → "N 件を「cur」→「sug」へ修正要確認"。"""
        from skills.export.excel_report.template_engine import _parent_row_check_result
        findings = [
            _mk_finding(sub_code="TC-03a", area="A5", account_name="給与手当",
                        current_value="課対仕入10%", suggested_value="対象外"),
        ]
        result = _parent_row_check_result(_group_from(findings))
        assert "1 件を「課対仕入10%」→「対象外」へ修正要確認" == result

    def test_pattern_a_empty_suggested_directs_to_child_message(self):
        """Pattern A かつ sug が空 → "N 件: 現状「cur」— 詳細は子行 E 列(メッセージ)を確認"。

        旧実装の「→「-」へ修正要確認」(or "-" フォールバック) は出ない。
        """
        from skills.export.excel_report.template_engine import _parent_row_check_result
        findings = [
            _mk_finding(sub_code="TC-04e", area="A8", account_name="雑収入",
                        current_value="課税売上10%", suggested_value=""),
        ]
        result = _parent_row_check_result(_group_from(findings))
        assert "→「-」" not in result, f"旧フォールバック復活禁止: {result!r}"
        assert "現状「課税売上10%」" in result
        assert "子行" in result and "E 列" in result


# ═════════════════════════════════════════════════════════════════════
# Group 7. C-3: _parent_row_summary の 3 分岐
# ═════════════════════════════════════════════════════════════════════

class TestC3ParentRowSummary:

    def test_pattern_b_variants_ge2_uses_mixing_tail(self):
        from skills.export.excel_report.template_engine import _parent_row_summary
        findings = [
            _mk_finding(sub_code="TC-06a", area="A12", account_name="租税公課",
                        current_value="課対仕入10%", suggested_value="対象外", wallet_txn_id="A"),
            _mk_finding(sub_code="TC-06a", area="A12", account_name="租税公課",
                        current_value="非課仕入", suggested_value="対象外", wallet_txn_id="B"),
        ]
        result = _parent_row_summary(_group_from(findings))
        assert "（税区分混在）" in result

    def test_pattern_a_master_suggested_uses_arrow_tail(self):
        from skills.export.excel_report.template_engine import _parent_row_summary
        findings = [
            _mk_finding(sub_code="TC-03a", area="A5", account_name="給与手当",
                        current_value="課対仕入10%", suggested_value="対象外"),
        ]
        result = _parent_row_summary(_group_from(findings))
        assert "（課対仕入10%→対象外）" in result

    def test_pattern_a_empty_suggested_uses_yohandan_tail(self):
        from skills.export.excel_report.template_engine import _parent_row_summary
        findings = [
            _mk_finding(sub_code="TC-04e", area="A8", account_name="雑収入",
                        current_value="課税売上10%", suggested_value=""),
        ]
        result = _parent_row_summary(_group_from(findings))
        assert "（課税売上10%→要判断）" in result
