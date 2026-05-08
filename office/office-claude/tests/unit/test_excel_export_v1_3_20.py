"""V1-3-20 (TC-INV カテゴリ) の Excel 出力テスト.

E5-4 Phase 1b-ii で新設。Phase 1b-i で配線した:
    build_output → _group_v1_3_20_findings → adapt_invoice_groups
    → _fill_detail_sheet_with_groups → _write_parent_row / _write_child_row
の経路をカバーする。

テスト分類:
    - 単体: _group_v1_3_20_findings を直接呼ぶ (5 件)
    - 結合: build_output を呼んで A14 シートを openpyxl で読み戻す (3 件)

Phase 2 / 3 未実装 (子行 partner / format_invoice_memo 反映、親行 V1-3-20
対応) に依存するアサートは含めない。Phase 2 / 3 完了後に拡張可能。
"""
from __future__ import annotations

from datetime import date

import pytest
from openpyxl import load_workbook

from skills._common.schema import Finding, LinkHints
from skills.export.excel_report.template_engine import (
    DEFAULT_TEMPLATE_PATH,
    _SUM_TC_COL_CODE,
    _SUM_TC_COL_HIGH,
    _SUM_TC_COL_INFO,
    _SUM_TC_COL_LOW,
    _SUM_TC_COL_MED,
    _SUM_TC_COL_TOTAL,
    _SUM_TC_START_ROW,
    _fill_summary,
    _group_v1_3_20_findings,
    _write_child_row,
    build_output,
)


# 列定数 (template_engine の _D_DATE / _D_PARTNER / _D_MEMO に対応)
_COL_DATE = 8     # H
_COL_PARTNER = 10  # J
_COL_MEMO = 13    # M


@pytest.fixture
def v1_3_20_finding_factory():
    """V1-3-20 Finding を最小限の引数で作るファクトリ.

    classification と debit_amount だけ指定すれば残りはデフォルト。
    sub_code 等を変えたい場合は overrides で渡す。
    """
    def _make(classification, debit_amount=1000, **overrides):
        defaults = dict(
            tc_code="V1-3-20",
            sub_code="01",
            area="A14",
            sort_priority=30,
            severity="🟠 High",
            error_type="invoice_warning",
            review_level="🟠 重点確認",
            message="test message",
            classification=classification,
            debit_amount=debit_amount,
            credit_amount=0,
        )
        defaults.update(overrides)
        return Finding(**defaults)
    return _make


# ═══════════════════════════════════════════════════════════════
# 単体テスト: _group_v1_3_20_findings
# ═══════════════════════════════════════════════════════════════

class TestGroupV1320Findings:
    """_group_v1_3_20_findings の単体テスト."""

    def test_empty_input_returns_empty_list(self):
        """空入力なら空リストを返す."""
        result = _group_v1_3_20_findings([])
        assert result == []

    def test_single_classification_returns_one_group(self, v1_3_20_finding_factory):
        """同じ classification の Finding 3 件は 1 グループになる."""
        findings = [
            v1_3_20_finding_factory("qualified_but_transitional_tax", debit_amount=100),
            v1_3_20_finding_factory("qualified_but_transitional_tax", debit_amount=200),
            v1_3_20_finding_factory("qualified_but_transitional_tax", debit_amount=300),
        ]
        result = _group_v1_3_20_findings(findings)
        assert len(result) == 1
        assert result[0].count == 3
        assert result[0].total_debit == 600
        assert result[0].group_key == "V1-3-20|01|A14"
        assert len(result[0].findings) == 3

    def test_three_classifications_returns_three_groups(self, v1_3_20_finding_factory):
        """3 種類の classification の Finding は 3 グループになる.

        Phase 1b-i のスモーク確認 (python -c) と同じセットアップを
        pytest テストに昇格させたケース。
        """
        f1 = v1_3_20_finding_factory(
            "qualified_but_transitional_tax", debit_amount=1000
        )
        f2 = v1_3_20_finding_factory(
            "nonqualified_but_full_deduction_tax", debit_amount=2000, sub_code="02"
        )
        f3 = v1_3_20_finding_factory(
            "partner_unknown", debit_amount=3000, sub_code="03"
        )
        result = _group_v1_3_20_findings([f1, f2, f3])
        assert len(result) == 3
        assert result[0].group_key == "V1-3-20|01|A14"
        assert result[0].count == 1
        assert result[0].total_debit == 1000
        assert result[1].count == 1
        assert result[1].total_debit == 2000
        assert result[2].count == 1
        assert result[2].total_debit == 3000

    def test_classification_order_is_preserved(self, v1_3_20_finding_factory):
        """classification の出現順がグループ順に反映される."""
        f1 = v1_3_20_finding_factory("partner_unknown", sub_code="03")
        f2 = v1_3_20_finding_factory("qualified_but_transitional_tax", sub_code="01")
        result = _group_v1_3_20_findings([f1, f2])
        assert len(result) == 2
        assert result[0].sub_code == "03"
        assert result[1].sub_code == "01"

    def test_classification_none_falls_back_to_unknown(self, v1_3_20_finding_factory):
        """classification が None の Finding は "_unknown" グループに入る."""
        f = v1_3_20_finding_factory(None)
        result = _group_v1_3_20_findings([f])
        assert len(result) == 1
        assert result[0].count == 1


# ═══════════════════════════════════════════════════════════════
# 結合テスト: build_output 経由で A14 シート出力
# ═══════════════════════════════════════════════════════════════

class TestBuildOutputV1320:
    """build_output 経由で A14 シートに V1-3-20 Finding が出力されることを検証."""

    def test_v1_3_20_findings_produce_a14_sheet(
        self, v1_3_20_finding_factory, tmp_path
    ):
        """V1-3-20 Finding を渡すと A14 シートに行が書き込まれる."""
        findings = [
            v1_3_20_finding_factory("qualified_but_transitional_tax"),
            v1_3_20_finding_factory(
                "nonqualified_but_full_deduction_tax", sub_code="02"
            ),
        ]
        output_path = tmp_path / "test_v1_3_20.xlsx"
        build_output(findings, output_path)

        wb = load_workbook(output_path)
        a14_sheet_name = next(
            (s for s in wb.sheetnames if s.startswith("A14")),
            None,
        )
        assert a14_sheet_name is not None, (
            f"A14 シートが見つからない: {wb.sheetnames}"
        )

        ws = wb[a14_sheet_name]
        # row 1-3 はヘッダ、row 4 から実データ。最低限 row 4 にデータが
        # 書き込まれていることを確認 (親行に severity ラベル等が入る)。
        row4_has_value = any(
            ws.cell(row=4, column=c).value is not None for c in range(1, 24)
        )
        assert row4_has_value, "A14 シートの row 4 に何も書き込まれていない"

    def test_no_v1_3_20_findings_no_a14_sheet(self, tmp_path):
        """V1-3-20 Finding がない場合、A14 シートは出力されない (削除される).

        build_output は area_findings に含まれない area のシートを
        wb.remove する仕様。
        """
        output_path = tmp_path / "test_empty.xlsx"
        build_output([], output_path)

        wb = load_workbook(output_path)
        a14_sheets = [s for s in wb.sheetnames if s.startswith("A14")]
        assert len(a14_sheets) == 0, (
            f"A14 シートが想定外に残っている: {a14_sheets}"
        )

    def test_a14_parent_row_contains_classification_grouping(
        self, v1_3_20_finding_factory, tmp_path
    ):
        """A14 シートに classification 単位のグルーピングが反映されている.

        2 種類の classification を持つ Finding (3 件) を渡すと、A14 シート
        に親行 2 つ + 子行 3 つ = 計 5 行のデータが書かれることを確認する。
        """
        findings = [
            v1_3_20_finding_factory(
                "qualified_but_transitional_tax", debit_amount=500
            ),
            v1_3_20_finding_factory(
                "qualified_but_transitional_tax", debit_amount=500
            ),
            v1_3_20_finding_factory(
                "nonqualified_but_full_deduction_tax",
                debit_amount=2000,
                sub_code="02",
            ),
        ]
        output_path = tmp_path / "test_grouping.xlsx"
        build_output(findings, output_path)

        wb = load_workbook(output_path)
        a14_sheet_name = next(s for s in wb.sheetnames if s.startswith("A14"))
        ws = wb[a14_sheet_name]

        last_row = ws.max_row
        non_empty_rows = sum(
            1 for r in range(4, last_row + 1)
            if any(ws.cell(row=r, column=c).value is not None for c in range(1, 24))
        )
        # 親 2 + 子 3 = 5 行が理想。テンプレ残置等の影響を考慮し最低 5 行を確認。
        assert non_empty_rows >= 5, (
            f"想定行数以上のデータが書かれていない: {non_empty_rows}"
        )


# ═══════════════════════════════════════════════════════════════
# Phase 2 単体テスト: _write_child_row への直接属性反映
# ═══════════════════════════════════════════════════════════════

@pytest.fixture
def loaded_a14_sheet():
    """TC_template.xlsx を load し、A14 シートを返す.

    _write_child_row は NamedStyle (child_row_style) に依存するため、
    フレッシュ Workbook ではなく実テンプレを load する必要がある。
    """
    wb = load_workbook(DEFAULT_TEMPLATE_PATH)
    return wb["A14 インボイス"]


@pytest.fixture
def loaded_a4_sheet():
    """TC_template.xlsx を load し、A4 シート (V1-3-10 経路用) を返す."""
    wb = load_workbook(DEFAULT_TEMPLATE_PATH)
    return wb["A4 家賃・地代"]


class TestWriteChildRowV1320:
    """_write_child_row への V1-3-20 直接属性反映の単体テスト (Phase 2)."""

    def test_v1_3_20_child_row_partner_uses_finding_attribute(
        self, v1_3_20_finding_factory, loaded_a14_sheet
    ):
        """V1-3-20 Finding の partner 直接属性が J 列に書かれる."""
        finding = v1_3_20_finding_factory(
            "qualified_but_transitional_tax",
            partner="スモークテスト株式会社",
        )
        _write_child_row(loaded_a14_sheet, 5, finding, txn_index={})
        assert (
            loaded_a14_sheet.cell(row=5, column=_COL_PARTNER).value
            == "スモークテスト株式会社"
        )

    def test_v1_3_20_child_row_transaction_date_uses_finding_attribute(
        self, v1_3_20_finding_factory, loaded_a14_sheet
    ):
        """V1-3-20 Finding の transaction_date 直接属性が H 列に書かれる."""
        finding = v1_3_20_finding_factory(
            "qualified_but_transitional_tax",
            transaction_date="2025-12-01",
        )
        _write_child_row(loaded_a14_sheet, 5, finding, txn_index={})
        assert (
            loaded_a14_sheet.cell(row=5, column=_COL_DATE).value == "2025-12-01"
        )

    def test_v1_3_20_child_row_memo_uses_format_invoice_memo(
        self, v1_3_20_finding_factory, loaded_a14_sheet
    ):
        """V1-3-20 Finding の classification + is_qualified_invoice が M 列に整形される."""
        finding = v1_3_20_finding_factory(
            "qualified_but_transitional_tax",
            is_qualified_invoice=True,
        )
        _write_child_row(loaded_a14_sheet, 5, finding, txn_index={})
        assert (
            loaded_a14_sheet.cell(row=5, column=_COL_MEMO).value
            == "qualified_but_transitional_tax · 適格=true"
        )

    def test_v1_3_10_child_row_falls_back_to_txn_index_when_finding_attrs_none(
        self, loaded_a4_sheet
    ):
        """V1-3-10 Finding (直接属性 None + txn_index 空) は H/J/M が空文字になる.

        Phase 2 の or チェーンが既存挙動 (空文字) を保持していることの回帰防止。
        """
        finding = Finding(
            tc_code="V1-3-10",
            sub_code="TC-02a",
            severity="🟢 Low",
            review_level="🟢 参考確認",
            error_type="balance_mismatch",
            area="A4",
            sort_priority=10,
            message="V1-3-10 test",
            wallet_txn_id="t-v1310",
        )
        _write_child_row(loaded_a4_sheet, 5, finding, txn_index={})
        assert loaded_a4_sheet.cell(row=5, column=_COL_DATE).value == ""
        assert loaded_a4_sheet.cell(row=5, column=_COL_PARTNER).value == ""
        assert loaded_a4_sheet.cell(row=5, column=_COL_MEMO).value == ""

    def test_v1_3_20_child_row_falls_back_when_transaction_date_empty(
        self, v1_3_20_finding_factory, loaded_a14_sheet
    ):
        """transaction_date が空文字のとき _txn_date(finding) にフォールバックする.

        V1-3-20 checker は row.transaction_date が None のとき空文字 ""
        を Finding に入れる (checker.py:168-172)。空文字は falsy なので
        or チェーンで _txn_date(finding) (link_hints.period_start 経由)
        にフォールバックする。
        """
        finding = v1_3_20_finding_factory(
            "qualified_but_transitional_tax",
            transaction_date="",
            link_hints=LinkHints(
                target="general_ledger",
                period_start=date(2025, 12, 1),
            ),
        )
        _write_child_row(loaded_a14_sheet, 5, finding, txn_index={})
        # _txn_date の出力形式は "%Y/%m/%d" (slash 区切り)
        assert loaded_a14_sheet.cell(row=5, column=_COL_DATE).value == "2025/12/01"


# ═══════════════════════════════════════════════════════════════
# E5-5 後修正 17a: サマリー TC-INV 行への V1-3-20 集計
# ═══════════════════════════════════════════════════════════════

@pytest.fixture
def loaded_summary_sheet():
    """TC_template.xlsx を load し、サマリーシートを返す."""
    wb = load_workbook(DEFAULT_TEMPLATE_PATH)
    return wb["サマリー"]


class TestFillSummaryV1320:
    """サマリーシートの TC-INV 行への V1-3-20 Finding 集計を検証する単体テスト群."""

    def test_fill_summary_aggregates_v1_3_20_into_tc_inv_row(
        self, v1_3_20_finding_factory, loaded_summary_sheet
    ):
        """V1-3-20 Finding (tc_code='V1-3-20') が TC-INV 行に 4 区分独立で集計される.

        _TC_NAMES の末尾 (TC-INV) は _SUM_TC_START_ROW + 7 = Row 17 に書かれる.
        V1-3-20 Finding 3 件 (Critical=🔴 / High=🟠 / Low=🟢) を渡し、TODO-T 18a の
        4 区分独立集計仕様に基づき、D=1 (🔴) / E=1 (🟠) / F=0 (🟡 なし) /
        G=1 (🟢) / H=3 (合計) が TC-INV 行に出ることを assert.
        """
        findings = [
            v1_3_20_finding_factory(
                "qualified_but_transitional_tax", severity="🔴 Critical"
            ),
            v1_3_20_finding_factory(
                "qualified_but_transitional_tax", severity="🟠 High"
            ),
            v1_3_20_finding_factory(
                "qualified_but_transitional_tax", severity="🟢 Low"
            ),
        ]
        _fill_summary(
            loaded_summary_sheet,
            findings,
            company_name="テスト会社",
            period="2025-12",
            area_sheet_map={"A14": "A14 インボイス"},
            lower_row_styles=[],
        )
        # TC-INV 行 = _SUM_TC_START_ROW (10) + 7 (8 番目) = Row 17
        tc_inv_row = _SUM_TC_START_ROW + 7
        assert (
            loaded_summary_sheet.cell(tc_inv_row, _SUM_TC_COL_CODE).value == "TC-INV"
        )
        # 4 区分独立 (TODO-T 18a):
        assert loaded_summary_sheet.cell(tc_inv_row, _SUM_TC_COL_HIGH).value == 1   # D: 🔴
        assert loaded_summary_sheet.cell(tc_inv_row, _SUM_TC_COL_MED).value == 1    # E: 🟠
        assert loaded_summary_sheet.cell(tc_inv_row, _SUM_TC_COL_LOW).value == 0    # F: 🟡 なし
        assert loaded_summary_sheet.cell(tc_inv_row, _SUM_TC_COL_INFO).value == 1   # G: 🟢
        assert loaded_summary_sheet.cell(tc_inv_row, _SUM_TC_COL_TOTAL).value == 3  # H: 合計


# ═══════════════════════════════════════════════════════════════
# TODO-T 18b: 4 区分独立化の恒久検証 (回帰防止)
# ═══════════════════════════════════════════════════════════════

class TestFillSummary4SeverityIndependence:
    """4 区分独立集計の恒久検証 (TODO-T 18a で導入).

    旧仕様 (3 区分集約、🟠 + 🟡 を `_is_medium_or_orange` で E 列にまとめ書き)
    への逆戻りを検出する。新仕様では 🔴/🟠/🟡/🟢 を D/E/F/G 列に独立集計する。
    """

    def test_fill_summary_aggregates_4_severities_independently(
        self, v1_3_20_finding_factory, loaded_summary_sheet
    ):
        """🔴/🟠/🟡/🟢 を 1 件ずつ与えて D/E/F/G 列にそれぞれ独立に 1 が出る."""
        findings = [
            v1_3_20_finding_factory(
                "qualified_but_transitional_tax", severity="🔴 Critical",
                wallet_txn_id="r-1",
            ),
            v1_3_20_finding_factory(
                "qualified_but_transitional_tax", severity="🟠 High",
                wallet_txn_id="o-1",
            ),
            v1_3_20_finding_factory(
                "qualified_but_transitional_tax", severity="🟡 Med",
                wallet_txn_id="y-1",
            ),
            v1_3_20_finding_factory(
                "qualified_but_transitional_tax", severity="🟢 Low",
                wallet_txn_id="g-1",
            ),
        ]
        _fill_summary(
            loaded_summary_sheet,
            findings,
            company_name="独立テスト",
            period="2025-12",
            area_sheet_map={"A14": "A14 インボイス"},
            lower_row_styles=[],
        )
        tc_inv_row = _SUM_TC_START_ROW + 7  # Row 17
        # 4 区分が独立して D/E/F/G に 1 ずつ集計される
        # (もし 🟠 + 🟡 まとめ書きに逆戻りしたら F=0, E=2 になり fail する)
        assert loaded_summary_sheet.cell(tc_inv_row, _SUM_TC_COL_HIGH).value == 1   # D: 🔴 のみ
        assert loaded_summary_sheet.cell(tc_inv_row, _SUM_TC_COL_MED).value == 1    # E: 🟠 のみ
        assert loaded_summary_sheet.cell(tc_inv_row, _SUM_TC_COL_LOW).value == 1    # F: 🟡 のみ
        assert loaded_summary_sheet.cell(tc_inv_row, _SUM_TC_COL_INFO).value == 1   # G: 🟢 のみ
        assert loaded_summary_sheet.cell(tc_inv_row, _SUM_TC_COL_TOTAL).value == 4  # H: 合計
