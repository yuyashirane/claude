"""Phase 8-B 親子行描画の新規テスト（18 件）。

既存 test_excel_export.py は書き換え済み（カテゴリ A〜D、23 件）。本ファイルは
仕様案 Z の検証に特化した新規テストを追加する。

構成:
    Section A. 親子行描画（6 件）
    Section B. Named Style 適用（4 件）
    Section C. 仕様案 Z（子行白地）検証（2 件）
    Section D. 範囲外の不変性（3 件）
    Section E. 統合（3 件）

配置: tests/unit/test_template_engine_phase8b.py
出典: Phase 8-B Step 3 GO 指示 §Phase 3-2
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


# ─────────────────────────────────────────────────────────────────────
# 共通 Finding ビルダー（test_excel_export.py と同じロジック）
# ─────────────────────────────────────────────────────────────────────

def _load_schema():
    import importlib.util
    if "schema" not in sys.modules:
        _root = Path(__file__).parent.parent.parent
        _schema_path = (
            _root / "skills" / "verify" / "V1-3-rule"
            / "check-tax-classification" / "schema.py"
        )
        spec = importlib.util.spec_from_file_location("schema", _schema_path)
        mod = importlib.util.module_from_spec(spec)
        sys.modules["schema"] = mod
        spec.loader.exec_module(mod)
    return sys.modules["schema"]


def _mk_finding(
    sub_code: str = "TC-03a",
    area: str = "A5",
    severity: str = "🔴 High",
    current_value: str = "課対仕入10%",
    suggested_value: str = "対象外",
    account_name: str = "給与手当",
    debit_amount: int | None = 100000,
    credit_amount: int | None = None,
    wallet_txn_id: str = "txn-1",
    confidence: int = 80,
    deal_id: str | None = None,
    message: str = "テスト",
    period_start: date = date(2025, 12, 1),
    period_end: date = date(2025, 12, 31),
):
    schema = _load_schema()
    lh = schema.LinkHints(
        target="general_ledger",
        account_name=account_name,
        period_start=period_start,
        period_end=period_end,
        fiscal_year_id="9842248",
        company_id="3525430",
    )
    return schema.Finding(
        tc_code=sub_code[:5],
        sub_code=sub_code,
        severity=severity,
        error_type="direct_error",
        review_level="🔴必修",
        area=area,
        sort_priority=10,
        wallet_txn_id=wallet_txn_id,
        current_value=current_value,
        suggested_value=suggested_value,
        confidence=confidence,
        message=message,
        show_by_default=True,
        link_hints=lh,
        debit_amount=debit_amount,
        credit_amount=credit_amount,
        deal_id=deal_id,
    )


# ═════════════════════════════════════════════════════════════════════
# Section A. 親子行描画（6 件）
# ═════════════════════════════════════════════════════════════════════

class TestParentChildLayout:
    """R3 二層責務: 親行 1 + 子行 N の基本レイアウト検証。"""

    def test_single_finding_produces_one_parent_and_one_child(self, tmp_path):
        """1 Finding → count=1 のグループ → 親行 + 子行 = 2 行。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding()
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        # 親行 Row 4
        assert ws.cell(4, 2).value == "TC-03a"  # 親行 B 列に sub_code
        assert ws.cell(4, 1).style.startswith("parent_row_style_")
        # 子行 Row 5
        assert ws.cell(5, 1).style == "child_row_style"
        # Row 6 は空（データ終了）
        assert ws.cell(6, 2).value in (None, "")

    def test_three_findings_same_group_produce_one_parent_three_children(self, tmp_path):
        """同 sub_code 3 件 → 1 グループ → 親 1 + 子 3 = 4 行。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [_mk_finding() for _ in range(3)]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        ws = load_workbook(output)["A5 人件費"]
        # 親行 Row 4
        assert ws.cell(4, 1).style.startswith("parent_row_style_")
        # 子行 Row 5, 6, 7
        for r in (5, 6, 7):
            assert ws.cell(r, 1).style == "child_row_style"
        # Row 8 以降は空
        assert ws.cell(8, 2).value in (None, "")

    def test_three_different_subcodes_produce_three_parent_child_pairs(self, tmp_path):
        """異 sub_code 3 件 → 3 グループ → 親子交互に 6 行（親4/子5/親6/子7/親8/子9）。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [
            _mk_finding(sub_code="TC-07a", area="A10"),
            _mk_finding(sub_code="TC-07b", area="A10"),
            _mk_finding(sub_code="TC-07c", area="A10"),
        ]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        ws = load_workbook(output)["A10 その他経費"]
        # 親行: rows 4, 6, 8
        for r in (4, 6, 8):
            assert ws.cell(r, 1).style.startswith("parent_row_style_"), \
                f"Row {r} expected parent style, got {ws.cell(r, 1).style!r}"
        # 子行: rows 5, 7, 9
        for r in (5, 7, 9):
            assert ws.cell(r, 1).style == "child_row_style", \
                f"Row {r} expected child_row_style, got {ws.cell(r, 1).style!r}"

    def test_parent_row_c_column_contains_count_and_total(self, tmp_path):
        """親行 C 列に件数と合計金額を含むサマリーが入る（C-β-3 form）。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [_mk_finding(debit_amount=100000) for _ in range(3)]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        ws = load_workbook(output)["A5 人件費"]
        summary = ws.cell(4, 3).value
        assert summary
        assert "3 件" in summary, f"件数表示なし: {summary!r}"
        assert "300,000" in summary, f"合計金額表示なし: {summary!r}"
        assert "給与手当" in summary, f"科目名表示なし: {summary!r}"

    def test_parent_row_o_column_is_sum_of_child_debits(self, tmp_path):
        """親行 O 列 = 子 Finding 借方金額の合計。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [
            _mk_finding(debit_amount=100000),
            _mk_finding(debit_amount=200000),
            _mk_finding(debit_amount=50000),
        ]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        ws = load_workbook(output)["A5 人件費"]
        assert ws.cell(4, 15).value == 350000  # 親行 O 列 = 合計
        # 子行の個別値
        assert ws.cell(5, 15).value == 100000
        assert ws.cell(6, 15).value == 200000
        assert ws.cell(7, 15).value == 50000

    def test_child_rows_preserve_insertion_order(self, tmp_path):
        """子行は Finding の挿入順（grouper の決定的動作）を維持する。"""
        from skills.export.excel_report.exporter import export_to_excel
        # 同 sub_code だが walletTxnId で区別可能な 3 件
        findings = [
            _mk_finding(wallet_txn_id="txn-A"),
            _mk_finding(wallet_txn_id="txn-B"),
            _mk_finding(wallet_txn_id="txn-C"),
        ]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        ws = load_workbook(output)["A5 人件費"]
        # 子行 rows 5-7 の W 列が挿入順
        assert ws.cell(5, 23).value == "txn-A"
        assert ws.cell(6, 23).value == "txn-B"
        assert ws.cell(7, 23).value == "txn-C"


# ═════════════════════════════════════════════════════════════════════
# Section B. Named Style 適用（4 件）
# ═════════════════════════════════════════════════════════════════════

class TestNamedStyleApplication:
    """severity 別の親行スタイルが正しい色を持つことを検証（add_named_styles.py §3）。"""

    def test_critical_severity_uses_critical_parent_style(self, tmp_path):
        """🔴 High/Critical → parent_row_style_critical (FCEBEB)。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(severity="🔴 High")
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        cell = ws.cell(4, 1)
        assert cell.style == "parent_row_style_critical"
        rgb = str(cell.fill.start_color.rgb).upper()
        assert "FCEBEB" in rgb, f"critical bg: got {rgb!r}"

    def test_warning_severity_uses_warning_parent_style(self, tmp_path):
        """🟠 Warning → parent_row_style_warning (FAEEDA)。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(severity="🟠 Warning")
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        cell = ws.cell(4, 1)
        assert cell.style == "parent_row_style_warning"
        rgb = str(cell.fill.start_color.rgb).upper()
        assert "FAEEDA" in rgb, f"warning bg: got {rgb!r}"

    def test_medium_severity_uses_medium_parent_style(self, tmp_path):
        """🟡 Medium → parent_row_style_medium (FEF5D6)。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(severity="🟡 Medium")
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        cell = ws.cell(4, 1)
        assert cell.style == "parent_row_style_medium"
        rgb = str(cell.fill.start_color.rgb).upper()
        assert "FEF5D6" in rgb, f"medium bg: got {rgb!r}"

    def test_low_severity_uses_low_parent_style(self, tmp_path):
        """🟢 Low → parent_row_style_low (EAF3DE)。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(severity="🟢 Low")
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        cell = ws.cell(4, 1)
        assert cell.style == "parent_row_style_low"
        rgb = str(cell.fill.start_color.rgb).upper()
        assert "EAF3DE" in rgb, f"low bg: got {rgb!r}"


# ═════════════════════════════════════════════════════════════════════
# Section C. 仕様案 Z（子行白地）検証（2 件）
# ═════════════════════════════════════════════════════════════════════

class TestSpecZWhiteChildRows:
    """仕様案 Z: 子行は白地（severity_fills 重ね塗りなし）。"""

    def test_child_row_has_no_fill(self, tmp_path):
        """子行は fill_type が 'none' または 'solid' でも白地（FFFFFF）。

        child_row_style 定義は fill 未設定（NamedStyle に fill=None）なので
        fill_type は None または "none"。
        """
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(severity="🔴 High")  # 親は重大色
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        child_cell = ws.cell(5, 1)
        # child_row_style が適用されている
        assert child_cell.style == "child_row_style"
        # 子行は塗りなし（白地）
        ft = child_cell.fill.fill_type
        assert ft in (None, "none"), \
            f"子行は塗りなし想定: got fill_type={ft!r}"

    def test_child_row_severity_column_is_blank(self, tmp_path):
        """子行 A 列（severity 表示列）は空欄（仕様案 Z: 親行のみ色と emoji）。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(severity="🔴 High")
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        # 親行 A = 重大
        assert ws.cell(4, 1).value == "重大"
        # 子行 A は空
        assert ws.cell(5, 1).value in (None, "")


# ═════════════════════════════════════════════════════════════════════
# Section D. 範囲外の不変性（3 件）
# ═════════════════════════════════════════════════════════════════════

class TestOutOfScopeInvariance:
    """Phase 8-B は詳細シートのみ改修。サマリー・参考は不変。"""

    def test_summary_sheet_structure_unchanged(self, tmp_path):
        """Phase 8-B 後もサマリーシートの主要セル位置が維持される。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [_mk_finding(severity="🔴 High")]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output, company_name="テスト株式会社")
        ws = load_workbook(output)["サマリー"]
        # Row 1: タイトル
        assert "テスト株式会社" in ws.cell(1, 1).value
        # Row 9: TC 集計ヘッダー
        assert ws.cell(9, 1).value == "項目"
        # Row 12: TC-03 行（重大=1）
        assert ws.cell(12, 4).value == 1

    def test_sankou_sheet_preserved_after_phase8b(self, tmp_path):
        """参考シートが Phase 8-B 実装後もテンプレから維持される。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [_mk_finding()]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        wb = load_workbook(output)
        assert "参考" in wb.sheetnames
        ws = wb["参考"]
        # 参考シートの何らかのセルにデータがあることを確認
        assert ws.max_row > 1

    def test_empty_findings_produces_summary_only(self, tmp_path):
        """空 findings でも親子行処理は実行されず、サマリーのみ出力。"""
        from skills.export.excel_report.exporter import export_to_excel
        output = tmp_path / "out.xlsx"
        export_to_excel([], output)
        wb = load_workbook(output)
        assert "サマリー" in wb.sheetnames
        assert "参考" in wb.sheetnames
        # 詳細シートは全削除
        for a in ("A4 家賃・地代", "A5 人件費", "A10 その他経費"):
            assert a not in wb.sheetnames


# ═════════════════════════════════════════════════════════════════════
# Section E. 統合（3 件）
# ═════════════════════════════════════════════════════════════════════

class TestIntegration:
    """Pattern B 検証・マルチエリア・大規模グループの統合テスト。"""

    def test_pattern_b_tc06_produces_mixing_summary(self, tmp_path):
        """TC-06 (Pattern B 混在検知) は「税区分混在」サマリーを生成する。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [
            _mk_finding(
                sub_code="TC-06a", area="A12",
                account_name="租税公課", current_value="課対仕入10%",
                suggested_value="対象外", debit_amount=100000,
            ),
            _mk_finding(
                sub_code="TC-06a", area="A12",
                account_name="租税公課", current_value="非課仕入",
                suggested_value="対象外", debit_amount=200000,
            ),
            _mk_finding(
                sub_code="TC-06a", area="A12",
                account_name="租税公課", current_value="対象外",
                suggested_value="対象外", debit_amount=300000,
            ),
        ]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        wb = load_workbook(output)
        a12_name = next(n for n in wb.sheetnames if n.startswith("A12"))
        ws = wb[a12_name]
        # 親行 C 列に「税区分混在」
        parent_c = ws.cell(4, 3).value
        assert parent_c and "税区分混在" in parent_c, \
            f"Pattern B C サマリー: {parent_c!r}"
        # 親行 D 列: 同一科目に税区分混在
        assert ws.cell(4, 4).value == "同一科目に税区分混在"
        # 親行 E 列: N 種類の税区分混在 — ...
        parent_e = ws.cell(4, 5).value
        assert parent_e and "種類" in parent_e and "混在" in parent_e
        # 子行 3 件: rows 5, 6, 7
        for r in (5, 6, 7):
            assert ws.cell(r, 1).style == "child_row_style"

    def test_multiple_areas_each_has_parent_child_layout(self, tmp_path):
        """複数エリアの Finding が各エリアシートで親子行構造で展開される。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [
            _mk_finding(sub_code="TC-03a", area="A5", account_name="給与手当"),
            _mk_finding(sub_code="TC-07a", area="A10", account_name="福利厚生費"),
        ]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        wb = load_workbook(output)
        # A5 人件費
        ws_a5 = wb["A5 人件費"]
        assert ws_a5.cell(4, 1).style.startswith("parent_row_style_")
        assert ws_a5.cell(5, 1).style == "child_row_style"
        # A10 その他経費
        ws_a10 = wb["A10 その他経費"]
        assert ws_a10.cell(4, 1).style.startswith("parent_row_style_")
        assert ws_a10.cell(5, 1).style == "child_row_style"

    def test_large_group_five_findings_produces_one_parent_five_children(self, tmp_path):
        """5 件同一 sub_code → 1 親 + 5 子 = 6 行。count と total が正しく集計される。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [_mk_finding(debit_amount=100000) for _ in range(5)]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        ws = load_workbook(output)["A5 人件費"]
        # 親行 O 列 = 500000
        assert ws.cell(4, 15).value == 500000
        # 親行 C 列に "5 件" / "500,000"
        summary = ws.cell(4, 3).value
        assert "5 件" in summary
        assert "500,000" in summary
        # 子行 5 つ
        for r in range(5, 10):
            assert ws.cell(r, 1).style == "child_row_style"
        # Row 10 は空
        assert ws.cell(10, 2).value in (None, "")


# ─────────────────────────────────────────────────────────────────────
# Section F. 実務レビュー受スタイル修正 (2026-04-22)
# ─────────────────────────────────────────────────────────────────────
# 悠皓さんの E2E 成果物レビューで検出された視覚的改善 4 点のうち、
# 今回実装した ①③⑤⑥ を契約化するテスト。Phase 8-C で ②④ を実装する
# 際に本仕様が壊れないことを保証する。
#
#   ① 子行 F/G/H/I/J/K 中央揃え
#   ③ 親行 C/D/E wrap_text=True
#   ⑤ 親子両方の O/P 右揃え
#   ⑥ 親子両方の A/B/Q/R 中央揃え
# ─────────────────────────────────────────────────────────────────────

class TestStyleRefinements:
    """Phase 8-B 実務レビュー受スタイル修正 (2026-04-22) の契約テスト。"""

    def test_parent_row_c_d_e_has_wrap_text(self, tmp_path):
        """③ 親行 C/D/E 列は wrap_text=True (長文でも見切れないよう折り返し)。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [_mk_finding()]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        ws = load_workbook(output)["A5 人件費"]
        for col in (3, 4, 5):  # C / D / E
            align = ws.cell(4, col).alignment
            assert align.wrap_text is True, (
                f"親行 col={col} wrap_text={align.wrap_text} (期待 True)"
            )

    def test_parent_and_child_amount_columns_right_aligned(self, tmp_path):
        """⑤ 親子両方の O/P 列が right (数値列として自然)。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [_mk_finding(debit_amount=100000)]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        ws = load_workbook(output)["A5 人件費"]
        for row in (4, 5):  # parent / child
            for col in (15, 16):  # O / P
                align = ws.cell(row, col).alignment
                assert align.horizontal == "right", (
                    f"row={row} col={col} horizontal={align.horizontal!r} (期待 right)"
                )

    def test_child_row_tax_columns_center_aligned(self, tmp_path):
        """① 子行 F/G/H/I/J/K 列が center。⑥ A/B/Q/R 列も center。"""
        from skills.export.excel_report.exporter import export_to_excel
        findings = [_mk_finding()]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        ws = load_workbook(output)["A5 人件費"]
        child_row = 5
        # ① F/G/H/I/J/K = 中央
        for col in (6, 7, 8, 9, 10, 11):
            align = ws.cell(child_row, col).alignment
            assert align.horizontal == "center", (
                f"子行 col={col} horizontal={align.horizontal!r} (期待 center)"
            )
        # ⑥ A/B/Q/R = 中央 (親子両方)
        for row in (4, 5):
            for col in (1, 2, 17, 18):
                align = ws.cell(row, col).alignment
                assert align.horizontal == "center", (
                    f"row={row} col={col} horizontal={align.horizontal!r} "
                    f"(期待 center)"
                )


# ═════════════════════════════════════════════════════════════════════
# Phase 8-C 新規テスト: Section H (親行 GL リンク) / I (子行 D/E) / J (S/T)
# ═════════════════════════════════════════════════════════════════════

class _MockCtx:
    """Phase 8-C: build_group_gl_link の ctx パラメタ用 duck-typed オブジェクト。

    CheckContext の import は skill 側 (hyphen path) で複雑なため、テストでは
    必要な属性だけを持つ軽量モックを使う。build_group_gl_link は getattr 経由で
    属性参照するため本モックで十分。

    Phase 8-C Fix v2: tax_code_master を追加。
    _extract_tax_group_code が Finding.current_value → 税区分コード を逆引きする
    際に参照する。デフォルトで代表的な税区分マッピングを持たせ、テスト側で
    毎回指定する負担を減らす。
    """
    def __init__(self, period_start=None, period_end=None,
                 fiscal_year_id="9842248", company_id="3525430",
                 tax_code_master=None):
        self.period_start = period_start
        self.period_end = period_end
        self.fiscal_year_id = fiscal_year_id
        self.company_id = company_id
        self.tax_code_master = tax_code_master if tax_code_master is not None else {
            "課対仕入10%": "136",
            "課対仕入8%軽減": "137",
            "非課仕入":      "21",
            "対象外":        "2",
            "課税売上10%":   "8",
            "非課売上":      "11",
        }


# ─────────────────────────────────────────────────────────────────────
# Section H. 親行 Q 列 GL リンク（Phase 8-C ②、+4 tests）
# ─────────────────────────────────────────────────────────────────────

class TestParentRowGlHyperlink:
    """Phase 8-C ②: 親行 Q 列にグループ全体の GL リンクを追加。

    Phase 8-C Fix v2 確定仕様:
      - 期間: 単月（グループ代表 Finding の link_hints.period_start/end）
      - 税区分: グループ内の全税区分を tax_group_codes として複数指定
              (ctx.tax_code_master 経由で current_value から逆引き)
      - ラベル: "GL" 固定
    """

    def test_parent_row_has_gl_hyperlink(self, tmp_path):
        """親行 Q 列にハイパーリンクが設定される。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding()
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        parent_q = ws.cell(4, 17)
        assert parent_q.hyperlink is not None, "親行 Q4 に GL ハイパーリンクが付く"
        assert "general_ledgers/show" in parent_q.hyperlink.target

    def test_parent_row_gl_label_is_gl(self, tmp_path):
        """親行 Q 列の表示値は "GL" 固定（判断 4）。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding()
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        assert ws.cell(4, 17).value == "GL"

    def test_parent_row_gl_url_uses_single_month_period(self, tmp_path):
        """Phase 8-C Fix v2: 親行 GL URL の期間は単月 (link_hints 期間)。

        ctx.period_start/end (会計期間) は使わず、グループ代表 Finding の
        link_hints.period_start/end (取引月) を使う。
        """
        import urllib.parse
        from skills.export.excel_report.exporter import export_to_excel
        # ctx の期間と link_hints の期間を意図的に差別化
        f = _mk_finding(
            period_start=date(2025, 12, 1),
            period_end=date(2025, 12, 31),
        )
        ctx = _MockCtx(
            period_start=date(2025, 4, 1),   # 会計期首 (URL には使わない)
            period_end=date(2026, 3, 31),    # 会計期末 (URL には使わない)
        )
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output, ctx=ctx)
        ws = load_workbook(output)["A5 人件費"]
        parent_url = ws.cell(4, 17).hyperlink.target
        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(parent_url).query)
        # 親行は単月 (link_hints)。ctx の会計期間は反映されない。
        assert parsed["start_date"] == ["2025-12-01"], (
            f"親行 GL URL start_date は link_hints 単月: got {parsed.get('start_date')}"
        )
        assert parsed["end_date"] == ["2025-12-31"]
        # 子行も link_hints 期間 (単月) で同じ
        child_url = ws.cell(5, 17).hyperlink.target
        child_parsed = urllib.parse.parse_qs(urllib.parse.urlparse(child_url).query)
        assert child_parsed["start_date"] == ["2025-12-01"]

    def test_parent_row_gl_url_includes_single_tax_code(self, tmp_path):
        """グループ内の税区分が 1 種類の場合、tax_group_codes を 1 つ付与。"""
        import urllib.parse
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(current_value="課対仕入10%")
        ctx = _MockCtx()  # tax_code_master デフォルト: 課対仕入10% → 136
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output, ctx=ctx)
        ws = load_workbook(output)["A5 人件費"]
        parent_url = ws.cell(4, 17).hyperlink.target
        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(parent_url).query)
        assert parsed.get("tax_group_codes") == ["136"], (
            f"単一税区分でも tax_group_codes が 1 件付与: got {parsed.get('tax_group_codes')}"
        )

    def test_parent_row_gl_url_includes_multiple_tax_codes(self, tmp_path):
        """グループ内の税区分が複数の場合、tax_group_codes を複数付与 (ソート済)。

        freee の URL 形式: ?tax_group_codes=2&tax_group_codes=21&tax_group_codes=136
        """
        import urllib.parse
        from skills.export.excel_report.exporter import export_to_excel
        # Pattern B: 同一科目に 3 種類の税区分が混在 (TC-06 相当)
        findings = [
            _mk_finding(
                sub_code="TC-06a", area="A12",
                account_name="租税公課", current_value="課対仕入10%",
                suggested_value="対象外", wallet_txn_id="A",
            ),
            _mk_finding(
                sub_code="TC-06a", area="A12",
                account_name="租税公課", current_value="非課仕入",
                suggested_value="対象外", wallet_txn_id="B",
            ),
            _mk_finding(
                sub_code="TC-06a", area="A12",
                account_name="租税公課", current_value="対象外",
                suggested_value="対象外", wallet_txn_id="C",
            ),
        ]
        ctx = _MockCtx()  # 課対仕入10% → 136, 非課仕入 → 21, 対象外 → 2
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output, ctx=ctx)
        wb = load_workbook(output)
        a12_name = next(n for n in wb.sheetnames if n.startswith("A12"))
        ws = wb[a12_name]
        parent_url = ws.cell(4, 17).hyperlink.target
        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(parent_url).query)
        codes = parsed.get("tax_group_codes", [])
        assert set(codes) == {"2", "21", "136"}, (
            f"3 税区分すべて付与: got {codes}"
        )
        assert codes == sorted(codes), f"ソート順安定化: got {codes}"

    def test_parent_row_gl_url_omits_tax_filter_when_unavailable(self, tmp_path):
        """税区分コードが全て逆引き失敗の場合、tax_group_codes は付与されない。"""
        import urllib.parse
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(current_value="存在しない税区分ラベル")
        ctx = _MockCtx()  # tax_code_master に該当キーなし
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output, ctx=ctx)
        ws = load_workbook(output)["A5 人件費"]
        parent_url = ws.cell(4, 17).hyperlink.target
        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(parent_url).query)
        assert "tax_group_codes" not in parsed, (
            f"逆引き失敗時はフィルタなし: got {parsed}"
        )

    def test_parent_row_gl_url_uses_tax_code_master_for_lookup(self, tmp_path):
        """ctx.tax_code_master 経由で Finding.current_value → コードが解決される。"""
        import urllib.parse
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(current_value="課対仕入10%")
        # カスタムマスタで意図的に違うコードを返し、マスタ経由解決を検証
        ctx = _MockCtx(tax_code_master={"課対仕入10%": "999"})
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output, ctx=ctx)
        ws = load_workbook(output)["A5 人件費"]
        parent_url = ws.cell(4, 17).hyperlink.target
        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(parent_url).query)
        assert parsed.get("tax_group_codes") == ["999"], (
            f"tax_code_master 経由で解決: got {parsed.get('tax_group_codes')}"
        )

    def test_parent_row_gl_url_when_tax_code_master_missing_label(self, tmp_path):
        """一部の Finding のみ逆引きできる場合、解決できたコードのみで URL 生成。"""
        import urllib.parse
        from skills.export.excel_report.exporter import export_to_excel
        findings = [
            _mk_finding(
                sub_code="TC-06a", area="A12",
                account_name="租税公課", current_value="課対仕入10%",  # 解決可
                wallet_txn_id="A",
            ),
            _mk_finding(
                sub_code="TC-06a", area="A12",
                account_name="租税公課", current_value="未登録ラベル",  # 解決不可
                wallet_txn_id="B",
            ),
        ]
        ctx = _MockCtx()  # 課対仕入10% のみ解決可
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output, ctx=ctx)
        wb = load_workbook(output)
        a12_name = next(n for n in wb.sheetnames if n.startswith("A12"))
        ws = wb[a12_name]
        parent_url = ws.cell(4, 17).hyperlink.target
        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(parent_url).query)
        codes = parsed.get("tax_group_codes", [])
        assert codes == ["136"], (
            f"解決できたコードのみ URL 反映: got {codes}"
        )

    def test_parent_row_gl_url_when_ctx_has_no_tax_code_master(self, tmp_path):
        """ctx に tax_code_master 属性がない場合、フィルタなし URL (旧 ctx 互換)。"""
        import urllib.parse
        from skills.export.excel_report.exporter import export_to_excel

        class _LegacyCtx:
            """tax_code_master を持たない古い ctx の模倣。"""
            period_start = date(2025, 4, 1)
            period_end = date(2026, 3, 31)
            fiscal_year_id = "9842248"
            company_id = "3525430"

        f = _mk_finding(current_value="課対仕入10%")
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output, ctx=_LegacyCtx())
        ws = load_workbook(output)["A5 人件費"]
        parent_url = ws.cell(4, 17).hyperlink.target
        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(parent_url).query)
        assert "tax_group_codes" not in parsed, (
            f"tax_code_master なしの ctx ではフィルタなし: got {parsed}"
        )

    def test_parent_row_gl_url_multiple_codes_sorted_stable(self, tmp_path):
        """複数税区分がソート順で決定論的に安定する (同入力 → 同 URL)。"""
        import urllib.parse
        from skills.export.excel_report.exporter import export_to_excel
        findings = [
            _mk_finding(
                sub_code="TC-06a", area="A12", account_name="租税公課",
                current_value="課対仕入10%", wallet_txn_id="A",
            ),
            _mk_finding(
                sub_code="TC-06a", area="A12", account_name="租税公課",
                current_value="非課仕入", wallet_txn_id="B",
            ),
            _mk_finding(
                sub_code="TC-06a", area="A12", account_name="租税公課",
                current_value="対象外", wallet_txn_id="C",
            ),
        ]
        ctx = _MockCtx()
        urls = []
        for i in range(5):
            output = tmp_path / f"out{i}.xlsx"
            export_to_excel(findings, output, ctx=ctx)
            wb = load_workbook(output)
            a12_name = next(n for n in wb.sheetnames if n.startswith("A12"))
            ws = wb[a12_name]
            urls.append(ws.cell(4, 17).hyperlink.target)
        assert all(u == urls[0] for u in urls), (
            f"決定論性破綻: unique URLs = {set(urls)}"
        )
        parsed = urllib.parse.parse_qs(urllib.parse.urlparse(urls[0]).query)
        codes = parsed.get("tax_group_codes", [])
        assert codes == sorted(codes)

    def test_parent_row_q_column_empty_when_account_missing(self, tmp_path):
        """link_hints=None の Finding では親行 Q 列にリンクなし・空欄。"""
        from skills.export.excel_report.exporter import export_to_excel
        schema = _load_schema()
        # link_hints を手動で None に差し替えた Finding
        f = schema.Finding(
            tc_code="TC-03",
            sub_code="TC-03a",
            severity="🔴 High",
            error_type="direct_error",
            review_level="🔴必修",
            area="A5",
            sort_priority=10,
            wallet_txn_id="txn-1",
            current_value="課対仕入10%",
            suggested_value="対象外",
            confidence=80,
            message="",
            show_by_default=True,
            link_hints=None,  # ← 勘定科目が取れない
            debit_amount=100000,
        )
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        parent_q = ws.cell(4, 17)
        assert parent_q.value in (None, ""), (
            f"親行 Q4 は空欄: got {parent_q.value!r}"
        )
        assert parent_q.hyperlink is None


# ─────────────────────────────────────────────────────────────────────
# Section I. 子行 D/E 列の message 復活（Phase 8-C ④、+4 tests）
# ─────────────────────────────────────────────────────────────────────

class TestChildRowDetailRevival:
    """Phase 8-C Fix v2 確定仕様: 子行 D/E 列の役割。

    - D: 常に空欄（親行 D がグループ観点を担うため、子行では冗長）
    - E: Finding.message 全文
         ・message 空 → 空欄
         ・直前子行と同一 message → "同上" 圧縮
    """

    def test_child_row_d_column_always_empty(self, tmp_path):
        """Phase 8-C Fix v2: 子行 D 列は常に空欄（長文 message でも）。"""
        from skills.export.excel_report.exporter import export_to_excel
        long_msg = (
            "給与・賞与・役員報酬等は対象外です。労働の対価は資産の譲渡等"
            "に該当しないため、消費税の課税対象になりません。"
        )
        f = _mk_finding(message=long_msg)
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        assert ws.cell(5, 4).value in (None, ""), (
            f"子行 D5 は常に空欄: got {ws.cell(5, 4).value!r}"
        )
        # E 列には全文
        assert ws.cell(5, 5).value == long_msg

    def test_child_row_e_column_has_full_message(self, tmp_path):
        """E 列には Finding.message の全文が入る。"""
        from skills.export.excel_report.exporter import export_to_excel
        msg = "給与・賞与・役員報酬等は対象外です（労働の対価は資産の譲渡等に該当しません）。"
        f = _mk_finding(message=msg)
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        assert ws.cell(5, 5).value == msg

    def test_child_row_dou_jou_compression(self, tmp_path):
        """同一グループ内の同一 message は、2 件目以降 E 列のみ "同上" に圧縮。

        Phase 8-C Fix v2: D 列は常に空欄なので "同上" も E 列のみ。
        """
        from skills.export.excel_report.exporter import export_to_excel
        msg = "給与は対象外"
        findings = [
            _mk_finding(wallet_txn_id="A", message=msg),
            _mk_finding(wallet_txn_id="B", message=msg),  # 同一
            _mk_finding(wallet_txn_id="C", message=msg),  # 同一
        ]
        output = tmp_path / "out.xlsx"
        export_to_excel(findings, output)
        ws = load_workbook(output)["A5 人件費"]
        # Row 5 (初出): D 空欄 / E 全文
        assert ws.cell(5, 4).value in (None, ""), "子行 D5 は空欄"
        assert ws.cell(5, 5).value == msg
        # Row 6/7: D 空欄 / E = "同上"
        for r in (6, 7):
            assert ws.cell(r, 4).value in (None, ""), f"子行 D{r} は空欄"
            assert ws.cell(r, 5).value == "同上", f"子行 E{r} = 同上"

    def test_child_row_empty_message_fallback(self, tmp_path):
        """message が空の Finding では D/E 両方空欄（フォールバック）。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding(message="")
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        assert ws.cell(5, 4).value in (None, "")
        assert ws.cell(5, 5).value in (None, "")


# ─────────────────────────────────────────────────────────────────────
# Section J. S/T 列中央揃え（Phase 8-C ⑦、+2 tests）
# ─────────────────────────────────────────────────────────────────────

class TestConfidenceErrorTypeCenterAlign:
    """Phase 8-C ⑦: 確信度 (S=19) / エラー型 (T=20) 列の中央寄せ。"""

    def test_parent_row_s_t_columns_center_aligned(self, tmp_path):
        """親行 S/T 列が horizontal=center。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding()
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        for col in (19, 20):
            align = ws.cell(4, col).alignment
            assert align.horizontal == "center", (
                f"親行 col={col} horizontal={align.horizontal!r} (期待 center)"
            )

    def test_child_row_s_t_columns_center_aligned(self, tmp_path):
        """子行 S/T 列が horizontal=center。"""
        from skills.export.excel_report.exporter import export_to_excel
        f = _mk_finding()
        output = tmp_path / "out.xlsx"
        export_to_excel([f], output)
        ws = load_workbook(output)["A5 人件費"]
        for col in (19, 20):
            align = ws.cell(5, col).alignment
            assert align.horizontal == "center", (
                f"子行 col={col} horizontal={align.horizontal!r} (期待 center)"
            )
