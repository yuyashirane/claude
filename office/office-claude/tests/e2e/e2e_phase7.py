import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
from datetime import datetime
sys.path.insert(0, '.')
skill_dir = Path('skills/verify/V1-3-rule/check-tax-classification')
sys.path.insert(0, str(skill_dir))
from scripts.e2e.freee_to_context import build_check_context
import importlib.util
spec = importlib.util.spec_from_file_location('checker', skill_dir / 'checker.py')
m = importlib.util.module_from_spec(spec)
spec.loader.exec_module(m)
from skills.export.excel_report.exporter import export_to_excel
from openpyxl import load_workbook

base = Path('data/e2e/3525430/202512')
ctx = build_check_context(
    deals_path=base / 'deals_202512.json',
    partners_path=base / 'partners_all.json',
    account_items_path=base / 'account_items_all.json',
    company_info_path=base / 'company_info.json',
    taxes_codes_path=base / 'taxes_codes.json',
)
findings = m.run(ctx)
ts = datetime.now().strftime('%Y%m%d_%H%M%S')
output = base / f'e2e_report_phase7_{ts}.xlsx'
period = base.name  # "202512" — deals ファイルのディレクトリ名がそのまま対象月
export_to_excel(findings, output, company_name=ctx.company_name, period=period)

print(f'Generated: {output.resolve()}')
print(f'Size: {output.stat().st_size:,} bytes')
print(f'Findings: {len(findings)} total')

wb = load_workbook(output)
print('\n=== Q/R 列ハイパーリンク確認 ===')
for sheet_name in wb.sheetnames:
    if not (sheet_name.startswith('A') and len(sheet_name) > 1):
        continue
    ws = wb[sheet_name]
    rows_with_links = []
    for row in range(4, ws.max_row + 1):
        q = ws.cell(row=row, column=17)
        r = ws.cell(row=row, column=18)
        q_url = q.hyperlink.target if q.hyperlink else None
        r_url = r.hyperlink.target if r.hyperlink else None
        if q_url or r_url:
            rows_with_links.append((row, q_url, r_url))
    if rows_with_links:
        print(f'\n[{sheet_name}]')
        for row, q_url, r_url in rows_with_links:
            print(f'  row{row}:')
            print(f'    Q(GL)  = {q_url}')
            print(f'    R(JNL) = {r_url}')

print('\n=== サマリー確認 ===')
ws_sum = wb['\u30b5\u30de\u30ea\u30fc']
print(f'  B3 (\u4f1a\u793e\u540d): {ws_sum.cell(3,2).value}')
print(f'  B4 (\u5bfe\u8c61\u6708): {ws_sum.cell(4,2).value}')
