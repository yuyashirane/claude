from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

FONT = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color="1F4E78")
SUBTITLE_FONT = Font(name=FONT, bold=True, size=11, color="1F4E78")
NORMAL = Font(name=FONT, size=10)
BOLD = Font(name=FONT, size=10, bold=True)
RED_FONT = Font(name=FONT, size=10, color="C00000", bold=True)
ORANGE_FILL = PatternFill("solid", start_color="FFE699")
RED_FILL = PatternFill("solid", start_color="F8CBAD")
GREEN_FILL = PatternFill("solid", start_color="C6EFCE")
GRAY_FILL = PatternFill("solid", start_color="F2F2F2")

thin = Side(border_style="thin", color="999999")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws, row, cols):
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

# ============= Sheet1: サマリー =============
ws1 = wb.active
ws1.title = "サマリー"

ws1["A1"] = "株式会社コムネットシステム  2026年3月 消費税区分チェック結果"
ws1["A1"].font = TITLE_FONT
ws1.merge_cells("A1:F1")

info = [
    ("会社名", "株式会社コムネットシステム"),
    ("freee事業所ID", "1362187"),
    ("会計期間", "2025-04-01 〜 2026-03-31（3月決算）"),
    ("対象月", "2026年3月（決算月）"),
    ("経理方式", "本則課税・個別対応方式（tax_method=2）"),
    ("業種", "システム開発（IT）"),
    ("チェック実施日", "2026-04-27"),
]
r = 3
for k, v in info:
    ws1.cell(row=r, column=1, value=k).font = BOLD
    ws1.cell(row=r, column=1).fill = GRAY_FILL
    ws1.cell(row=r, column=2, value=v).font = NORMAL
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    for c in range(1, 7):
        ws1.cell(row=r, column=c).border = BORDER
    r += 1

r += 1
ws1.cell(row=r, column=1, value="■ 全体所見").font = SUBTITLE_FONT
r += 1

overall = [
    ("仮払消費税 期末残高", "¥54,400,455", "3月発生：借方¥468,037 / 貸方¥13,454"),
    ("仮受消費税 期末残高", "¥61,486,238", "3月発生：貸方¥881のみ"),
    ("仮受 − 仮払（差額）", "+¥7,085,783", "納税側"),
    ("3月単月の取引件数", "雑収入3件 / 賃借料6件", "決算月としては少なめ"),
]
ws1.cell(row=r, column=1, value="項目").font = HEADER_FONT
ws1.cell(row=r, column=2, value="金額").font = HEADER_FONT
ws1.cell(row=r, column=3, value="備考").font = HEADER_FONT
for c in range(1, 7):
    ws1.cell(row=r, column=c).fill = HEADER_FILL
    ws1.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=r, column=c).border = BORDER
ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
r += 1
for k, v, note in overall:
    ws1.cell(row=r, column=1, value=k).font = NORMAL
    ws1.cell(row=r, column=2, value=v).font = BOLD
    ws1.cell(row=r, column=3, value=note).font = NORMAL
    ws1.merge_cells(start_row=r, start_column=3, end_row=r, end_column=6)
    for c in range(1, 7):
        ws1.cell(row=r, column=c).border = BORDER
    r += 1

r += 1
ws1.cell(row=r, column=1, value="■ チェック結果サマリー").font = SUBTITLE_FONT
r += 1

result_header = ["No", "重要度", "科目", "件名", "金額", "判定"]
for i, h in enumerate(result_header, 1):
    ws1.cell(row=r, column=i, value=h)
style_header(ws1, r, range(1, 7))
r += 1

results = [
    (1, "中", "雑収入", "足立区利子補給：非課税→不課税(対象外)に修正要", "¥56,986", "要修正"),
    (2, "低", "雑収入", "貸付利息：科目を受取利息へ振替推奨（税区分はOK）", "¥27,665", "改善"),
    (3, "低", "賃借料", "tax_code 189(経過措置80%控除)の運用妥当性確認", "計¥486,700", "確認のみ"),
    (4, "—", "雑収入", "事務手数料(課税売上10%)", "¥8,708", "✅妥当"),
    (5, "—", "賃借料", "通常分(課対仕入10%, tax_code 136)", "¥203,500", "✅妥当"),
    (6, "—", "保険積立金", "非課税仕入(tax_code 2)", "¥136,319", "✅妥当"),
]
for row_data in results:
    for i, v in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=i, value=v)
        cell.font = NORMAL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    if row_data[1] == "中":
        ws1.cell(row=r, column=2).fill = ORANGE_FILL
        ws1.cell(row=r, column=2).font = RED_FONT
    elif row_data[1] == "低":
        ws1.cell(row=r, column=2).fill = GRAY_FILL
    elif row_data[5] == "✅妥当":
        ws1.cell(row=r, column=6).fill = GREEN_FILL
    r += 1

r += 1
ws1.cell(row=r, column=1, value="■ 最優先対応事項").font = SUBTITLE_FONT
r += 1
ws1.cell(row=r, column=1, value=(
    "雑収入「足立区利子補給 ¥56,986」を 非課税→不課税（対象外）へ修正してください。\n"
    "本則課税・個別対応方式のため、課税売上割合を通じて控除税額に波及します。"
)).font = Font(name=FONT, size=10, bold=True, color="C00000")
ws1.cell(row=r, column=1).fill = RED_FILL
ws1.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="center")
ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
ws1.row_dimensions[r].height = 45

widths = [22, 28, 24, 24, 24, 24]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ============= Sheet2: 要確認事項詳細 =============
ws2 = wb.create_sheet("要確認事項")
ws2["A1"] = "要確認事項 詳細"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:G1")

headers2 = ["No", "重要度", "仕訳ID", "計上日", "科目", "現状の税区分", "あるべき税区分 / コメント"]
r = 3
for i, h in enumerate(headers2, 1):
    ws2.cell(row=r, column=i, value=h)
style_header(ws2, r, range(1, 8))
r += 1

issues = [
    (1, "中", "3468018062", "2026-03-18", "雑収入 ¥56,986\n(東日本6 足立区利子補給)",
     "tax_code 2 (非課税売上)",
     "不課税(対象外)が正。地方公共団体からの利子補給金は課税対象外。\n"
     "個別対応方式のため、課税売上割合の分母（非課税売上）に算入されると控除対象仕入税額に影響する可能性あり。"),
    (2, "低", "3466212010", "2026-03-26", "雑収入 ¥27,665\n(貸付利息)",
     "tax_code 2 (非課税売上)",
     "税区分は適切。ただし勘定科目は『受取利息』が望ましい。"),
    (3, "低", "3467619976 等", "2026-03-25 他", "賃借料 計¥486,700\n(tax_code 189)",
     "課対仕入10%（控除80%・経過措置）",
     "個人大家・免税事業者からの賃借料に対する経過措置適用とみられ運用は適切。\n"
     "貸主のインボイス登録状況の最新化（登録番号取得）を確認推奨。"),
]
for row in issues:
    for i, v in enumerate(row, 1):
        cell = ws2.cell(row=r, column=i, value=v)
        cell.font = NORMAL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    if row[1] == "中":
        ws2.cell(row=r, column=2).fill = ORANGE_FILL
        ws2.cell(row=r, column=2).font = RED_FONT
    else:
        ws2.cell(row=r, column=2).fill = GRAY_FILL
    ws2.row_dimensions[r].height = 60
    r += 1

widths2 = [5, 10, 16, 12, 28, 24, 50]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ============= Sheet3: 3月仕訳明細 =============
ws3 = wb.create_sheet("3月仕訳明細")
ws3["A1"] = "2026年3月 主要科目 仕訳明細"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:H1")

headers3 = ["仕訳ID", "計上日", "取引タイプ", "勘定科目", "借/貸", "金額", "消費税(vat)", "税区分(tax_code)・摘要"]
r = 3
for i, h in enumerate(headers3, 1):
    ws3.cell(row=r, column=i, value=h)
style_header(ws3, r, range(1, 9))
r += 1

# 雑収入セクション
ws3.cell(row=r, column=1, value="【雑収入】3月合計 ¥93,359").font = SUBTITLE_FONT
ws3.cell(row=r, column=1).fill = GRAY_FILL
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1

zatsu = [
    ("3466355979", "2026-03-31", "expense", "雑収入", "貸方", 8708, 791, "tax_code 129 (課税売上10%) 3月分事務手数料 ✅"),
    ("3466212010", "2026-03-26", "income",  "雑収入", "貸方", 27665, 0, "tax_code 2 (非課税売上) 貸付利息 → 受取利息へ振替推奨"),
    ("3468018062", "2026-03-18", "income",  "雑収入", "貸方", 56986, 0, "tax_code 2 (非課税売上) 利子補給 → 不課税(対象外)へ修正要 ⚠️"),
]
for row in zatsu:
    for i, v in enumerate(row, 1):
        cell = ws3.cell(row=r, column=i, value=v)
        cell.font = NORMAL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    if "修正要" in row[7]:
        ws3.cell(row=r, column=8).fill = RED_FILL
    elif "振替推奨" in row[7]:
        ws3.cell(row=r, column=8).fill = ORANGE_FILL
    else:
        ws3.cell(row=r, column=8).fill = GREEN_FILL
    ws3.cell(row=r, column=6).number_format = '#,##0'
    ws3.cell(row=r, column=7).number_format = '#,##0'
    r += 1

# 賃借料セクション
ws3.cell(row=r, column=1, value="【賃借料】3月合計 ¥690,200").font = SUBTITLE_FONT
ws3.cell(row=r, column=1).fill = GRAY_FILL
ws3.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
r += 1

chinshaku = [
    ("3467981109", "2026-03-27", "expense", "賃借料", "借方", 33000, 3000, "tax_code 136 (課対仕入10%) ✅"),
    ("3467968999", "2026-03-27", "expense", "賃借料", "借方", 170500, 15500, "tax_code 136 (課対仕入10%) ✅"),
    ("3467619976", "2026-03-25", "expense", "賃借料", "借方", 13000, 945, "tax_code 189 (経過措置80%控除) 要確認"),
    ("3467619577", "2026-03-25", "expense", "賃借料", "借方", 43700, 3178, "tax_code 189 (経過措置80%控除) 要確認"),
    ("3436785529", "2026-03-25", "expense", "賃借料", "借方", 150000, 10909, "tax_code 189 (経過措置80%控除) 本社賃借料未払分 2025/10〜12"),
    ("3436775373", "2026-03-25", "expense", "賃借料", "借方", 250000, 18181, "tax_code 189 (経過措置80%控除) 本社賃借料"),
    ("3436775373", "2026-03-25", "expense", "賃借料", "借方", 30000, 2181, "tax_code 189 (経過措置80%控除)"),
]
for row in chinshaku:
    for i, v in enumerate(row, 1):
        cell = ws3.cell(row=r, column=i, value=v)
        cell.font = NORMAL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    if "要確認" in row[7] or "経過措置" in row[7]:
        ws3.cell(row=r, column=8).fill = ORANGE_FILL
    else:
        ws3.cell(row=r, column=8).fill = GREEN_FILL
    ws3.cell(row=r, column=6).number_format = '#,##0'
    ws3.cell(row=r, column=7).number_format = '#,##0'
    r += 1

widths3 = [14, 12, 11, 14, 8, 12, 12, 50]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ============= Sheet4: 決算月の追加確認 =============
ws4 = wb.create_sheet("決算月チェック推奨")
ws4["A1"] = "決算月としての追加確認推奨事項"
ws4["A1"].font = TITLE_FONT
ws4.merge_cells("A1:C1")

r = 3
headers4 = ["No", "確認項目", "内容・観点"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=r, column=i, value=h)
style_header(ws4, r, range(1, 4))
r += 1

decchecks = [
    (1, "課税売上割合の算定", "利子補給を不課税に修正後の数値で再計算する。"),
    (2, "個別対応方式の用途区分", "販管費・売上原価で『課税売上対応／非課税売上対応／共通対応』が適切に設定されているか。"),
    (3, "インボイス経過措置(80%控除)", "適用累計と対象取引先リストの整合性を確認。貸主・支払先のインボイス登録番号最新化。"),
    (4, "期末一括計上仕訳", "前払費用の取崩、未払費用計上、決算修正の税区分が継続性を保っているか。"),
    (5, "仮払/仮受消費税の精算", "差額¥7,085,783を未払消費税等へ振替。中間納付額（仮払消費税内）の精算確認。"),
]
for row in decchecks:
    for i, v in enumerate(row, 1):
        cell = ws4.cell(row=r, column=i, value=v)
        cell.font = NORMAL
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws4.row_dimensions[r].height = 30
    r += 1

ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 28
ws4.column_dimensions['C'].width = 70

out_path = r"C:\Users\yuya_\claude\office\office-claude\data\reports\1362187_株式会社コムネットシステム\消費税チェック_2026年3月.xlsx"
wb.save(out_path)
print(f"OK: {out_path}")
